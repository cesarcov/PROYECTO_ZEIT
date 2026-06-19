"""
Utilidad compartida para exportaciones Excel con estilos CeShark.
Usado por todos los módulos que generen archivos .xlsx descargables.
"""
import io
from datetime import datetime
from fastapi.responses import StreamingResponse

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None


# ── Design tokens ──────────────────────────────────────────────────────────────
PRIMARY_HEX = "0B2E33"
ACCENT_HEX  = "B8E3E9"
LIGHT_HEX   = "EEF7F8"
CAP_HEX     = "1a4a52"
ALT_HEX     = "F0F9FA"

# ── Estilos reutilizables ──────────────────────────────────────────────────────

def _fill(hex_color: str):
    return PatternFill("solid", fgColor=hex_color)

def _border():
    thin = Side(style="thin", color="C5D8DB")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _font_header():
    return Font(bold=True, color="FFFFFF", size=9)

def _font_title():
    return Font(bold=True, color="FFFFFF", size=12)

def _font_normal():
    return Font(size=9, color="1F2937")

def _font_bold_dark():
    return Font(bold=True, size=9, color=PRIMARY_HEX)

def _align(h="left", wrap=False):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)


# ── Función principal: aplica cabecera corporativa ────────────────────────────

def write_title_row(ws, title: str, n_cols: int, row: int = 1) -> None:
    """Fila 1: banda oscura con título del reporte."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    c = ws.cell(row=row, column=1, value=title)
    c.fill      = _fill(PRIMARY_HEX)
    c.font      = _font_title()
    c.alignment = _align("center")
    ws.row_dimensions[row].height = 26


def write_header_row(ws, headers: list[str], row: int = 2) -> None:
    """Fila de encabezados con fondo verde oscuro y texto blanco."""
    for col, header in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=header)
        c.fill      = _fill(CAP_HEX)
        c.font      = _font_header()
        c.alignment = _align("center")
        c.border    = _border()


def write_data_row(ws, row: int, values: list, alternate: bool = False) -> None:
    """Fila de datos con alternado de color para legibilidad."""
    fill = _fill(ALT_HEX) if alternate else _fill("FFFFFF")
    for col, val in enumerate(values, start=1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill      = fill
        c.font      = _font_normal()
        c.alignment = _align("left", wrap=False)
        c.border    = _border()


def write_total_row(ws, row: int, values: list, n_cols: int) -> None:
    """Fila de totales con fondo accent."""
    for col, val in enumerate(values, start=1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill      = _fill(ACCENT_HEX)
        c.font      = _font_bold_dark()
        c.alignment = _align("right" if col > 1 else "left")
        c.border    = _border()


def set_column_widths(ws, widths: list[int]) -> None:
    """Aplica anchos de columna en orden."""
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def fmt_date(val) -> str:
    """Formatea fecha/datetime a DD/MM/YYYY para celdas de texto."""
    if val is None:
        return ""
    if isinstance(val, str):
        try:
            val = datetime.fromisoformat(val.replace("Z", "+00:00"))
        except Exception:
            return val
    return val.strftime("%d/%m/%Y")


def fmt_num(val, decimals: int = 2) -> str:
    """Número con separador de miles y decimales fijos."""
    if val is None:
        return ""
    try:
        return f"{float(val):,.{decimals}f}"
    except Exception:
        return str(val)


# ── StreamingResponse helper ───────────────────────────────────────────────────

def excel_response(wb, filename: str) -> StreamingResponse:
    """Serializa el Workbook y devuelve StreamingResponse para descarga."""
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
