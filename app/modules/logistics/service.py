import app.core.security
import json
from decimal import Decimal
from typing import Optional, Dict
from uuid import UUID
from fastapi import HTTPException
from psycopg2 import sql

from openpyxl import Workbook
from openpyxl.styles import Font
from pathlib import Path
from datetime import datetime
from app.core.database import db_connection

from app.modules.logistics.schemas import (
    StockMovementCreate,
    StockLocationCreate,
    ToolAssignCreate,
    MaterialCreate,
    StockReceptionCreate
)
from app.modules.logistics.utils import read_excel

DISPATCH_FILES_DIR = Path("app/storage/dispatches")


# ============================================================
# 📦 CALCULAR STOCK TOTAL (KARDEX)
# ============================================================
def get_current_stock(
    material_id: str,
    warehouse_id: str,
    project_id: str | None = None,
) -> float:

    query = """
        SELECT movement_type, quantity, from_warehouse, to_warehouse
        FROM stock_movements
        WHERE material_id = %s
    """

    params = [material_id]

    if project_id:
        query += " AND project_id = %s"
        params.append(project_id)

    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(query, tuple(params))
            rows = cur.fetchall()

    stock = 0.0

    for m in rows:
        movement_type = m[0]
        quantity = float(m[1])
        from_wh = m[2]
        to_wh = m[3]

        if movement_type in ("IN", "RETURN") and to_wh == warehouse_id:
            stock += quantity

        elif movement_type == "OUT" and from_wh == warehouse_id:
            stock -= quantity

        elif movement_type == "TRANSFER":
            # 👉 Transferencia interna NO cambia el total del almacén
            # Solo moverá stock entre ubicaciones
            continue

    return stock

# ============================================================
# 📍 STOCK POR UBICACIÓN
# ============================================================
def get_location_stock(material_id, warehouse_id, rack, level, box, position):
    query = """
        SELECT quantity
        FROM stock_locations
        WHERE material_id = %s
          AND warehouse_id = %s
          AND rack = %s
          AND level = %s
          AND box = %s
          AND position = %s
        LIMIT 1;
    """

    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(query, (
                str(material_id),   # 🔑 CONVERTIR UUID → STRING
                str(warehouse_id),
                rack,
                level,
                box,
                position
            ))
            row = cur.fetchone()
            return float(row[0]) if row else 0.0



# ============================================================
# 📜 HISTORIAL
# ============================================================
def get_stock_movements_history(
    date_from: str,
    date_to: str,
    material_id: Optional[str] = None,
    project_id: Optional[str] = None,
    warehouse_id: Optional[str] = None,
    movement_type: Optional[str] = None,
    created_by: Optional[str] = None,
):
    query = """
        SELECT
            sm.id,
            sm.movement_type,
            sm.quantity,
            m.name          AS material_name,
            m.code          AS material_code,
            wf.name         AS from_warehouse_name,
            wt.name         AS to_warehouse_name,
            sm.reference,
            sm.notes,
            sm.created_by,
            sm.created_at,
            sm.material_id,
            sm.from_warehouse,
            sm.to_warehouse
        FROM stock_movements sm
        JOIN materials m ON m.id = sm.material_id
        LEFT JOIN warehouses wf ON wf.id = sm.from_warehouse
        LEFT JOIN warehouses wt ON wt.id = sm.to_warehouse
        WHERE sm.created_at BETWEEN %s AND %s
    """
    params = [f"{date_from} 00:00:00", f"{date_to} 23:59:59"]

    if material_id:
        query += " AND sm.material_id = %s"
        params.append(material_id)
    if project_id:
        query += " AND sm.project_id = %s"
        params.append(project_id)
    if warehouse_id:
        query += " AND (sm.from_warehouse = %s OR sm.to_warehouse = %s)"
        params.extend([warehouse_id, warehouse_id])
    if movement_type:
        query += " AND sm.movement_type = %s"
        params.append(movement_type)
    if created_by:
        query += " AND sm.created_by = %s"
        params.append(created_by)

    query += " ORDER BY sm.created_at DESC LIMIT 500"

    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(query, params)
            rows = cur.fetchall()

    return [
        {
            "id": str(r[0]),
            "movement_type": r[1],
            "quantity": float(r[2]),
            "material_name": r[3],
            "material_code": r[4],
            "from_warehouse": r[5],
            "to_warehouse": r[6],
            "reference": r[7],
            "notes": r[8],
            "created_by": r[9],
            "created_at": r[10],
            "material_id": str(r[11]),
            "from_warehouse_id": str(r[12]) if r[12] else None,
            "to_warehouse_id": str(r[13]) if r[13] else None,
        }
        for r in rows
    ]


# ============================================================
# 📊 RESUMEN
# ============================================================
def get_current_stock_summary(material_id: Optional[str] = None) -> Dict:
    query = "SELECT material_id, movement_type, quantity, from_warehouse, to_warehouse FROM stock_movements"
    params = []

    if material_id:
        query += " WHERE material_id = %s"
        params.append(material_id)

    stock: Dict[str, Dict[str, float]] = {}

    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(query, params)
            rows = cur.fetchall()

    for mat, m_type, qty, from_wh, to_wh in rows:
        stock.setdefault(mat, {})

        def add(wh, amount):
            if not wh:
                return
            stock[mat][wh] = stock[mat].get(wh, 0) + amount

        if m_type in ("IN", "RETURN"):
            add(to_wh, qty)
        elif m_type == "OUT":
            add(from_wh, -qty)
        elif m_type == "TRANSFER":
            add(from_wh, -qty)
            add(to_wh, qty)

    return stock


# ============================================================
# 🚨 STOCK NEGATIVO
# ============================================================
def get_negative_stock():
    stock = get_current_stock_summary()
    alerts = []

    for material, warehouses in stock.items():
        for wh, qty in warehouses.items():
            if qty < 0:
                alerts.append({
                    "material_id": material,
                    "warehouse_id": wh,
                    "quantity": qty,
                    "alert": "NEGATIVE_STOCK"
                })

    return alerts


# ============================================================
# 📍 UBICACIONES
# ============================================================
def upsert_stock_location(payload: StockLocationCreate):
    data = payload.model_dump()

    # 🔁 Convertir UUIDs a string (psycopg2 no acepta UUID directamente)
    for k, v in data.items():
        if isinstance(v, UUID):
            data[k] = str(v)

    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO stock_locations (
                    material_id, warehouse_id, rack, level, box, position, quantity
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT (material_id, warehouse_id, rack, level, box, position)
                DO UPDATE SET
                    quantity = EXCLUDED.quantity,
                    updated_at = NOW()
                RETURNING id, material_id, warehouse_id, rack, level, box, position,
                          quantity, created_at, updated_at;
                """,
                (
                    data["material_id"],
                    data["warehouse_id"],
                    data["rack"],
                    data["level"],
                    data["box"],
                    data["position"],
                    data["quantity"],
                )
            )

            row = cur.fetchone()
            conn.commit()

    return {
        "id": row[0],
        "material_id": row[1],
        "warehouse_id": row[2],
        "rack": row[3],
        "level": row[4],
        "box": row[5],
        "position": row[6],
        "quantity": float(row[7]),
        "created_at": row[8],
        "updated_at": row[9],
    }



def get_material_locations(material_id: str, warehouse_id: Optional[str] = None):
    query = """
        SELECT material_id, warehouse_id, rack, level, box, position, quantity
        FROM stock_locations
        WHERE material_id = %s
    """
    params = [material_id]

    if warehouse_id:
        query += " AND warehouse_id = %s"
        params.append(warehouse_id)

    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(query, params)
            rows = cur.fetchall()

    return rows


def consume_stock_location(
    material_id, warehouse_id, rack, level, box, position, quantity
):
    # 🔁 Convertir UUIDs a string para psycopg2
    if isinstance(material_id, UUID):
        material_id = str(material_id)
    if isinstance(warehouse_id, UUID):
        warehouse_id = str(warehouse_id)

    with db_connection() as conn:
        with conn.cursor() as cur:
            # 🔎 Buscar ubicación exacta
            cur.execute("""
                SELECT id, quantity FROM stock_locations
                WHERE material_id=%s AND warehouse_id=%s
                  AND rack=%s AND level=%s AND box=%s AND position=%s
            """, (material_id, warehouse_id, rack, level, box, position))

            row = cur.fetchone()

            if not row:
                raise ValueError("No existe stock en la ubicación indicada")

            location_id, current_qty = row

            if current_qty < quantity:
                raise ValueError("Stock insuficiente en la ubicación indicada")

            # ➖ Descontar cantidad
            cur.execute("""
                UPDATE stock_locations
                SET quantity = quantity - %s,
                    updated_at = NOW()
                WHERE id = %s
            """, (quantity, location_id))

        conn.commit()


def add_stock_location(
    material_id, warehouse_id, rack, level, box, position, quantity
):
    # 🔁 Convertir UUIDs a string para psycopg2
    if isinstance(material_id, UUID):
        material_id = str(material_id)
    if isinstance(warehouse_id, UUID):
        warehouse_id = str(warehouse_id)

    with db_connection() as conn:
        with conn.cursor() as cur:
            # 🔎 Verificar si ya existe esa ubicación
            cur.execute("""
                SELECT id, quantity FROM stock_locations
                WHERE material_id=%s AND warehouse_id=%s
                  AND rack=%s AND level=%s AND box=%s AND position=%s
            """, (material_id, warehouse_id, rack, level, box, position))

            row = cur.fetchone()

            if row:
                location_id, current_qty = row

                # ➕ Sumar cantidad existente
                cur.execute("""
                    UPDATE stock_locations
                    SET quantity = quantity + %s,
                        updated_at = NOW()
                    WHERE id = %s
                """, (quantity, location_id))
            else:
                # 🆕 Crear nueva ubicación
                cur.execute("""
                    INSERT INTO stock_locations (
                        material_id, warehouse_id, rack, level, box, position, quantity
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                """, (material_id, warehouse_id, rack, level, box, position, quantity))

        conn.commit()


def get_stock_by_project(project_id: str | None = None):
    base_query = """
        SELECT 
            project_id,
            material_id,
            SUM(
                CASE
                    WHEN movement_type = 'OUT' THEN quantity
                    WHEN movement_type = 'RETURN' THEN -quantity
                    ELSE 0
                END
            ) AS used_quantity
        FROM stock_movements
        WHERE project_id IS NOT NULL
    """

    params = []

    if project_id:
        base_query += " AND project_id = %s"
        params.append(project_id)

    base_query += " GROUP BY project_id, material_id;"

    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(base_query, params)
            rows = cur.fetchall()

    result = []
    for r in rows:
        result.append({
            "project_id": str(r[0]),
            "material_id": str(r[1]),
            "used_quantity": float(r[2])
        })

    return result

def get_low_stock_alerts():
    query = """
        SELECT 
            m.id AS material_id,
            COALESCE(SUM(
                CASE 
                    WHEN sm.movement_type IN ('IN','RETURN') THEN sm.quantity
                    WHEN sm.movement_type = 'OUT' THEN -sm.quantity
                    ELSE 0
                END
            ), 0) AS current_stock,
            m.min_stock
        FROM materials m
        LEFT JOIN stock_movements sm ON sm.material_id = m.id
        GROUP BY m.id, m.min_stock
        HAVING COALESCE(SUM(
                CASE 
                    WHEN sm.movement_type IN ('IN','RETURN') THEN sm.quantity
                    WHEN sm.movement_type = 'OUT' THEN -sm.quantity
                    ELSE 0
                END
            ), 0) <= m.min_stock;
    """

    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(query)
            rows = cur.fetchall()

    alerts = []
    for material_id, current_stock, min_stock in rows:
        alerts.append({
            "material_id": str(material_id),
            "current_stock": float(current_stock),
            "min_stock": float(min_stock),
            "alert": "LOW_STOCK"
        })

    return alerts

def get_most_used_materials(limit: int = 10):
    query = """
        SELECT 
            material_id,
            SUM(quantity) AS total_used
        FROM stock_movements
        WHERE movement_type = 'OUT'
        GROUP BY material_id
        ORDER BY total_used DESC
        LIMIT %s;
    """

    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(query, (limit,))
            rows = cur.fetchall()

    return [
        {
            "material_id": str(material_id),
            "total_used": float(total_used)
        }
        for material_id, total_used in rows
    ]

def assign_tool_service(payload: ToolAssignCreate):
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                INSERT INTO tool_assignments (
                    material_id, project_id, assigned_to, expected_return, condition_out
                )
                VALUES (%s,%s,%s,%s,%s)
                RETURNING id, material_id, project_id, assigned_to,
                          assigned_at, expected_return, status, condition_out;
            """, (
                str(payload.material_id),
                str(payload.project_id),
                payload.assigned_to,
                payload.expected_return,
                payload.condition_out,
            ))

            row = cur.fetchone()
            conn.commit()

    return {
        "id": str(row[0]),
        "material_id": str(row[1]),
        "project_id": str(row[2]),
        "assigned_to": row[3],
        "assigned_at": row[4],
        "expected_return": row[5],
        "status": row[6],
        "condition_out": row[7],
    }

# ============================================================
# 🔁 DEVOLVER HERRAMIENTA
# ============================================================
def return_tool_service(assignment_id: str):
    with db_connection() as conn:
        with conn.cursor() as cur:
            # 1️⃣ Buscar asignación activa
            cur.execute("""
                SELECT id, material_id, project_id, assigned_to, assigned_at,
                       expected_return, status
                FROM tool_assignments
                WHERE id = %s AND status = 'IN_USE';
            """, (assignment_id,))

            row = cur.fetchone()

            if not row:
                raise HTTPException(
                    status_code=404,
                    detail="Asignación no encontrada o ya devuelta"
                )

            # 2️⃣ Marcar como devuelta
            cur.execute("""
                UPDATE tool_assignments
                SET status = 'RETURNED',
                    returned_at = NOW()
                WHERE id = %s;
            """, (assignment_id,))

            conn.commit()

    return {
        "assignment_id": assignment_id,
        "status": "RETURNED",
        "message": "Herramienta devuelta correctamente"
    }

def get_assigned_tools_service():
    query = """
        SELECT id, material_id, project_id, assigned_to,
               assigned_at, expected_return, status
        FROM tool_assignments
        WHERE status = 'IN_USE';
    """

    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(query)
            rows = cur.fetchall()

    return [
        {
            "id": str(r[0]),
            "material_id": str(r[1]),
            "project_id": str(r[2]),
            "assigned_to": r[3],
            "assigned_at": r[4],
            "expected_return": r[5],
            "status": r[6],
        }
        for r in rows
    ]

def register_tool_maintenance(material_id, maintenance_type, last_maintenance, next_due, notes=None):

    # 🔑 Convertir UUID a string si viene como objeto UUID
    if isinstance(material_id, UUID):
        material_id = str(material_id)

    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                INSERT INTO tool_maintenance (
                    material_id, maintenance_type, last_maintenance, next_due, notes
                ) VALUES (%s,%s,%s,%s,%s)
                RETURNING id;
            """, (material_id, maintenance_type, last_maintenance, next_due, notes))
            maintenance_id = cur.fetchone()[0]
            conn.commit()

    return {"id": str(maintenance_id)}

def get_due_maintenance():
    query = """
        SELECT id, material_id, maintenance_type, next_due
        FROM tool_maintenance
        WHERE next_due <= CURRENT_DATE;
    """

    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(query)
            rows = cur.fetchall()

    return [
        {
            "id": str(r[0]),
            "material_id": str(r[1]),
            "maintenance_type": r[2],
            "next_due": r[3],
            "alert": "MAINTENANCE_DUE"
        }
        for r in rows
    ]

def get_tool_maintenance_alerts(days_ahead: int = 7):
    query = """
        SELECT 
            tm.id,
            tm.material_id,
            tm.maintenance_type,
            tm.last_maintenance,
            tm.next_due,
            tm.notes
        FROM tool_maintenance tm
        WHERE tm.next_due <= CURRENT_DATE + INTERVAL '%s days'
        ORDER BY tm.next_due ASC;
    """

    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(query, (days_ahead,))
            rows = cur.fetchall()

    alerts = []
    for r in rows:
        alerts.append({
            "id": str(r[0]),
            "material_id": str(r[1]),
            "maintenance_type": r[2],
            "last_maintenance": r[3],
            "next_due": r[4],
            "notes": r[5],
            "alert": "MAINTENANCE_DUE"
        })

    return alerts

# ============================================================
# 📦 CREAR MATERIAL
# ============================================================
def create_material_service(payload: MaterialCreate):
    data = payload.model_dump()

    if data.get("aliases") and len(data["aliases"]) > 3:
        raise HTTPException(400, "Máximo 3 nombres alternativos por material")

    # Convertir strings vacíos a None en campos opcionales
    for f in ("purchase_date", "warranty_expires", "brand", "model",
              "serial_number", "supplier_name", "supplier_contact"):
        if data.get(f) == "":
            data[f] = None

    with db_connection() as conn:
        with conn.cursor() as cur:

            cur.execute("SELECT id FROM materials WHERE code = %s;", (data["code"],))
            if cur.fetchone():
                raise HTTPException(400, "Ya existe un material con ese código")

            cur.execute("""
                INSERT INTO materials (
                    name, code, min_stock, category,
                    brand, model, serial_number,
                    supplier_name, supplier_contact, unit_cost,
                    useful_life_years, purchase_date, warranty_expires
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id, name, code, min_stock, category, created_at,
                          brand, model, serial_number,
                          supplier_name, supplier_contact, unit_cost,
                          useful_life_years, purchase_date, warranty_expires;
            """, (
                data["name"], data["code"], data["min_stock"], data["category"],
                data.get("brand"), data.get("model"), data.get("serial_number"),
                data.get("supplier_name"), data.get("supplier_contact"), data.get("unit_cost"),
                data.get("useful_life_years"),
                data.get("purchase_date") or None,
                data.get("warranty_expires") or None,
            ))

            row = cur.fetchone()
            material_id = row[0]

            aliases = data.get("aliases", [])
            for alias in aliases:
                cur.execute(
                    "INSERT INTO material_aliases (material_id, alias_name) VALUES (%s, %s);",
                    (material_id, alias)
                )

            conn.commit()

    return {
        "id": str(row[0]), "name": row[1], "code": row[2],
        "min_stock": float(row[3]), "category": row[4], "aliases": aliases, "created_at": row[5],
        "brand": row[6], "model": row[7], "serial_number": row[8],
        "supplier_name": row[9], "supplier_contact": row[10],
        "unit_cost": float(row[11]) if row[11] is not None else None,
        "useful_life_years": row[12],
        "purchase_date": row[13].isoformat() if row[13] else None,
        "warranty_expires": row[14].isoformat() if row[14] else None,
    }


# ============================================================
# 📋 LISTAR MATERIALES
# ============================================================
def get_materials_service(estado: str = None):
    with db_connection() as conn:
        with conn.cursor() as cur:
            where = "WHERE m.estado = %s" if estado else "WHERE m.estado != 'INACTIVO'"
            params = (estado,) if estado else ()
            cur.execute(f"""
                SELECT
                    m.id, m.name, m.code, m.min_stock, m.category, m.created_at,
                    COALESCE(array_agg(a.alias_name) FILTER (WHERE a.alias_name IS NOT NULL), '{{}}') AS aliases,
                    m.brand, m.model, m.serial_number,
                    m.supplier_name, m.supplier_contact, m.unit_cost,
                    m.useful_life_years, m.purchase_date, m.warranty_expires,
                    m.estado, m.precio_referencia, m.proveedor_referencia
                FROM materials m
                LEFT JOIN material_aliases a ON a.material_id = m.id
                {where}
                GROUP BY m.id
                ORDER BY m.estado, m.name;
            """, params)
            rows = cur.fetchall()

    return [
        {
            "id": str(r[0]), "name": r[1], "code": r[2],
            "min_stock": float(r[3]), "category": r[4], "created_at": r[5],
            "aliases": list(r[6]),
            "brand": r[7], "model": r[8], "serial_number": r[9],
            "supplier_name": r[10], "supplier_contact": r[11],
            "unit_cost": float(r[12]) if r[12] is not None else None,
            "useful_life_years": r[13],
            "purchase_date": r[14].isoformat() if r[14] else None,
            "warranty_expires": r[15].isoformat() if r[15] else None,
            "estado": r[16] if r[16] else "ACTIVO",
            "precio_referencia": float(r[17]) if r[17] is not None else None,
            "proveedor_referencia": r[18],
        }
        for r in rows
    ]

# ============================================================
# ✏️ EDITAR MATERIAL
# ============================================================
def update_material_service(material_id: str, payload):
    data = payload.model_dump(exclude_unset=True)

    # Convertir strings vacíos a None en campos opcionales
    for f in ("purchase_date", "warranty_expires", "brand", "model",
              "serial_number", "supplier_name", "supplier_contact"):
        if data.get(f) == "":
            data[f] = None

    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT id FROM materials WHERE id = %s;", (material_id,))
            if not cur.fetchone():
                raise HTTPException(404, "Material no encontrado")

            if "code" in data and data["code"]:
                cur.execute("SELECT id FROM materials WHERE code = %s AND id != %s;", (data["code"], material_id))
                if cur.fetchone():
                    raise HTTPException(400, "Ya existe un material con ese código")

            field_map = {k: v for k, v in data.items() if k != "aliases" and v is not None}
            if field_map:
                set_clause = sql.SQL(", ").join(
                    sql.SQL("{} = %s").format(sql.Identifier(k)) for k in field_map
                )
                values = list(field_map.values()) + [material_id]
                cur.execute(
                    sql.SQL("UPDATE materials SET {} WHERE id = %s;").format(set_clause),
                    values,
                )

            if "aliases" in data and data["aliases"] is not None:
                cur.execute("DELETE FROM material_aliases WHERE material_id = %s;", (material_id,))
                for alias in data["aliases"]:
                    if alias.strip():
                        cur.execute("INSERT INTO material_aliases (material_id, alias_name) VALUES (%s, %s);", (material_id, alias.strip()))

            cur.execute(
                "SELECT id, name, code, min_stock, category FROM materials WHERE id = %s;",
                (material_id,)
            )
            row = cur.fetchone()
            conn.commit()

    return {"id": str(row[0]), "name": row[1], "code": row[2], "min_stock": float(row[3]), "category": row[4]}


# ============================================================
# 🗑️ ELIMINAR MATERIAL
# ============================================================
def delete_material_service(material_id: str):
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT id, name FROM materials WHERE id = %s;", (material_id,))
            mat = cur.fetchone()
            if not mat:
                raise HTTPException(404, "Material no encontrado")

            cur.execute("SELECT COUNT(*) FROM stock_movements WHERE material_id = %s;", (material_id,))
            count = cur.fetchone()[0]
            if count > 0:
                raise HTTPException(400, f"No se puede eliminar: el material tiene {count} movimiento(s) registrados")

            cur.execute("DELETE FROM material_aliases WHERE material_id = %s;", (material_id,))
            cur.execute("DELETE FROM materials WHERE id = %s;", (material_id,))
            conn.commit()

    return {"message": f"Material '{mat[1]}' eliminado correctamente"}


# ============================================================
# 📦 A) Importar Materiales
# ============================================================

def import_materials_from_excel(file):
    df = read_excel(file)

    # Normalize Spanish column names to internal field names
    COL_MAP = {
        "Código": "code", "Nombre": "name", "Categoría": "category",
        "Stock mínimo": "min_stock", "Marca": "brand", "Modelo": "model",
        "Proveedor": "supplier_name", "Contacto proveedor": "supplier_contact",
        "Costo unitario": "unit_cost", "N° serie": "serial_number",
        "Vida útil (años)": "useful_life_years", "Fecha compra": "purchase_date",
        "Venc. garantía": "warranty_expires", "Aliases (sep. por coma)": "aliases",
    }
    df = df.rename(columns=COL_MAP)

    if "code" not in df.columns or "name" not in df.columns:
        raise ValueError("Faltan columnas obligatorias: 'Código' y 'Nombre'")

    inserted = 0
    updated = 0
    errors = []

    with db_connection() as conn:
        with conn.cursor() as cur:
            for idx, row in df.iterrows():
                try:
                    code = str(row["code"]).strip() if row.get("code") else None
                    name = str(row["name"]).strip() if row.get("name") else None
                    if not code or not name:
                        raise ValueError("Código y Nombre son obligatorios")

                    category   = str(row["category"]).strip()   if row.get("category")   else None
                    min_stock  = float(row["min_stock"])         if row.get("min_stock") is not None else 0.0
                    brand      = str(row["brand"]).strip()       if row.get("brand")      else None
                    model      = str(row["model"]).strip()       if row.get("model")      else None
                    sup_name   = str(row["supplier_name"]).strip()    if row.get("supplier_name")    else None
                    sup_con    = str(row["supplier_contact"]).strip()  if row.get("supplier_contact") else None
                    unit_cost  = float(row["unit_cost"])         if row.get("unit_cost") is not None else None
                    serial     = str(row["serial_number"]).strip()    if row.get("serial_number")    else None
                    life_yrs   = int(row["useful_life_years"])   if row.get("useful_life_years") is not None else None
                    pur_date   = str(row["purchase_date"]).strip()    if row.get("purchase_date")    else None
                    warranty   = str(row["warranty_expires"]).strip()  if row.get("warranty_expires")  else None
                    aliases_raw = str(row["aliases"]).strip()    if row.get("aliases")    else ""
                    aliases = [a.strip() for a in aliases_raw.split(",") if a.strip()] if aliases_raw else []

                    cur.execute("SELECT id FROM materials WHERE code = %s", (code,))
                    existing = cur.fetchone()

                    if existing:
                        mat_id = existing[0]
                        cur.execute("""
                            UPDATE materials
                            SET name=%s, category=%s, min_stock=%s,
                                brand=%s, model=%s, serial_number=%s,
                                supplier_name=%s, supplier_contact=%s, unit_cost=%s,
                                useful_life_years=%s, purchase_date=%s, warranty_expires=%s
                            WHERE id=%s
                        """, (name, category, min_stock, brand, model, serial,
                              sup_name, sup_con, unit_cost, life_yrs, pur_date, warranty, mat_id))
                        cur.execute("DELETE FROM material_aliases WHERE material_id = %s", (mat_id,))
                        updated += 1
                    else:
                        cur.execute("""
                            INSERT INTO materials (
                                name, code, category, min_stock,
                                brand, model, serial_number,
                                supplier_name, supplier_contact, unit_cost,
                                useful_life_years, purchase_date, warranty_expires
                            ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                            RETURNING id
                        """, (name, code, category, min_stock, brand, model, serial,
                              sup_name, sup_con, unit_cost, life_yrs, pur_date, warranty))
                        mat_id = cur.fetchone()[0]
                        inserted += 1

                    for alias in aliases:
                        cur.execute(
                            "INSERT INTO material_aliases (material_id, alias_name) VALUES (%s, %s) ON CONFLICT DO NOTHING",
                            (mat_id, alias)
                        )
                except Exception as e:
                    errors.append(f"Fila {int(idx)+2}: {e}")
                    continue

        conn.commit()

    return {
        "imported": inserted + updated,
        "inserted": inserted,
        "updated": updated,
        "skipped": 0,
        "errors": errors,
    }

# ============================================================
#  📥 B) Importar Stock IN (Ingreso)
# ============================================================

def import_stock_in_from_excel(file):
    df = read_excel(file)

    # NOTA: usamos warehouse_code (código humano), no UUID
    required_cols = ["material_code", "warehouse_id", "quantity"]
    for col in required_cols:
        if col not in df.columns:
            raise ValueError(f"Falta la columna: {col}")

    inserted = 0
    failed = 0
    errors = []

    with db_connection() as conn:
        with conn.cursor() as cur:
            for idx, row in df.iterrows():
                try:
                    # 1️⃣ Validar material por código
                    if not row.get("material_code"):
                        raise ValueError("Fila sin material_code")

                    cur.execute(
                        "SELECT id FROM materials WHERE code=%s;",
                        (row["material_code"],)
                    )
                    mat = cur.fetchone()
                    if not mat:
                        raise ValueError(f"Material no existe: {row['material_code']}")

                    material_id = mat[0]

                    # 2️⃣ Validar almacén por CÓDIGO (WH-CENTRAL, etc.)
                    if not row.get("warehouse_id"):
                        raise ValueError("Fila sin warehouse_id (código)")

                    cur.execute(
                        "SELECT id FROM warehouses WHERE code=%s;",
                        (row["warehouse_id"],)
                    )
                    wh = cur.fetchone()
                    if not wh:
                        raise ValueError(f"Almacén no existe: {row['warehouse_id']}")

                    warehouse_id = wh[0]

                    # 3️⃣ Insertar movimiento IN usando SOLO UUIDs
                    cur.execute("""
                        INSERT INTO stock_movements (
                            material_id, movement_type, quantity,
                            from_warehouse, to_warehouse,
                            reference, notes, created_by
                        )
                        VALUES (%s,'IN',%s,NULL,%s,%s,%s,'excel');
                    """, (
                        str(material_id),
                        float(row["quantity"]),
                        str(warehouse_id),
                        row.get("reference"),
                        row.get("notes"),
                    ))

                    inserted += 1

                except Exception as e:
                    failed += 1
                    errors.append({
                        "row": int(idx) + 2,  # +2 por encabezado y base 0
                        "material_code": row.get("material_code"),
                        "warehouse_id": row.get("warehouse_id"),
                        "error": str(e)
                    })
                    # Continúa con la siguiente fila
                    continue

        conn.commit()

    return {
        "status": "OK",
        "inserted": inserted,
        "failed": failed,
        "total": inserted + failed,
        "errors": errors
    }


# ============================================================
#  📤 C) Importar Stock OUT por Proyecto
# ============================================================

def import_stock_out_from_excel(file):
    df = read_excel(file)

    # NOTA: warehouse_id y project_id vienen como CÓDIGOS, no UUID
    required_cols = ["material_code", "warehouse_id", "project_id", "quantity"]
    for col in required_cols:
        if col not in df.columns:
            raise ValueError(f"Falta la columna: {col}")

    inserted = 0
    failed = 0
    errors = []

    with db_connection() as conn:
        with conn.cursor() as cur:
            for idx, row in df.iterrows():
                try:
                    # 1️⃣ Validar material por código
                    if not row.get("material_code"):
                        raise ValueError("Fila sin material_code")

                    cur.execute(
                        "SELECT id FROM materials WHERE code=%s;",
                        (row["material_code"],)
                    )
                    mat = cur.fetchone()
                    if not mat:
                        raise ValueError(f"Material no existe: {row['material_code']}")

                    material_id = mat[0]

                    # 2️⃣ Validar almacén por CÓDIGO
                    if not row.get("warehouse_id"):
                        raise ValueError("Fila sin warehouse_id (código)")

                    cur.execute(
                        "SELECT id FROM warehouses WHERE code=%s;",
                        (row["warehouse_id"],)
                    )
                    wh = cur.fetchone()
                    if not wh:
                        raise ValueError(f"Almacén no existe: {row['warehouse_id']}")

                    warehouse_id = wh[0]

                    # 3️⃣ Validar proyecto por CÓDIGO
                    if not row.get("project_id"):
                        raise ValueError("Fila sin project_id (código)")

                    cur.execute(
                        "SELECT id FROM projects WHERE code=%s;",
                        (row["project_id"],)
                    )
                    pr = cur.fetchone()
                    if not pr:
                        raise ValueError(f"Proyecto no existe: {row['project_id']}")

                    project_id = pr[0]

                    # 4️⃣ Insertar movimiento OUT usando SOLO UUIDs
                    cur.execute("""
                        INSERT INTO stock_movements (
                            material_id, movement_type, quantity,
                            from_warehouse, to_warehouse, project_id,
                            reference, notes, created_by
                        )
                        VALUES (%s,'OUT',%s,%s,NULL,%s,%s,%s,'excel');
                    """, (
                        str(material_id),
                        float(row["quantity"]),
                        str(warehouse_id),
                        str(project_id),
                        row.get("reference"),
                        row.get("notes"),
                    ))

                    inserted += 1

                except Exception as e:
                    failed += 1
                    errors.append({
                        "row": int(idx) + 2,  # +2 por encabezado y base 0
                        "material_code": row.get("material_code"),
                        "warehouse_id": row.get("warehouse_id"),
                        "project_id": row.get("project_id"),
                        "error": str(e)
                    })
                    continue

        conn.commit()

    return {
        "status": "OK",
        "inserted": inserted,
        "failed": failed,
        "total": inserted + failed,
        "errors": errors
    }


# ============================================================
# 🧹 RESET DE DATOS (SOLO PARA TESTING)
# ============================================================

def reset_logistics_data_service():
    with db_connection() as conn:
        with conn.cursor() as cur:
            # Movimientos y ubicaciones
            cur.execute("TRUNCATE TABLE stock_movements RESTART IDENTITY CASCADE;")
            cur.execute("TRUNCATE TABLE stock_locations RESTART IDENTITY CASCADE;")

            # Herramientas
            cur.execute("TRUNCATE TABLE tool_assignments RESTART IDENTITY CASCADE;")
            cur.execute("TRUNCATE TABLE tool_maintenance RESTART IDENTITY CASCADE;")

            # Materiales
            cur.execute("TRUNCATE TABLE material_aliases RESTART IDENTITY CASCADE;")
            cur.execute("TRUNCATE TABLE materials RESTART IDENTITY CASCADE;")

            # Proyectos y almacenes
            cur.execute("TRUNCATE TABLE projects RESTART IDENTITY CASCADE;")
            cur.execute("TRUNCATE TABLE warehouses RESTART IDENTITY CASCADE;")

            conn.commit()

    return {
        "status": "OK",
        "message": "Todos los datos de logística fueron eliminados (tablas conservadas)."
    }

# ============================================================
# 🏬 WAREHOUSES
# ============================================================

def create_warehouse_service(payload):
    data = payload.model_dump()

    with db_connection() as conn:
        with conn.cursor() as cur:
            # Evitar códigos duplicados
            cur.execute("SELECT id FROM warehouses WHERE code = %s;", (data["code"],))
            if cur.fetchone():
                raise HTTPException(400, "Ya existe un almacén con ese código")

            cur.execute("""
                INSERT INTO warehouses (code, name, location)
                VALUES (%s, %s, %s)
                RETURNING id, code, name, location;
            """, (
                data["code"],
                data["name"],
                data.get("location"),
            ))

            row = cur.fetchone()
            conn.commit()

    return {
        "id": str(row[0]),
        "code": row[1],
        "name": row[2],
        "location": row[3],
    }


def get_warehouses_service():
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT id, code, name, location FROM warehouses ORDER BY code;")
            rows = cur.fetchall()

    return [
        {
            "id": str(r[0]),
            "code": r[1],
            "name": r[2],
            "location": r[3],
        }
        for r in rows
    ]


def get_warehouse_inventory_service(warehouse_id: str):
    """
    Full inventory detail for a single warehouse:
    stock, location coding, tool condition, last movement.
    """
    with db_connection() as conn:
        with conn.cursor() as cur:
            # Warehouse info
            cur.execute("SELECT id, code, name, location FROM warehouses WHERE id = %s", (warehouse_id,))
            wh = cur.fetchone()
            if not wh:
                raise ValueError("Almacén no encontrado")

            # Full inventory
            cur.execute("""
                SELECT
                    m.id, m.code, m.name, m.description, m.unit, m.category,
                    m.brand, m.model, m.min_stock, m.unit_cost,
                    COALESCE(v.stock_available, 0)  AS stock_available,
                    sl.rack, sl.level, sl.box, sl.position AS location_ref,
                    lm.movement_type  AS last_mv_type,
                    lm.quantity       AS last_mv_qty,
                    lm.notes          AS last_mv_notes,
                    lm.created_at     AS last_mv_at,
                    ta.condition_in, ta.condition_out, ta.return_notes,
                    ta.status         AS assignment_status,
                    ta.returned_at
                FROM materials m
                JOIN vw_stock_availability v
                    ON v.material_id = m.id AND v.warehouse_id = %s
                LEFT JOIN stock_locations sl
                    ON sl.material_id = m.id AND sl.warehouse_id = %s
                LEFT JOIN LATERAL (
                    SELECT movement_type, quantity, notes, created_at
                    FROM stock_movements
                    WHERE material_id = m.id
                      AND (to_warehouse = %s OR from_warehouse = %s)
                    ORDER BY created_at DESC
                    LIMIT 1
                ) lm ON TRUE
                LEFT JOIN LATERAL (
                    SELECT condition_in, condition_out, return_notes, status, returned_at
                    FROM tool_assignments
                    WHERE material_id = m.id
                    ORDER BY COALESCE(returned_at, assigned_at) DESC
                    LIMIT 1
                ) ta ON TRUE
                ORDER BY m.category, m.name
            """, (warehouse_id, warehouse_id, warehouse_id, warehouse_id))
            rows = cur.fetchall()

    items = []
    for r in rows:
        items.append({
            "material_id":       str(r[0]),
            "code":              r[1],
            "name":              r[2],
            "description":       r[3],
            "unit":              r[4],
            "category":          r[5],
            "brand":             r[6],
            "model":             r[7],
            "min_stock":         float(r[8]) if r[8] is not None else None,
            "unit_cost":         float(r[9]) if r[9] is not None else None,
            "stock_available":   float(r[10]),
            "rack":              r[11],
            "level":             r[12],
            "box":               r[13],
            "location_ref":      r[14],
            "last_mv_type":      r[15],
            "last_mv_qty":       float(r[16]) if r[16] is not None else None,
            "last_mv_notes":     r[17],
            "last_mv_at":        r[18].isoformat() if r[18] else None,
            "condition_in":      r[19],
            "condition_out":     r[20],
            "return_notes":      r[21],
            "assignment_status": r[22],
            "returned_at":       r[23].isoformat() if r[23] else None,
        })

    return {
        "warehouse": {
            "id":       str(wh[0]),
            "code":     wh[1],
            "name":     wh[2],
            "location": wh[3],
        },
        "items": items,
        "summary": {
            "total_skus":   len(items),
            "total_units":  sum(i["stock_available"] for i in items),
            "zero_stock":   sum(1 for i in items if i["stock_available"] <= 0),
            "low_stock":    sum(1 for i in items if i["stock_available"] > 0 and i["min_stock"] and i["stock_available"] <= i["min_stock"]),
            "with_damage":  sum(1 for i in items if i["condition_in"] and i["condition_in"].upper() in ("DAÑADO", "DAÑADA", "REQUIERE MANTENIMIENTO", "MALO", "MALA")),
        },
    }


def get_stock_availability_service():
    """
    Flat list: material + warehouse + stock_available + location data.
    """
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT
                    m.id              AS material_id,
                    m.name            AS material_name,
                    m.code            AS material_code,
                    m.min_stock,
                    w.id              AS warehouse_id,
                    w.name            AS warehouse_name,
                    w.code            AS warehouse_code,
                    v.stock_available,
                    sl.rack,
                    sl.level,
                    sl.box,
                    sl.position
                FROM vw_stock_availability v
                JOIN materials  m ON m.id = v.material_id
                JOIN warehouses w ON w.id = v.warehouse_id
                LEFT JOIN stock_locations sl
                    ON sl.material_id = v.material_id
                   AND sl.warehouse_id = v.warehouse_id
                ORDER BY m.name, w.code;
            """)
            rows = cur.fetchall()

    return [
        {
            "material_id":    str(r[0]),
            "material_name":  r[1],
            "material_code":  r[2],
            "min_stock":      float(r[3]) if r[3] is not None else None,
            "warehouse_id":   str(r[4]),
            "warehouse_name": r[5],
            "warehouse_code": r[6],
            "stock_available": float(r[7]),
            "rack":           r[8],
            "level":          r[9],
            "box":            r[10],
            "location_ref":   r[11],
        }
        for r in rows
    ]


def update_stock_location_meta_service(
    material_id: str,
    warehouse_id: str,
    rack: str,
    level: str,
    box: str,
    position: str,
):
    """Update only the physical location coding for a material in a warehouse."""
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                UPDATE stock_locations
                SET rack = %s, level = %s, box = %s, position = %s, updated_at = NOW()
                WHERE material_id = %s AND warehouse_id = %s
                RETURNING id
            """, (rack, level, box, position, material_id, warehouse_id))
            row = cur.fetchone()
            if not row:
                # Create location entry if it doesn't exist yet
                cur.execute("""
                    INSERT INTO stock_locations (id, material_id, warehouse_id, rack, level, box, position, quantity, created_at, updated_at)
                    VALUES (gen_random_uuid(), %s, %s, %s, %s, %s, %s, 0, NOW(), NOW())
                    RETURNING id
                """, (material_id, warehouse_id, rack, level, box, position))
            conn.commit()
    return {"ok": True}


# ============================================================
# 🏗 PROJECTS
# ============================================================

def create_project_service(payload):
    data = payload.model_dump()

    with db_connection() as conn:
        with conn.cursor() as cur:
            # Evitar códigos duplicados
            cur.execute("SELECT id FROM projects WHERE code = %s;", (data["code"],))
            if cur.fetchone():
                raise HTTPException(400, "Ya existe un proyecto con ese código")

            cur.execute("""
                INSERT INTO projects (code, name)
                VALUES (%s, %s)
                RETURNING id, code, name;
            """, (
                data["code"],
                data["name"],
            ))

            row = cur.fetchone()
            conn.commit()

    return {
        "id": str(row[0]),
        "code": row[1],
        "name": row[2],
    }


def get_projects_service():
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT id, code, name FROM projects ORDER BY code;")
            rows = cur.fetchall()

    return [
        {
            "id": str(r[0]),
            "code": r[1],
            "name": r[2],
        }
        for r in rows
    ]


def get_project_summary_service(project_id: str):
    with db_connection() as conn:
        with conn.cursor() as cur:
            # Basic info
            cur.execute("SELECT id, code, name, created_at FROM projects WHERE id = %s", (project_id,))
            proj = cur.fetchone()
            if not proj:
                raise ValueError("Proyecto no encontrado")

            # Tools assigned
            cur.execute("""
                SELECT
                    ta.id, m.name AS tool_name, m.code AS tool_code,
                    ta.assigned_to, ta.assigned_at, ta.expected_return,
                    ta.status, ta.condition_out, ta.returned_at, ta.condition_in
                FROM tool_assignments ta
                JOIN materials m ON m.id = ta.material_id
                WHERE ta.project_id = %s
                ORDER BY ta.assigned_at DESC
            """, (project_id,))
            tools = [
                {
                    "id": str(r[0]), "tool_name": r[1], "tool_code": r[2],
                    "assigned_to": r[3], "assigned_at": r[4].isoformat() if r[4] else None,
                    "expected_return": r[5].isoformat() if r[5] else None,
                    "status": r[6], "condition_out": r[7],
                    "returned_at": r[8].isoformat() if r[8] else None,
                    "condition_in": r[9],
                }
                for r in cur.fetchall()
            ]

            # Material requests
            cur.execute("""
                SELECT
                    mr.id, mr.status, mr.priority, mr.quantity, mr.needed_by, mr.created_at,
                    u.username AS requested_by,
                    COALESCE(m.name, mr.reason) AS material_name,
                    COALESCE(m.code, '') AS material_code,
                    mr.notes
                FROM material_requests mr
                JOIN users u ON u.id = mr.requested_by
                LEFT JOIN materials m ON m.id = mr.related_material_id
                WHERE mr.project_id = %s
                ORDER BY mr.created_at DESC
            """, (project_id,))
            requests = [
                {
                    "id": str(r[0]), "status": r[1], "priority": r[2],
                    "quantity": float(r[3]) if r[3] else None,
                    "needed_by": r[4].isoformat() if r[4] else None,
                    "created_at": r[5].isoformat() if r[5] else None,
                    "requested_by": r[6], "material_name": r[7],
                    "material_code": r[8], "notes": r[9],
                }
                for r in cur.fetchall()
            ]

            # Dispatches
            cur.execute("""
                SELECT
                    sd.id, sd.status, sd.recipient_name, sd.dispatched_at, sd.created_at,
                    u.username AS dispatched_by,
                    COUNT(sdi.id) AS items_count
                FROM stock_dispatches sd
                LEFT JOIN users u ON u.id = sd.dispatched_by
                LEFT JOIN stock_dispatch_items sdi ON sdi.dispatch_id = sd.id
                WHERE sd.project_id = %s
                GROUP BY sd.id, u.username
                ORDER BY sd.created_at DESC
            """, (project_id,))
            dispatches = [
                {
                    "id": str(r[0]), "status": r[1], "recipient_name": r[2],
                    "dispatched_at": r[3].isoformat() if r[3] else None,
                    "created_at": r[4].isoformat() if r[4] else None,
                    "dispatched_by": r[5], "items_count": r[6],
                }
                for r in cur.fetchall()
            ]

            # Operations plans
            cur.execute("""
                SELECT
                    pp.id, pp.title, pp.status, pp.created_at,
                    u.username AS engineer_name,
                    COUNT(ppi.id) AS items_count
                FROM project_plans pp
                JOIN users u ON u.id = pp.engineer_id
                LEFT JOIN project_plan_items ppi ON ppi.plan_id = pp.id
                WHERE pp.project_id = %s
                GROUP BY pp.id, u.username
                ORDER BY pp.created_at DESC
            """, (project_id,))
            plans = [
                {
                    "id": str(r[0]), "title": r[1], "status": r[2],
                    "created_at": r[3].isoformat() if r[3] else None,
                    "engineer_name": r[4], "items_count": r[5],
                }
                for r in cur.fetchall()
            ]

    tools_in_field = sum(1 for t in tools if t["status"] == "ASSIGNED")
    return {
        "id": str(proj[0]),
        "code": proj[1],
        "name": proj[2],
        "created_at": proj[3].isoformat() if proj[3] else None,
        "kpis": {
            "tools_total": len(tools),
            "tools_in_field": tools_in_field,
            "requests_total": len(requests),
            "requests_pending": sum(1 for r in requests if r["status"] == "PENDING"),
            "dispatches_total": len(dispatches),
            "dispatches_delivered": sum(1 for d in dispatches if d["status"] == "DELIVERED"),
            "plans_total": len(plans),
        },
        "tools": tools,
        "requests": requests,
        "dispatches": dispatches,
        "plans": plans,
    }


def update_warehouse_service(warehouse_id: str, payload):
    data = {k: v for k, v in payload.model_dump().items() if v is not None}
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT id FROM warehouses WHERE id = %s;", (warehouse_id,))
            if not cur.fetchone():
                raise HTTPException(404, "Almacén no encontrado")
            if "code" in data:
                cur.execute("SELECT id FROM warehouses WHERE code = %s AND id != %s;", (data["code"], warehouse_id))
                if cur.fetchone():
                    raise HTTPException(400, "Ya existe un almacén con ese código")
            set_clause = sql.SQL(", ").join(
                sql.SQL("{} = %s").format(sql.Identifier(k)) for k in data
            )
            values = list(data.values()) + [warehouse_id]
            cur.execute(
                sql.SQL("UPDATE warehouses SET {} WHERE id = %s RETURNING id, code, name, location;").format(set_clause),
                values,
            )
            row = cur.fetchone()
            conn.commit()
    return {"id": str(row[0]), "code": row[1], "name": row[2], "location": row[3]}


def delete_warehouse_service(warehouse_id: str):
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT COUNT(*) FROM stock_movements WHERE from_warehouse = %s OR to_warehouse = %s;",
                (warehouse_id, warehouse_id)
            )
            count = cur.fetchone()[0]
            if count > 0:
                raise HTTPException(400, f"No se puede eliminar: el almacén tiene {count} movimientos registrados")
            cur.execute("DELETE FROM warehouses WHERE id = %s RETURNING id;", (warehouse_id,))
            if not cur.fetchone():
                raise HTTPException(404, "Almacén no encontrado")
            conn.commit()
    return {"ok": True}


def update_project_service(project_id: str, payload):
    data = {k: v for k, v in payload.model_dump().items() if v is not None}
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT id FROM projects WHERE id = %s;", (project_id,))
            if not cur.fetchone():
                raise HTTPException(404, "Proyecto no encontrado")
            if "code" in data:
                cur.execute("SELECT id FROM projects WHERE code = %s AND id != %s;", (data["code"], project_id))
                if cur.fetchone():
                    raise HTTPException(400, "Ya existe un proyecto con ese código")
            set_clause = sql.SQL(", ").join(
                sql.SQL("{} = %s").format(sql.Identifier(k)) for k in data
            )
            values = list(data.values()) + [project_id]
            cur.execute(
                sql.SQL("UPDATE projects SET {} WHERE id = %s RETURNING id, code, name;").format(set_clause),
                values,
            )
            row = cur.fetchone()
            conn.commit()
    return {"id": str(row[0]), "code": row[1], "name": row[2]}


def delete_project_service(project_id: str):
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM projects WHERE id = %s RETURNING id;", (project_id,))
            if not cur.fetchone():
                raise HTTPException(404, "Proyecto no encontrado")
            conn.commit()
    return {"ok": True}


# ============================================================
# 📥 IMPORTAR WAREHOUSES (EXCEL)
# ============================================================

def import_warehouses_from_excel(file):
    df = read_excel(file)

    required_cols = ["code", "name", "location"]
    for col in required_cols:
        if col not in df.columns:
            raise ValueError(f"Falta la columna: {col}")

    inserted = 0
    updated = 0
    failed = 0
    errors = []

    with db_connection() as conn:
        with conn.cursor() as cur:
            for idx, row in df.iterrows():
                try:
                    if not row.get("code"):
                        raise ValueError("Fila sin code")

                    cur.execute("SELECT id FROM warehouses WHERE code = %s;", (row["code"],))
                    existing = cur.fetchone()

                    if existing:
                        cur.execute("""
                            UPDATE warehouses
                            SET name = %s,
                                location = %s
                            WHERE code = %s;
                        """, (
                            row["name"],
                            row.get("location"),
                            row["code"],
                        ))
                        updated += 1
                    else:
                        cur.execute("""
                            INSERT INTO warehouses (code, name, location)
                            VALUES (%s, %s, %s);
                        """, (
                            row["code"],
                            row["name"],
                            row.get("location"),
                        ))
                        inserted += 1

                except Exception as e:
                    failed += 1
                    errors.append({
                        "row": int(idx) + 2,
                        "code": row.get("code"),
                        "error": str(e)
                    })
                    continue

        conn.commit()

    return {
        "status": "OK",
        "inserted": inserted,
        "updated": updated,
        "failed": failed,
        "total": inserted + updated + failed,
        "errors": errors
    }


# ============================================================
# 📥 IMPORTAR PROJECTS (EXCEL)
# ============================================================

def import_projects_from_excel(file):
    df = read_excel(file)

    required_cols = ["code", "name"]
    for col in required_cols:
        if col not in df.columns:
            raise ValueError(f"Falta la columna: {col}")

    inserted = 0
    updated = 0
    failed = 0
    errors = []

    with db_connection() as conn:
        with conn.cursor() as cur:
            for idx, row in df.iterrows():
                try:
                    if not row.get("code"):
                        raise ValueError("Fila sin code")

                    cur.execute("SELECT id FROM projects WHERE code = %s;", (row["code"],))
                    existing = cur.fetchone()

                    if existing:
                        cur.execute("""
                            UPDATE projects
                            SET name = %s
                            WHERE code = %s;
                        """, (
                            row["name"],
                            row["code"],
                        ))
                        updated += 1
                    else:
                        cur.execute("""
                            INSERT INTO projects (code, name)
                            VALUES (%s, %s);
                        """, (
                            row["code"],
                            row["name"],
                        ))
                        inserted += 1

                except Exception as e:
                    failed += 1
                    errors.append({
                        "row": int(idx) + 2,
                        "code": row.get("code"),
                        "error": str(e)
                    })
                    continue

        conn.commit()

    return {
        "status": "OK",
        "inserted": inserted,
        "updated": updated,
        "failed": failed,
        "total": inserted + updated + failed,
        "errors": errors
    }

def consume_stock_location_tx(
    cur,
    material_id,
    warehouse_id,
    rack,
    level,
    box,
    position,
    quantity
):
    cur.execute("""
        SELECT id, quantity
        FROM stock_locations
        WHERE material_id=%s AND warehouse_id=%s
          AND rack=%s AND level=%s AND box=%s AND position=%s
        FOR UPDATE
    """, (material_id, warehouse_id, rack, level, box, position))

    row = cur.fetchone()
    if not row:
        raise ValueError("No existe stock en la ubicación")

    location_id, current_qty = row

    if current_qty < quantity:
        raise ValueError("Stock insuficiente")

    cur.execute("""
        UPDATE stock_locations
        SET quantity = quantity - %s,
            updated_at = NOW()
        WHERE id = %s
    """, (quantity, location_id))

def add_stock_location_tx(
    cur,
    material_id,
    warehouse_id,
    rack,
    level,
    box,
    position,
    quantity
):
    cur.execute("""
        SELECT id FROM stock_locations
        WHERE material_id=%s AND warehouse_id=%s
          AND rack=%s AND level=%s AND box=%s AND position=%s
        FOR UPDATE
    """, (material_id, warehouse_id, rack, level, box, position))

    row = cur.fetchone()

    if row:
        cur.execute("""
            UPDATE stock_locations
            SET quantity = quantity + %s,
                updated_at = NOW()
            WHERE id = %s
        """, (quantity, row[0]))
    else:
        cur.execute("""
            INSERT INTO stock_locations (
                material_id, warehouse_id, rack, level, box, position, quantity
            )
            VALUES (%s,%s,%s,%s,%s,%s,%s)
        """, (material_id, warehouse_id, rack, level, box, position, quantity))

def save_audit_log_tx(
    cur,
    *,
    user: dict,
    action: str,
    module: str,
    entity: str | None,
    entity_id,
    new_data: dict | None
):
    cur.execute("""
        INSERT INTO audit_logs (
            user_id,
            username,
            roles,
            action,
            endpoint,
            module,
            payload,
            ip_address,
            status,
            method,
            entity,
            entity_id,
            new_data,
            user_agent
        )
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
    """, (
        user.get("id"),
        user.get("username"),
        user.get("roles") or [],           # ✅ ARRAY
        action,
        user.get("endpoint", "UNKNOWN"),   # ✅ NOT NULL
        module,
        None,                               # payload (opcional)
        user.get("ip_address"),
        "SUCCESS",
        user.get("method", "UNKNOWN"),
        entity,
        entity_id,
        json.dumps(new_data) if new_data else None,
        user.get("user_agent"),
    ))


#=========================================================================#
# AQUI
#=========================================================================

def create_stock_movement_service(payload: StockMovementCreate, current_user: dict):
    with db_connection() as conn:
        try:
            with conn.cursor() as cur:

                # =========================================================
                # 🔹 VALIDACIONES
                # =========================================================
                if payload.movement_type in ("OUT", "ADJUST", "TRANSFER"):
                    if not all([payload.rack, payload.level, payload.box]):
                        raise HTTPException(
                            status_code=400,
                            detail="Debe indicar ubicación (rack, level, box)"
                        )

                # =========================================================
                # 🔹 DETERMINAR FROM / TO WAREHOUSE
                # =========================================================
                warehouse_id = str(payload.warehouse_id)

                if payload.movement_type == "IN":
                    from_wh = None
                    to_wh = warehouse_id
                elif payload.movement_type == "OUT":
                    from_wh = warehouse_id
                    to_wh = None
                else:  # TRANSFER / ADJUST
                    from_wh = warehouse_id
                    to_wh = warehouse_id

                # =========================================================
                # 🔹 INSERT STOCK MOVEMENT
                # =========================================================
                cur.execute("""
                    INSERT INTO stock_movements (
                        material_id,
                        movement_type,
                        quantity,
                        from_warehouse,
                        to_warehouse,
                        project_id,
                        reference,
                        notes,
                        created_by
                    )
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    RETURNING id;
                """, (
                    str(payload.material_id),
                    payload.movement_type,
                    payload.quantity,
                    from_wh,
                    to_wh,
                    str(payload.project_id) if payload.project_id else None,
                    payload.reference,
                    payload.notes,
                    current_user.get("username"),
                ))

                movement_id = cur.fetchone()[0]

                # =========================================================
                # 🔹 IMPACTO EN UBICACIONES (MISMA TRANSACCIÓN)
                # =========================================================
                if payload.movement_type == "TRANSFER":
                    consume_stock_location_tx(
                        cur,
                        str(payload.material_id),
                        warehouse_id,
                        payload.rack,
                        payload.level,
                        payload.box,
                        payload.position,
                        payload.quantity,
                    )

                    add_stock_location_tx(
                        cur,
                        str(payload.material_id),
                        warehouse_id,
                        payload.to_rack,
                        payload.to_level,
                        payload.to_box,
                        payload.to_position,
                        payload.quantity,
                    )

                elif payload.movement_type == "IN":
                    add_stock_location_tx(
                        cur,
                        str(payload.material_id),
                        warehouse_id,
                        payload.rack,
                        payload.level,
                        payload.box,
                        payload.position,
                        payload.quantity,
                    )

                # =========================================================
                # 🔹 AUDITORÍA (DOMAIN LEVEL – MISMA TRANSACCIÓN)
                # =========================================================
                audit_data = payload.model_dump()
                for k, v in audit_data.items():
                    if isinstance(v, UUID):
                        audit_data[k] = str(v)

                save_audit_log_tx(
                    cur,
                    user=current_user,
                    action="CREATE",
                    module="logistics",
                    entity="stock_movement",
                    entity_id=str(movement_id),
                    new_data=audit_data,
                )

                # =========================================================
                # 🔹 RESPUESTA COMPLETA (PARA response_model)
                # =========================================================
                cur.execute("""
                    SELECT
                        id,
                        material_id,
                        movement_type,
                        quantity,
                        from_warehouse,
                        to_warehouse,
                        project_id,
                        reference,
                        notes,
                        created_by,
                        created_at
                    FROM stock_movements
                    WHERE id = %s;
                """, (movement_id,))

                row = cur.fetchone()

            conn.commit()

            return {
                "id": row[0],
                "material_id": row[1],
                "movement_type": row[2],
                "quantity": float(row[3]),
                "from_warehouse": row[4],
                "to_warehouse": row[5],
                "project_id": row[6],
                "reference": row[7],
                "notes": row[8],
                "created_by": row[9],
                "created_at": row[10],
            }

        except Exception:
            conn.rollback()
            raise

# ============================================================
# 🚚 DISPATCHES — FLUJO COMPLETO (PENDING → READY → IN_TRANSIT → DELIVERED)
# ============================================================

def list_dispatches_service():
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT
                    sd.id,
                    sd.status,
                    sd.notes,
                    sd.created_at,
                    sd.dispatched_at,
                    sd.delivered_at,
                    sd.receipt_notes,
                    sd.recipient_name,
                    m.name          AS material_name,
                    m.code          AS material_code,
                    w.name          AS warehouse_name,
                    sr.quantity,
                    u_rec.username  AS recipient_username,
                    sd.reservation_id,
                    sd.recipient_user_id,
                    sd.request_id
                FROM stock_dispatches sd
                LEFT JOIN stock_reservations sr ON sr.id = sd.reservation_id
                LEFT JOIN materials  m  ON m.id  = sr.material_id
                LEFT JOIN warehouses w  ON w.id  = sr.warehouse_id
                LEFT JOIN users u_rec   ON u_rec.id = sd.recipient_user_id
                ORDER BY sd.created_at DESC
                LIMIT 200
            """)
            rows = cur.fetchall()

    return [
        {
            "id": str(r[0]),
            "status": r[1],
            "notes": r[2],
            "created_at": r[3],
            "dispatched_at": r[4],
            "delivered_at": r[5],
            "receipt_notes": r[6],
            "recipient_name": r[7],
            "material_name": r[8],
            "material_code": r[9],
            "warehouse_name": r[10],
            "quantity": float(r[11]) if r[11] is not None else None,
            "recipient_username": r[12],
            "reservation_id": str(r[13]) if r[13] else None,
            "recipient_user_id": str(r[14]) if r[14] else None,
            "request_id": str(r[15]) if r[15] else None,
        }
        for r in rows
    ]


def create_dispatch_service(payload, user: dict):
    with db_connection() as conn:
        try:
            with conn.cursor() as cur:
                cur.execute("""
                    SELECT status, material_id, warehouse_id, quantity, project_id, material_request_id
                    FROM stock_reservations
                    WHERE id = %s
                    FOR UPDATE
                """, (str(payload.reservation_id),))

                res = cur.fetchone()
                if not res:
                    raise HTTPException(404, "Reserva no encontrada")

                res_status, material_id, warehouse_id, quantity, project_id, material_request_id = res

                if res_status != "CONFIRMED":
                    raise HTTPException(400, "Solo se puede despachar desde reservas CONFIRMED")

                cur.execute("""
                    INSERT INTO stock_dispatches (
                        reservation_id,
                        warehouse_id,
                        project_id,
                        dispatched_by,
                        recipient_user_id,
                        recipient_name,
                        request_id,
                        notes,
                        status
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 'PENDING')
                    RETURNING id, created_at
                """, (
                    str(payload.reservation_id),
                    str(warehouse_id),
                    str(project_id) if project_id else None,
                    str(user["id"]),
                    str(payload.recipient_user_id) if payload.recipient_user_id else None,
                    payload.recipient_name,
                    str(material_request_id) if material_request_id else None,
                    payload.notes,
                ))

                dispatch_id, created_at = cur.fetchone()
            conn.commit()

        except Exception:
            conn.rollback()
            raise

    return {"id": str(dispatch_id), "status": "PENDING", "created_at": created_at}


def update_dispatch_status_service(dispatch_id: str, new_status: str, user: dict):
    valid_transitions = {"PENDING": "READY", "READY": "IN_TRANSIT"}

    with db_connection() as conn:
        try:
            with conn.cursor() as cur:
                cur.execute("""
                    SELECT status, reservation_id, warehouse_id, project_id
                    FROM stock_dispatches
                    WHERE id = %s
                    FOR UPDATE
                """, (dispatch_id,))

                row = cur.fetchone()
                if not row:
                    raise HTTPException(404, "Despacho no encontrado")

                current_status, reservation_id, warehouse_id, project_id = row

                if valid_transitions.get(current_status) != new_status:
                    raise HTTPException(400, f"Transición inválida: {current_status} → {new_status}")

                if new_status == "IN_TRANSIT":
                    cur.execute("""
                        SELECT material_id, quantity, status
                        FROM stock_reservations
                        WHERE id = %s
                        FOR UPDATE
                    """, (str(reservation_id),))

                    res = cur.fetchone()
                    if not res:
                        raise HTTPException(404, "Reserva no encontrada")

                    material_id, quantity, res_status = res

                    if res_status != "CONFIRMED":
                        raise HTTPException(400, "La reserva no está CONFIRMED")

                    cur.execute("""
                        SELECT COALESCE(SUM(quantity), 0)
                        FROM stock_locations
                        WHERE material_id = %s AND warehouse_id = %s
                    """, (material_id, warehouse_id))

                    physical_stock = float(cur.fetchone()[0])
                    if physical_stock < float(quantity):
                        raise HTTPException(400, f"Stock físico insuficiente: {physical_stock:.2f} disponible, {float(quantity):.2f} requerido")

                    remaining = float(quantity)
                    cur.execute("""
                        SELECT rack, level, box, position, quantity
                        FROM stock_locations
                        WHERE material_id = %s AND warehouse_id = %s AND quantity > 0
                        ORDER BY quantity DESC
                        FOR UPDATE
                    """, (material_id, warehouse_id))

                    for rack, level, box, position, loc_qty in cur.fetchall():
                        if remaining <= 0:
                            break
                        consume = min(remaining, float(loc_qty))
                        cur.execute("""
                            UPDATE stock_locations
                            SET quantity = quantity - %s, updated_at = NOW()
                            WHERE material_id = %s AND warehouse_id = %s
                              AND rack = %s AND level = %s AND box = %s AND position = %s
                        """, (consume, material_id, warehouse_id, rack, level, box, position))
                        remaining -= consume

                    cur.execute("""
                        INSERT INTO stock_movements (
                            material_id, movement_type, quantity,
                            from_warehouse, to_warehouse, project_id,
                            reference, notes, created_by
                        )
                        VALUES (%s, 'OUT', %s, %s, NULL, %s, %s, %s, %s)
                    """, (
                        material_id, quantity, warehouse_id, project_id,
                        f"DISPATCH-{dispatch_id}",
                        "Salida por despacho",
                        user.get("username"),
                    ))

                    cur.execute("""
                        UPDATE stock_reservations
                        SET status = 'CONSUMED', released_at = NOW()
                        WHERE id = %s
                    """, (str(reservation_id),))

                    cur.execute("""
                        UPDATE stock_dispatches
                        SET status = %s, dispatched_at = NOW()
                        WHERE id = %s
                    """, (new_status, dispatch_id))

                else:
                    cur.execute("""
                        UPDATE stock_dispatches SET status = %s WHERE id = %s
                    """, (new_status, dispatch_id))

            conn.commit()

        except Exception:
            conn.rollback()
            raise

    return {"dispatch_id": dispatch_id, "status": new_status}


def confirm_dispatch_receipt_service(dispatch_id: str, user: dict, receipt_notes: str | None):
    with db_connection() as conn:
        try:
            with conn.cursor() as cur:
                cur.execute("""
                    SELECT status FROM stock_dispatches WHERE id = %s FOR UPDATE
                """, (dispatch_id,))

                row = cur.fetchone()
                if not row:
                    raise HTTPException(404, "Despacho no encontrado")

                if row[0] != "IN_TRANSIT":
                    raise HTTPException(400, f"El despacho no está EN_TRÁNSITO (estado: {row[0]})")

                cur.execute("""
                    UPDATE stock_dispatches
                    SET status = 'DELIVERED',
                        delivered_at = NOW(),
                        receipt_notes = %s,
                        received_by = %s
                    WHERE id = %s
                """, (receipt_notes, str(user["id"]), dispatch_id))

            conn.commit()

        except Exception:
            conn.rollback()
            raise

    return {"dispatch_id": dispatch_id, "status": "DELIVERED"}


def receive_stock_service(payload: StockReceptionCreate, current_user):
    with db_connection() as conn:
        with conn.cursor() as cur:

            # 1️⃣ Obtener movimiento OUT original
            cur.execute("""
                SELECT material_id, quantity, project_id
                FROM stock_movements
                WHERE id = %s
                  AND movement_type = 'OUT'
            """, (str(payload.movement_id),))

            row = cur.fetchone()
            if not row:
                raise HTTPException(
                    404,
                    "Movimiento OUT no encontrado o inválido"
                )

            material_id, quantity, project_id = row

            # 2️⃣ Crear movimiento IN (recepción)
            cur.execute("""
                INSERT INTO stock_movements (
                    material_id,
                    movement_type,
                    quantity,
                    from_warehouse,
                    to_warehouse,
                    project_id,
                    notes,
                    created_by
                )
                VALUES (%s,'IN',%s,NULL,%s,%s,%s,%s)
                RETURNING id
            """, (
                material_id,
                quantity,
                str(payload.warehouse_id),
                project_id,
                payload.notes,
                current_user["username"]
            ))

            in_movement_id = cur.fetchone()[0]

            # 3️⃣ Registrar ubicación física
            add_stock_location_tx(
                cur,
                str(material_id),
                str(payload.warehouse_id),
                payload.rack,
                payload.level,
                payload.box,
                payload.position,
                quantity
            )

        conn.commit()

    return {
        "status": "RECEIVED",
        "movement_in_id": str(in_movement_id)
    }


def generate_dispatch_excel(dispatch_id: str) -> str:
    DISPATCH_FILES_DIR.mkdir(parents=True, exist_ok=True)

    with db_connection() as conn:
        with conn.cursor() as cur:

            # 1️⃣ Cabecera del despacho
            cur.execute("""
                SELECT
                    sd.id,
                    sd.created_at,
                    sd.dispatched_at,
                    sd.notes,
                    w.name AS warehouse_name,
                    p.name AS project_name,
                    sd.dispatched_by,
                    sd.received_by
                FROM stock_dispatches sd
                JOIN warehouses w ON w.id = sd.warehouse_id
                JOIN projects p ON p.id = sd.project_id
                WHERE sd.id = %s
            """, (dispatch_id,))

            header = cur.fetchone()
            if not header:
                raise Exception("Despacho no encontrado")

            (
                _id,
                created_at,
                dispatched_at,
                notes,
                warehouse_name,
                project_name,
                dispatched_by,
                received_by
            ) = header

            # 2️⃣ Ítems agrupados por categoría
            cur.execute("""
                SELECT
                    i.category,
                    i.item_name,
                    i.quantity,
                    i.unit
                FROM stock_dispatch_items i
                WHERE i.dispatch_id = %s
                ORDER BY i.category, i.item_name
            """, (dispatch_id,))

            items = cur.fetchall()

    # 3️⃣ Crear Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Despacho"

    bold = Font(bold=True)

    # Encabezado
    ws.append(["DESPACHO DE MATERIALES"])
    ws["A1"].font = Font(bold=True, size=14)

    ws.append([])
    ws.append(["Despacho ID:", dispatch_id])
    ws.append(["Proyecto / Destino:", project_name])
    ws.append(["Almacén:", warehouse_name])
    ws.append(["Despachado por:", dispatched_by])
    ws.append(["Recibido por:", received_by])
    ws.append(["Fecha despacho:", dispatched_at or created_at])
    ws.append(["Observaciones:", notes or ""])

    ws.append([])

    # 4️⃣ Detalle por categoría
    current_category = None

    for category, name, qty, unit in items:
        if category != current_category:
            ws.append([])
            ws.append([category.upper()])
            ws[f"A{ws.max_row}"].font = bold
            ws.append(["Ítem", "Cantidad", "Unidad"])
            for col in ["A", "B", "C"]:
                ws[f"{col}{ws.max_row}"].font = bold
            current_category = category

        ws.append([name, float(qty), unit])

    # 5️⃣ Guardar archivo
    filename = f"dispatch_{dispatch_id}.xlsx"
    filepath = DISPATCH_FILES_DIR / filename
    wb.save(filepath)

    return str(filepath)



# =========================================================
# ➕ AGREGAR ITEM A DESPACHO
# =========================================================

def add_dispatch_item_service(dispatch_id: str, payload, current_user: dict):
    from fastapi import HTTPException

    with db_connection() as conn:
        try:
            with conn.cursor() as cur:

                # 1️⃣ Validar despacho
                cur.execute("""
                    SELECT status
                    FROM stock_dispatches
                    WHERE id = %s
                    FOR UPDATE
                """, (dispatch_id,))
                row = cur.fetchone()

                if not row:
                    raise HTTPException(404, "Despacho no encontrado")

                if row[0] != "CREATED":
                    raise HTTPException(
                        400,
                        "Solo se pueden agregar ítems a despachos en estado CREATED"
                    )

                # 2️⃣ Obtener datos del material
                cur.execute("""
                    SELECT name, category
                    FROM materials
                    WHERE id = %s
                """, (str(payload.material_id),))

                material = cur.fetchone()

                if not material:
                    raise HTTPException(404, "Material no encontrado")

                material_name, category = material

                # 3️⃣ Insertar ítem
                cur.execute("""
                    INSERT INTO stock_dispatch_items (
                        dispatch_id,
                        material_id,
                        item_name,
                        category,
                        quantity,
                        unit
                    )
                    VALUES (%s,%s,%s,%s,%s,%s)
                """, (
                    dispatch_id,
                    str(payload.material_id),
                    material_name,
                    category,
                    payload.quantity,
                    "UND"
                ))

            conn.commit()

            return {
                "dispatch_id": dispatch_id,
                "material_id": str(payload.material_id),
                "quantity": float(payload.quantity),
                "status": "ITEM_ADDED"
            }

        except Exception:
            conn.rollback()
            raise




def dispatch_stock_service(dispatch_id: str, current_user: dict):
    from fastapi import HTTPException
    from app.core.database import db_connection

    with db_connection() as conn:
        try:
            with conn.cursor() as cur:

                # =========================================================
                # 1️⃣ VALIDAR DESPACHO
                # =========================================================
                cur.execute("""
                    SELECT status, reservation_id, warehouse_id, project_id
                    FROM stock_dispatches
                    WHERE id = %s
                    FOR UPDATE
                """, (dispatch_id,))

                row = cur.fetchone()

                if not row:
                    raise HTTPException(404, "Despacho no encontrado")

                status, reservation_id, warehouse_id, project_id = row

                if status != "CREATED":
                    raise HTTPException(
                        400,
                        "El despacho no está en estado CREATED"
                    )

                # =========================================================
                # 2️⃣ VALIDAR RESERVA CONFIRMED
                # =========================================================
                cur.execute("""
                    SELECT material_id, quantity
                    FROM stock_reservations
                    WHERE id = %s
                      AND status = 'CONFIRMED'
                    FOR UPDATE
                """, (reservation_id,))

                reservation = cur.fetchone()

                if not reservation:
                    raise HTTPException(
                        400,
                        "La reserva no está CONFIRMED"
                    )

                reserved_material_id, reserved_quantity = reservation
                reserved_quantity = float(reserved_quantity)

                # =========================================================
                # 3️⃣ OBTENER ÍTEMS DEL DESPACHO
                # =========================================================
                cur.execute("""
                    SELECT material_id, quantity
                    FROM stock_dispatch_items
                    WHERE dispatch_id = %s
                """, (dispatch_id,))

                items = cur.fetchall()

                if not items:
                    raise HTTPException(
                        400,
                        "El despacho no tiene ítems"
                    )

                # =========================================================
                # 🔒 4️⃣ VALIDACIÓN ANTI-SOBREDESPACHO
                # =========================================================
                total_to_dispatch = 0

                for material_id, quantity in items:

                    if str(material_id) != str(reserved_material_id):
                        raise HTTPException(
                            400,
                            "El material del despacho no coincide con la reserva"
                        )

                    total_to_dispatch += float(quantity)

                # ¿Cuánto ya se despachó antes?
                cur.execute("""
                    SELECT COALESCE(SUM(quantity), 0)
                    FROM stock_movements
                    WHERE reference = %s
                      AND material_id = %s
                """, (
                    f"DISPATCH-{dispatch_id}",
                    reserved_material_id
                ))

                already_dispatched = float(cur.fetchone()[0])

                if already_dispatched + total_to_dispatch > reserved_quantity:
                    raise HTTPException(
                        400,
                        "Intentas despachar más cantidad que la reservada"
                    )
                
                cur.execute("""
                    SELECT COALESCE(SUM(quantity), 0)
                    FROM stock_locations
                    WHERE material_id = %s
                      AND warehouse_id = %s
                """, (reserved_material_id, warehouse_id))

                total_available = float(cur.fetchone()[0])

                if total_available < total_to_dispatch:
                    raise HTTPException(
                        400,
                        "Stock físico insuficiente para el despacho"
                    )


                # =========================================================
                # 🔥 5️⃣ DESCONTAR STOCK FÍSICO
                # =========================================================
                for material_id, quantity in items:

                    remaining = float(quantity)

                    cur.execute("""
                        SELECT rack, level, box, position, quantity
                        FROM stock_locations
                        WHERE material_id = %s
                          AND warehouse_id = %s
                          AND quantity > 0
                        ORDER BY quantity DESC
                        FOR UPDATE
                    """, (material_id, warehouse_id))

                    locations = cur.fetchall()

                    if not locations:
                        raise HTTPException(
                            400,
                            "No hay stock físico disponible"
                        )

                    for rack, level, box, position, available_qty in locations:

                        if remaining <= 0:
                            break

                        consume_qty = min(remaining, float(available_qty))

                        cur.execute("""
                            UPDATE stock_locations
                            SET quantity = quantity - %s
                            WHERE material_id = %s
                              AND warehouse_id = %s
                              AND rack = %s
                              AND level = %s
                              AND box = %s
                              AND position = %s
                        """, (
                            consume_qty,
                            material_id,
                            warehouse_id,
                            rack,
                            level,
                            box,
                            position
                        ))

                        remaining -= consume_qty

                    if remaining > 0:
                        raise HTTPException(
                            400,
                            "Stock insuficiente en ubicaciones físicas"
                        )

                    # Registrar movimiento OUT
                    cur.execute("""
                        INSERT INTO stock_movements (
                            material_id,
                            movement_type,
                            quantity,
                            from_warehouse,
                            to_warehouse,
                            project_id,
                            reference,
                            notes,
                            created_by
                        )
                        VALUES (%s,'OUT',%s,%s,NULL,%s,%s,%s,%s)
                    """, (
                        material_id,
                        quantity,
                        warehouse_id,
                        project_id,
                        f"DISPATCH-{dispatch_id}",
                        "Salida por despacho",
                        current_user.get("username"),
                    ))

                # =========================================================
                # 6️⃣ GENERAR EXCEL
                # =========================================================
                file_path = generate_dispatch_excel(dispatch_id)

                # =========================================================
                # 7️⃣ MARCAR COMO DESPACHADO
                # =========================================================
                cur.execute("""
                    UPDATE stock_dispatches
                    SET status = 'DISPATCHED',
                        dispatched_at = NOW(),
                        file_path = %s
                    WHERE id = %s
                """, (file_path, dispatch_id))
                cur.execute("""
                    UPDATE stock_reservations
                    SET status = 'CONSUMED',
                        released_at = NOW()
                    WHERE id = %s
                """, (reservation_id,))
            conn.commit()

            return {
                "dispatch_id": dispatch_id,
                "status": "DISPATCHED",
                "file_path": file_path
            }

        except Exception:
            conn.rollback()
            raise


# ============================================================
# 🔍 MATERIALES PENDIENTES DE VALIDACIÓN
# ============================================================

def list_pending_materials_service():
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT
                    m.id, m.name, m.code, m.category,
                    m.unit_cost, m.logistics_notes,
                    m.proposed_at, m.precio_referencia,
                    m.proveedor_referencia, m.propuesto_desde,
                    m.cotizacion_origen_id,
                    u.username AS proposed_by_name
                FROM materials m
                LEFT JOIN users u ON u.id = m.proposed_by
                WHERE m.estado = 'PENDIENTE'
                ORDER BY m.proposed_at DESC NULLS LAST
            """)
            rows = cur.fetchall()

    return [
        {
            "id": str(r[0]),
            "name": r[1],
            "code": r[2],
            "category": r[3],
            "unit_cost": float(r[4]) if r[4] is not None else None,
            "logistics_notes": r[5],
            "proposed_at": r[6],
            "precio_referencia": float(r[7]) if r[7] is not None else None,
            "proveedor_referencia": r[8],
            "propuesto_desde": r[9],
            "cotizacion_origen_id": str(r[10]) if r[10] else None,
            "proposed_by": r[11],
        }
        for r in rows
    ]


def count_pending_materials_service():
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT COUNT(*) FROM materials WHERE estado = 'PENDIENTE'")
            row = cur.fetchone()
    return {"count": row[0] if row else 0}


# ============================================================
# ✅ VALIDAR MATERIAL PROPUESTO
# ============================================================

def validate_material_service(material_id: str, payload, user):
    with db_connection() as conn:
        with conn.cursor() as cur:

            cur.execute(
                "SELECT id, estado FROM materials WHERE id = %s",
                (material_id,)
            )
            row = cur.fetchone()
            if not row:
                raise HTTPException(404, "Material no encontrado")
            if row[1] != "PENDIENTE":
                raise HTTPException(400, "Solo se pueden validar materiales en estado PENDIENTE")

            data = payload.model_dump(exclude_unset=True)

            allowed = {"name", "category", "unit_cost", "supplier_name",
                       "supplier_contact", "logistics_notes", "brand", "model"}
            updates = {k: v for k, v in data.items() if k in allowed and v is not None}
            updates["estado"] = "ACTIVO"
            updates["validation_status"] = "VALIDATED"
            updates["validated_by"] = str(user["id"])

            set_parts = [f"{k} = %s" for k in updates] + ["validated_at = NOW()"]
            set_clause = ", ".join(set_parts)
            values = list(updates.values()) + [material_id]

            cur.execute(
                f"UPDATE materials SET {set_clause} WHERE id = %s",
                values
            )
            conn.commit()

    return {"status": "validated", "material_id": material_id}


# ============================================================
# 💡 PROPONER MATERIAL NUEVO (desde APU / Operaciones)
# ============================================================

def proponer_material_service(payload, user):
    import re

    nombre = payload.nombre.strip()
    if not nombre:
        raise HTTPException(400, "El nombre del material es obligatorio")

    # Generar código automático: primeras letras de categoría + nombre + timestamp
    cat = (payload.categoria or "MAT")[:2].upper()
    name_clean = re.sub(r"[^a-zA-Z0-9]", "", nombre)[:4].upper()
    base = f"{cat}{name_clean}"

    with db_connection() as conn:
        with conn.cursor() as cur:
            # Buscar el siguiente número libre para ese prefijo
            cur.execute(
                "SELECT code FROM materials WHERE code LIKE %s ORDER BY code",
                (f"{base}%",)
            )
            existing = [r[0] for r in cur.fetchall()]
            suffix = len(existing) + 1
            code = f"{base}{str(suffix).zfill(2)}"

            cur.execute("""
                INSERT INTO materials (
                    name, code, min_stock, category,
                    estado, precio_referencia, proveedor_referencia,
                    propuesto_desde, cotizacion_origen_id,
                    proposed_by, proposed_at, validation_status
                ) VALUES (
                    %s, %s, 0, %s,
                    'PENDIENTE', %s, %s,
                    %s, %s,
                    %s, NOW(), 'PENDING'
                )
                RETURNING id, name, code, category, estado, precio_referencia, proveedor_referencia, created_at
            """, (
                nombre,
                code,
                payload.categoria or "Material",
                payload.precio_referencia,
                payload.proveedor_referencia,
                "COTIZACION" if payload.plan_id else "MANUAL",
                payload.plan_id,
                str(user["id"]),
            ))
            row = cur.fetchone()
            conn.commit()

    return {
        "id": str(row[0]),
        "name": row[1],
        "code": row[2],
        "category": row[3],
        "estado": row[4],
        "precio_referencia": float(row[5]) if row[5] else None,
        "proveedor_referencia": row[6],
        "created_at": row[7],
    }


# ============================================================
# 📋 LISTAR SUBMISSIONS DE PROYECTO (vista logística)
# ============================================================

def list_project_submissions_service():
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT
                    s.id,
                    s.submission_number,
                    s.status,
                    s.reason,
                    s.submitted_at,
                    s.reviewed_at,
                    s.logistics_notes,
                    pp.id                                                          AS plan_id,
                    COALESCE(p.name, pp.custom_project_name, pp.title)             AS project_name,
                    COALESCE(p.code, pp.project_code)                              AS project_code,
                    u.username                                                      AS engineer_name,
                    COUNT(si.id)                                                    AS item_count,
                    COUNT(si.id) FILTER (WHERE si.logistics_status = 'APPROVED')   AS approved,
                    COUNT(si.id) FILTER (WHERE si.logistics_status = 'REJECTED')   AS rejected,
                    COUNT(si.id) FILTER (WHERE si.logistics_status = 'PARTIAL')    AS partial
                FROM project_plan_submissions s
                JOIN project_plans pp      ON pp.id = s.plan_id
                LEFT JOIN projects p       ON p.id  = pp.project_id
                JOIN users u               ON u.id  = pp.engineer_id
                LEFT JOIN project_plan_submission_items si ON si.submission_id = s.id
                GROUP BY s.id, pp.id, p.id, u.id
                ORDER BY s.submitted_at DESC
                LIMIT 200
            """)
            rows = cur.fetchall()

    return [
        {
            "id": str(r[0]),
            "submission_number": r[1],
            "status": r[2],
            "reason": r[3],
            "submitted_at": r[4],
            "reviewed_at": r[5],
            "logistics_notes": r[6],
            "plan_id": str(r[7]),
            "project_name": r[8],
            "project_code": r[9],
            "engineer_name": r[10],
            "item_count": int(r[11]),
            "approved": int(r[12]),
            "rejected": int(r[13]),
            "partial": int(r[14]),
        }
        for r in rows
    ]


# ============================================================
# 🔍 DETALLE DE SUBMISSION
# ============================================================

def get_submission_detail_service(submission_id: str):
    with db_connection() as conn:
        with conn.cursor() as cur:

            cur.execute("""
                SELECT
                    s.id, s.submission_number, s.status, s.reason,
                    s.submitted_at, s.reviewed_at, s.logistics_notes,
                    pp.id, COALESCE(p.name, pp.custom_project_name, pp.title),
                    COALESCE(p.code, pp.project_code),
                    u.username
                FROM project_plan_submissions s
                JOIN project_plans pp   ON pp.id = s.plan_id
                LEFT JOIN projects p   ON p.id  = pp.project_id
                JOIN users u           ON u.id  = pp.engineer_id
                WHERE s.id = %s
            """, (submission_id,))
            row = cur.fetchone()
            if not row:
                raise HTTPException(404, "Requerimiento no encontrado")

            sub = {
                "id": str(row[0]),
                "submission_number": row[1],
                "status": row[2],
                "reason": row[3],
                "submitted_at": row[4],
                "reviewed_at": row[5],
                "logistics_notes": row[6],
                "plan_id": str(row[7]),
                "project_name": row[8],
                "project_code": row[9],
                "engineer_name": row[10],
            }

            cur.execute("""
                SELECT
                    si.id, si.material_id, si.material_name, si.material_code,
                    si.category, si.quantity, si.unit_cost, si.wear_percentage,
                    si.stock_available, si.logistics_status,
                    si.approved_quantity, si.logistics_notes, si.reviewed_at
                FROM project_plan_submission_items si
                WHERE si.submission_id = %s
                ORDER BY si.category, si.material_name
            """, (submission_id,))
            item_rows = cur.fetchall()

    items = []
    for r in item_rows:
        unit_cost = float(r[6]) if r[6] is not None else 0.0
        qty = float(r[5])
        wear = float(r[7])
        items.append({
            "id": str(r[0]),
            "material_id": str(r[1]),
            "material_name": r[2],
            "material_code": r[3],
            "category": r[4],
            "quantity": qty,
            "unit_cost": unit_cost,
            "wear_percentage": wear,
            "effective_cost": qty * unit_cost * (wear / 100.0),
            "stock_available": float(r[8]),
            "logistics_status": r[9],
            "approved_quantity": float(r[10]) if r[10] is not None else None,
            "logistics_notes": r[11],
            "reviewed_at": r[12],
        })

    sub["items"] = items
    sub["total_cost"] = sum(i["effective_cost"] for i in items)
    return sub


# ============================================================
# ✅ REVISAR ÍTEM DE SUBMISSION (aprobar / rechazar)
# ============================================================

def review_submission_item_service(submission_id: str, item_id: str, payload, user):
    with db_connection() as conn:
        with conn.cursor() as cur:

            cur.execute(
                "SELECT id FROM project_plan_submissions WHERE id = %s",
                (submission_id,)
            )
            if not cur.fetchone():
                raise HTTPException(404, "Requerimiento no encontrado")

            cur.execute(
                "SELECT id, quantity FROM project_plan_submission_items WHERE id = %s AND submission_id = %s",
                (item_id, submission_id)
            )
            item = cur.fetchone()
            if not item:
                raise HTTPException(404, "Ítem no encontrado en este requerimiento")

            status = payload.logistics_status
            approved_qty = payload.approved_quantity
            notes = payload.logistics_notes

            # If PARTIAL, approved_quantity is required
            if status == "PARTIAL" and (approved_qty is None or float(approved_qty) <= 0):
                raise HTTPException(400, "Para aprobación parcial debes indicar la cantidad aprobada")
            if status == "APPROVED":
                approved_qty = float(item[1])

            cur.execute("""
                UPDATE project_plan_submission_items
                SET logistics_status = %s,
                    approved_quantity = %s,
                    logistics_notes   = %s,
                    reviewed_at       = NOW()
                WHERE id = %s
            """, (status, approved_qty, notes, item_id))

            # Update plan item submission_status to match
            item_status_map = {
                "APPROVED": "APPROVED",
                "PARTIAL":  "PARTIAL",
                "REJECTED": "REJECTED",
            }
            if status in item_status_map:
                cur.execute("""
                    UPDATE project_plan_items
                    SET submission_status = %s
                    WHERE id = (
                        SELECT plan_item_id FROM project_plan_submission_items WHERE id = %s
                    )
                """, (item_status_map[status], item_id))

            # Auto-update submission status based on all items
            cur.execute("""
                SELECT
                    COUNT(*) AS total,
                    COUNT(*) FILTER (WHERE logistics_status = 'PENDING')  AS pending,
                    COUNT(*) FILTER (WHERE logistics_status = 'APPROVED') AS approved,
                    COUNT(*) FILTER (WHERE logistics_status = 'REJECTED') AS rejected,
                    COUNT(*) FILTER (WHERE logistics_status = 'PARTIAL')  AS partial
                FROM project_plan_submission_items
                WHERE submission_id = %s
            """, (submission_id,))
            counts = cur.fetchone()
            total, pending, approved, rejected, partial = (int(x) for x in counts)

            if pending == 0:
                if rejected == total:
                    new_sub_status = "REJECTED"
                elif approved == total:
                    new_sub_status = "APPROVED"
                else:
                    new_sub_status = "PARTIAL"
                cur.execute(
                    "UPDATE project_plan_submissions SET status = %s, reviewed_at = NOW(), reviewed_by = %s WHERE id = %s",
                    (new_sub_status, str(user["id"]), submission_id)
                )

            conn.commit()

    return {"status": "ok", "item_id": item_id}


# ============================================================
# CALIBRATION RECORDS
# ============================================================

def get_calibration_alerts():
    """Materials with calibration_required=True expiring within 30 days or no record."""
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                WITH latest AS (
                    SELECT DISTINCT ON (material_id)
                        material_id, calibrated_at, expires_at,
                        certificate_url, technician, notes, id AS record_id
                    FROM calibration_records
                    ORDER BY material_id, expires_at DESC
                )
                SELECT
                    m.id, m.code, m.name, m.category,
                    m.calibration_interval_days,
                    l.calibrated_at, l.expires_at, l.certificate_url,
                    l.technician, l.notes, l.record_id,
                    (l.expires_at - CURRENT_DATE) AS days_left
                FROM materials m
                LEFT JOIN latest l ON l.material_id = m.id
                WHERE m.calibration_required = TRUE
                  AND (l.expires_at IS NULL OR l.expires_at <= CURRENT_DATE + INTERVAL '30 days')
                ORDER BY COALESCE(l.expires_at, '1900-01-01'::date) ASC
            """)
            rows = cur.fetchall()
    return [
        {
            "material_id":     str(r[0]),
            "material_code":   r[1],
            "material_name":   r[2],
            "category":        r[3],
            "interval_days":   r[4],
            "calibrated_at":   r[5].isoformat() if r[5] else None,
            "expires_at":      r[6].isoformat() if r[6] else None,
            "certificate_url": r[7],
            "technician":      r[8],
            "notes":           r[9],
            "last_record_id":  str(r[10]) if r[10] else None,
            "days_left":       int(r[11]) if r[11] is not None else None,
        }
        for r in rows
    ]


def get_all_calibrations_list():
    """All calibration-required materials with their latest calibration status."""
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                WITH latest AS (
                    SELECT DISTINCT ON (material_id)
                        material_id, calibrated_at, expires_at,
                        certificate_url, technician, id AS record_id
                    FROM calibration_records
                    ORDER BY material_id, expires_at DESC
                )
                SELECT
                    m.id, m.code, m.name, m.category,
                    m.calibration_interval_days,
                    l.calibrated_at, l.expires_at, l.certificate_url,
                    l.technician, l.record_id,
                    (l.expires_at - CURRENT_DATE) AS days_left
                FROM materials m
                LEFT JOIN latest l ON l.material_id = m.id
                WHERE m.calibration_required = TRUE
                ORDER BY COALESCE(l.expires_at, '1900-01-01'::date) ASC
            """)
            rows = cur.fetchall()
    return [
        {
            "material_id":     str(r[0]),
            "material_code":   r[1],
            "material_name":   r[2],
            "category":        r[3],
            "interval_days":   r[4],
            "calibrated_at":   r[5].isoformat() if r[5] else None,
            "expires_at":      r[6].isoformat() if r[6] else None,
            "certificate_url": r[7],
            "technician":      r[8],
            "last_record_id":  str(r[9]) if r[9] else None,
            "days_left":       int(r[10]) if r[10] is not None else None,
        }
        for r in rows
    ]


def get_calibration_history(material_id: str):
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT cr.id, cr.calibrated_at, cr.expires_at,
                       cr.certificate_url, cr.technician, cr.notes,
                       cr.created_by, cr.created_at, m.name, m.code
                FROM calibration_records cr
                JOIN materials m ON m.id = cr.material_id
                WHERE cr.material_id = %s
                ORDER BY cr.expires_at DESC
            """, (material_id,))
            rows = cur.fetchall()
    return [
        {
            "id":              str(r[0]),
            "calibrated_at":   r[1].isoformat() if r[1] else None,
            "expires_at":      r[2].isoformat() if r[2] else None,
            "certificate_url": r[3],
            "technician":      r[4],
            "notes":           r[5],
            "created_by":      r[6],
            "created_at":      r[7].isoformat() if r[7] else None,
            "material_name":   r[8],
            "material_code":   r[9],
        }
        for r in rows
    ]


def add_calibration_record(payload: dict, user_id: str):
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                INSERT INTO calibration_records
                    (material_id, calibrated_at, expires_at,
                     certificate_url, technician, notes, created_by)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
                RETURNING id
            """, (
                payload["material_id"],
                payload["calibrated_at"],
                payload["expires_at"],
                payload.get("certificate_url"),
                payload.get("technician"),
                payload.get("notes"),
                user_id,
            ))
            record_id = str(cur.fetchone()[0])
            if payload.get("certificate_url"):
                cur.execute(
                    "UPDATE materials SET calibration_cert_url=%s WHERE id=%s",
                    (payload["certificate_url"], payload["material_id"])
                )
            conn.commit()
    return {"id": record_id, "ok": True}


def set_material_calibration_flag(material_id: str, required: bool, interval_days: int = None):
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "UPDATE materials SET calibration_required=%s, calibration_interval_days=%s WHERE id=%s",
                (required, interval_days, material_id)
            )
            conn.commit()
    return {"ok": True}


# ============================================================
# PURCHASE ITEMS (Lista de compras)
# ============================================================

def get_purchase_items(project_id: str = None, status: str = None):
    with db_connection() as conn:
        with conn.cursor() as cur:
            query = """
                SELECT
                    pi.id, pi.material_id, pi.material_name_free,
                    pi.qty_needed, pi.unit, pi.project_id,
                    pi.source, pi.reason, pi.priority, pi.status,
                    pi.supplier_notes, pi.created_by, pi.created_at, pi.updated_at,
                    m.name  AS material_name,
                    m.code  AS material_code,
                    m.unit  AS material_unit,
                    p.code  AS project_code,
                    p.name  AS project_name
                FROM purchase_items pi
                LEFT JOIN materials m ON m.id = pi.material_id
                LEFT JOIN projects  p ON p.id = pi.project_id
                WHERE 1=1
            """
            params = []
            if project_id:
                query += " AND pi.project_id = %s"
                params.append(project_id)
            if status:
                query += " AND pi.status = %s"
                params.append(status)
            query += " ORDER BY pi.created_at DESC"
            cur.execute(query, params)
            rows = cur.fetchall()

    return [
        {
            "id":                str(r[0]),
            "material_id":       str(r[1]) if r[1] else None,
            "material_name_free":r[2],
            "qty_needed":        float(r[3]),
            "unit":              r[4] or r[16],
            "project_id":        str(r[5]) if r[5] else None,
            "source":            r[6],
            "reason":            r[7],
            "priority":          r[8],
            "status":            r[9],
            "supplier_notes":    r[10],
            "created_by":        r[11],
            "created_at":        r[12].isoformat() if r[12] else None,
            "updated_at":        r[13].isoformat() if r[13] else None,
            "material_name":     r[14] or r[2],
            "material_code":     r[15],
            "project_code":      r[17],
            "project_name":      r[18],
        }
        for r in rows
    ]


def add_purchase_item(payload: dict, user_id: str):
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                INSERT INTO purchase_items
                    (material_id, material_name_free, qty_needed, unit,
                     project_id, source, reason, priority, supplier_notes, created_by)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id
            """, (
                payload.get("material_id"),
                payload.get("material_name_free"),
                payload.get("qty_needed", 1),
                payload.get("unit"),
                payload.get("project_id"),
                payload.get("source", "MANUAL"),
                payload.get("reason"),
                payload.get("priority", "NORMAL"),
                payload.get("supplier_notes"),
                user_id,
            ))
            item_id = str(cur.fetchone()[0])
            conn.commit()
    return {"id": item_id, "ok": True}


def bulk_add_purchase_items(items: list, user_id: str):
    """Bulk insert from requirements gap, skipping duplicates."""
    inserted = 0
    with db_connection() as conn:
        with conn.cursor() as cur:
            for item in items:
                cur.execute("""
                    SELECT id FROM purchase_items
                    WHERE material_id=%s AND project_id=%s
                      AND status NOT IN ('RECEIVED','CANCELLED')
                """, (item.get("material_id"), item.get("project_id")))
                if cur.fetchone():
                    continue
                cur.execute("""
                    INSERT INTO purchase_items
                        (material_id, material_name_free, qty_needed, unit,
                         project_id, source, reason, priority, created_by)
                    VALUES (%s, %s, %s, %s, %s, 'AUTO_GAP', %s, %s, %s)
                """, (
                    item.get("material_id"),
                    item.get("material_name_free"),
                    item.get("qty_needed", 1),
                    item.get("unit"),
                    item.get("project_id"),
                    item.get("reason", "Faltante en analisis de brecha de proyecto"),
                    item.get("priority", "NORMAL"),
                    user_id,
                ))
                inserted += 1
            conn.commit()
    return {"inserted": inserted, "ok": True}


def update_purchase_item_status(item_id: str, status: str, notes: str = None):
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                UPDATE purchase_items
                SET status=%s,
                    supplier_notes=COALESCE(%s, supplier_notes),
                    updated_at=NOW()
                WHERE id=%s
            """, (status, notes, item_id))
            conn.commit()
    return {"ok": True}


def delete_purchase_item(item_id: str):
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM purchase_items WHERE id=%s", (item_id,))
            conn.commit()
    return {"ok": True}


# ============================================================
# REQUIREMENTS GAP (Brecha requerimientos por proyecto)
# ============================================================

def get_project_requirements_gap(project_id: str):
    """Compares material_requests for project vs available stock."""
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT id, code, name FROM projects WHERE id=%s", (project_id,))
            proj = cur.fetchone()
            if not proj:
                raise ValueError("Proyecto no encontrado")

            cur.execute("""
                SELECT
                    mr.related_material_id,
                    SUM(mr.quantity)                              AS total_requested,
                    MAX(mr.priority)                              AS priority,
                    MIN(mr.needed_by)                            AS needed_by,
                    m.code, m.name, m.unit, m.category,
                    COALESCE(stock.total_available, 0)           AS stock_available,
                    GREATEST(0, SUM(mr.quantity) - COALESCE(stock.total_available, 0)) AS shortage
                FROM material_requests mr
                LEFT JOIN materials m ON m.id = mr.related_material_id
                LEFT JOIN (
                    SELECT material_id, SUM(stock_available) AS total_available
                    FROM vw_stock_availability
                    GROUP BY material_id
                ) stock ON stock.material_id = mr.related_material_id
                WHERE mr.project_id = %s
                  AND mr.status NOT IN ('REJECTED', 'CANCELLED')
                  AND mr.related_material_id IS NOT NULL
                GROUP BY mr.related_material_id, m.code, m.name, m.unit, m.category, stock.total_available
                ORDER BY shortage DESC, priority DESC
            """, (project_id,))
            rows = cur.fetchall()

            cur.execute("""
                SELECT material_id FROM purchase_items
                WHERE project_id=%s AND status NOT IN ('RECEIVED','CANCELLED')
            """, (project_id,))
            in_purchase_set = {str(r[0]) for r in cur.fetchall()}

    items = []
    for r in rows:
        mat_id    = str(r[0]) if r[0] else None
        requested = float(r[1])
        available = float(r[8])
        shortage  = float(r[9])
        items.append({
            "material_id":      mat_id,
            "total_requested":  requested,
            "priority":         r[2],
            "needed_by":        r[3].isoformat() if r[3] else None,
            "material_code":    r[4],
            "material_name":    r[5],
            "unit":             r[6],
            "category":         r[7],
            "stock_available":  available,
            "shortage":         shortage,
            "covered_pct":      min(100, round(available / requested * 100, 1)) if requested > 0 else 100,
            "in_purchase_list": mat_id in in_purchase_set if mat_id else False,
        })

    fully_covered = sum(1 for i in items if i["shortage"] == 0)
    partial       = sum(1 for i in items if 0 < i["shortage"] < i["total_requested"])
    missing       = sum(1 for i in items if i["stock_available"] == 0 and i["total_requested"] > 0)

    return {
        "project": {"id": str(proj[0]), "code": proj[1], "name": proj[2]},
        "items":   items,
        "summary": {
            "total_materials":  len(items),
            "fully_covered":    fully_covered,
            "partial_coverage": partial,
            "not_in_stock":     missing,
        },
    }


# ══════════════════════════════════════════════════════════════════════════════
# 📥 EXPORTAR MATERIALES A EXCEL
# ══════════════════════════════════════════════════════════════════════════════

def export_materials_excel_service():
    from app.core.export_utils import (
        write_title_row, write_header_row, write_data_row,
        set_column_widths, fmt_date, fmt_num, excel_response,
    )
    import openpyxl

    materials = get_materials_service()
    fecha = datetime.now().strftime("%d/%m/%Y %H:%M")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Catálogo de Materiales"

    headers = [
        "Código", "Nombre", "Categoría", "Stock mínimo",
        "Marca", "Modelo", "N° Serie",
        "Proveedor", "Contacto Proveedor",
        "Costo Unitario", "Vida Útil (años)",
        "Fecha Compra", "Vence Garantía",
        "Aliases",
    ]
    widths = [12, 38, 18, 13, 16, 16, 16, 28, 22, 14, 14, 14, 14, 30]

    write_title_row(ws, f"CeShark ERP — Catálogo de Materiales — {fecha}", len(headers))
    write_header_row(ws, headers, row=2)
    set_column_widths(ws, widths)
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 18

    for i, m in enumerate(materials, start=1):
        write_data_row(ws, i + 2, [
            m.get("code", ""),
            m.get("name", ""),
            m.get("category", ""),
            fmt_num(m.get("min_stock"), 0) if m.get("min_stock") else "0",
            m.get("brand", ""),
            m.get("model", ""),
            m.get("serial_number", ""),
            m.get("supplier_name", ""),
            m.get("supplier_contact", ""),
            fmt_num(m.get("unit_cost")) if m.get("unit_cost") is not None else "",
            m.get("useful_life_years", ""),
            fmt_date(m.get("purchase_date")),
            fmt_date(m.get("warranty_expires")),
            ", ".join(m.get("aliases", [])),
        ], alternate=(i % 2 == 0))

    # Hoja de resumen
    ws2 = wb.create_sheet("Resumen")
    from app.core.export_utils import _fill, _font_bold_dark, _font_normal, _align, PRIMARY_HEX
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 14
    ws2["A1"] = "Total de materiales registrados"
    ws2["B1"] = len(materials)
    ws2["A1"].font = _font_bold_dark()
    ws2["B1"].font = _font_bold_dark()

    categorias = {}
    for m in materials:
        cat = m.get("category") or "Sin categoría"
        categorias[cat] = categorias.get(cat, 0) + 1
    ws2["A2"] = ""
    ws2["A3"] = "Por categoría:"
    ws2["A3"].font = _font_bold_dark()
    for row_i, (cat, count) in enumerate(sorted(categorias.items()), start=4):
        ws2[f"A{row_i}"] = f"  {cat}"
        ws2[f"B{row_i}"] = count
        ws2[f"A{row_i}"].font = _font_normal()
        ws2[f"B{row_i}"].font = _font_normal()

    return excel_response(wb, f"materiales_{datetime.now().strftime('%Y%m%d')}.xlsx")


# ══════════════════════════════════════════════════════════════════════════════
# 📦 EXPORTAR REPORTE DE STOCK A EXCEL
# ══════════════════════════════════════════════════════════════════════════════

def export_stock_excel_service():
    from app.core.export_utils import (
        write_title_row, write_header_row, write_data_row,
        set_column_widths, fmt_num, excel_response, _fill, _font_bold_dark, ACCENT_HEX,
    )
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    fecha = datetime.now().strftime("%d/%m/%Y %H:%M")

    with db_connection() as conn:
        with conn.cursor() as cur:
            # Stock actual por material + almacén desde stock_locations
            cur.execute("""
                SELECT
                    m.code,
                    m.name,
                    m.category,
                    m.min_stock,
                    w.code  AS warehouse_code,
                    w.name  AS warehouse_name,
                    COALESCE(SUM(sl.quantity), 0) AS stock_actual
                FROM materials m
                CROSS JOIN warehouses w
                LEFT JOIN stock_locations sl
                    ON sl.material_id = m.id AND sl.warehouse_id = w.id
                GROUP BY m.id, m.code, m.name, m.category, m.min_stock,
                         w.id, w.code, w.name
                HAVING COALESCE(SUM(sl.quantity), 0) > 0
                   OR m.min_stock > 0
                ORDER BY m.name, w.name
            """)
            rows = cur.fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stock Disponible"

    headers = ["Cód. Material", "Nombre Material", "Categoría",
               "Stock Mínimo", "Cód. Almacén", "Almacén",
               "Stock Actual", "Estado"]
    widths = [13, 38, 16, 13, 13, 28, 13, 14]

    write_title_row(ws, f"CeShark ERP — Reporte de Stock — {fecha}", len(headers))
    write_header_row(ws, headers, row=2)
    set_column_widths(ws, widths)
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 18

    for i, r in enumerate(rows, start=1):
        code, name, cat, min_stock, wh_code, wh_name, stock = r
        min_s = float(min_stock) if min_stock else 0
        stock_f = float(stock)
        if stock_f <= 0:
            estado = "SIN STOCK"
        elif min_s > 0 and stock_f < min_s:
            estado = "STOCK BAJO"
        else:
            estado = "OK"

        write_data_row(ws, i + 2, [
            code, name, cat or "",
            fmt_num(min_s, 0),
            wh_code, wh_name,
            fmt_num(stock_f, 2),
            estado,
        ], alternate=(i % 2 == 0))

        # Color semáforo en columna Estado (col 8)
        estado_cell = ws.cell(row=i + 2, column=8)
        if estado == "SIN STOCK":
            estado_cell.font = Font(size=9, bold=True, color="991B1B")
            estado_cell.fill = PatternFill("solid", fgColor="FEE2E2")
        elif estado == "STOCK BAJO":
            estado_cell.font = Font(size=9, bold=True, color="92400E")
            estado_cell.fill = PatternFill("solid", fgColor="FEF3C7")
        else:
            estado_cell.font = Font(size=9, bold=True, color="065F46")
            estado_cell.fill = PatternFill("solid", fgColor="D1FAE5")

    return excel_response(wb, f"stock_{datetime.now().strftime('%Y%m%d')}.xlsx")
