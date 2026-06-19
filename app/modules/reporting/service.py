from app.core.database import db_connection


def _f(x):
    """float() seguro: None/NULL -> 0.0 (evita 500 cuando una vista agrega sin datos)."""
    return float(x) if x is not None else 0.0


def get_material_requests_sla_kpi_service():
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT
                    total_decided_requests,
                    within_sla,
                    overdue,
                    sla_compliance_rate,
                    avg_decision_time_hours
                FROM vw_kpi_material_requests_sla
            """)

            row = cur.fetchone()

            if not row:
                return {
                    "total_decided_requests": 0,
                    "within_sla": 0,
                    "overdue": 0,
                    "sla_compliance_rate": 0.0,
                    "avg_decision_time_hours": 0.0,
                }

            return {
                "total_decided_requests": row[0],
                "within_sla": row[1],
                "overdue": row[2],
                "sla_compliance_rate": _f(row[3]),
                "avg_decision_time_hours": _f(row[4]),
            }


def get_material_requests_summary_kpi_service():
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT
                    total_requests,
                    pending_requests,
                    approved_requests,
                    rejected_requests,
                    overdue_requests
                FROM vw_kpi_material_requests_summary
            """)

            row = cur.fetchone()

            if not row:
                return {
                    "total_requests": 0,
                    "pending_requests": 0,
                    "approved_requests": 0,
                    "rejected_requests": 0,
                    "overdue_requests": 0,
                }

            return {
                "total_requests": row[0],
                "pending_requests": row[1],
                "approved_requests": row[2],
                "rejected_requests": row[3],
                "overdue_requests": row[4],
            }

def get_material_requests_lead_time_kpi_service():
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT
                    avg_lead_time_hours,
                    min_lead_time_hours,
                    max_lead_time_hours,
                    p95_lead_time_hours
                FROM vw_kpi_material_requests_lead_time
            """)

            row = cur.fetchone()

            if not row:
                return {
                    "avg_lead_time_hours": 0.0,
                    "min_lead_time_hours": 0.0,
                    "max_lead_time_hours": 0.0,
                    "p95_lead_time_hours": 0.0,
                }

            return {
                "avg_lead_time_hours": _f(row[0]),
                "min_lead_time_hours": _f(row[1]),
                "max_lead_time_hours": _f(row[2]),
                "p95_lead_time_hours": _f(row[3]),
            }

def get_material_requests_by_approver_kpi_service():
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT
                    approver_id,
                    total_decisions,
                    approved_count,
                    rejected_count,
                    sla_compliance_rate,
                    avg_decision_time_hours
                FROM vw_kpi_material_requests_by_approver
                ORDER BY total_decisions DESC
            """)

            rows = cur.fetchall()

            return [
                {
                    "approver_id": r[0],
                    "total_decisions": r[1],
                    "approved_count": r[2],
                    "rejected_count": r[3],
                    "sla_compliance_rate": _f(r[4]),
                    "avg_decision_time_hours": _f(r[5]),
                }
                for r in rows
            ]

def get_material_requests_monthly_kpi_service():
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT
                    month,
                    total_requests,
                    approved_requests,
                    rejected_requests,
                    pending_requests,
                    overdue_requests,
                    sla_compliance_rate
                FROM vw_kpi_material_requests_monthly
                ORDER BY month
            """)

            rows = cur.fetchall()

            return [
                {
                    "month": r[0],
                    "total_requests": r[1],
                    "approved_requests": r[2],
                    "rejected_requests": r[3],
                    "pending_requests": r[4],
                    "overdue_requests": r[5],
                    "sla_compliance_rate": _f(r[6]),
                }
                for r in rows
            ]

def get_requests_kpi_summary_service():
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT
                    total_requests,
                    pending_requests,
                    approved_requests,
                    rejected_requests,
                    requests_with_reservation,
                    requests_without_reservation
                FROM vw_requests_kpi_summary
            """)

            row = cur.fetchone()

            if not row:
                return {}

            return {
                "total_requests": row[0],
                "pending_requests": row[1],
                "approved_requests": row[2],
                "rejected_requests": row[3],
                "requests_with_reservation": row[4],
                "requests_without_reservation": row[5],
            }


def get_dashboard_kpis_service(desde: str = None, hasta: str = None):
    # 1. Logistics KPIs (reuse existing service functions)
    logistics_summary = get_material_requests_summary_kpi_service()
    logistics_sla = get_material_requests_sla_kpi_service()
    logistics_lead = get_material_requests_lead_time_kpi_service()

    date_filter_oc = ""
    date_params = []
    if desde:
        date_filter_oc += " AND created_at >= %s::date"
        date_params.append(desde)
    if hasta:
        date_filter_oc += " AND created_at <= %s::date + INTERVAL '1 day'"
        date_params.append(hasta)

    with db_connection() as conn:
        with conn.cursor() as cur:
            # 2. Operations KPIs
            cur.execute("""
                SELECT status, COUNT(*)
                FROM ordenes_trabajo
                GROUP BY status
            """)
            ot_status_counts = {r[0]: r[1] for r in cur.fetchall()}

            cur.execute("""
                SELECT id, code, titulo, status, fecha_fin_plan, asignado_a
                FROM ordenes_trabajo
                WHERE status NOT IN ('COMPLETADA', 'CERRADA', 'CANCELADA')
                  AND fecha_fin_plan < NOW()
                ORDER BY fecha_fin_plan ASC
            """)
            delayed_ots = [
                {
                    "id": str(r[0]),
                    "code": r[1],
                    "titulo": r[2],
                    "status": r[3],
                    "fecha_fin_plan": r[4].isoformat() if r[4] else None,
                    "asignado_a": str(r[5]) if r[5] else None,
                }
                for r in cur.fetchall()
            ]

            # Cotizaciones por estado
            cur.execute("""
                SELECT status, COUNT(*)
                FROM presupuesto_config
                WHERE numero_cotizacion IS NOT NULL
                GROUP BY status
            """)
            cotizaciones_status = {r[0]: r[1] for r in cur.fetchall()}

            # Compras — OCs por estado y gasto
            cur.execute(f"""
                SELECT status, COUNT(*)
                FROM ordenes_compra
                WHERE 1=1 {date_filter_oc}
                GROUP BY status
            """, date_params)
            oc_status_counts = {r[0]: r[1] for r in cur.fetchall()}

            cur.execute(f"""
                SELECT COALESCE(SUM(oci.cantidad_pedida * oci.precio_unitario), 0)
                FROM ordenes_compra_items oci
                JOIN ordenes_compra oc ON oc.id = oci.oc_id
                WHERE oc.status IN ('RECIBIDA', 'CERRADA') {date_filter_oc.replace('created_at', 'oc.created_at')}
            """, date_params)
            gasto_compras = float(cur.fetchone()[0] or 0)

            # Logística extra — materiales bajo stock y despachos pendientes
            cur.execute("""
                SELECT COUNT(*) FROM (
                    SELECT m.id
                    FROM materials m
                    LEFT JOIN stock_locations sl ON sl.material_id = m.id
                    GROUP BY m.id, m.min_stock
                    HAVING COALESCE(SUM(sl.quantity), 0) <= COALESCE(m.min_stock, 0)
                ) sub
            """)
            materiales_bajo_stock = cur.fetchone()[0]

            cur.execute("""
                SELECT COUNT(*) FROM stock_dispatches
                WHERE status IN ('PENDING', 'READY')
            """)
            despachos_pendientes = cur.fetchone()[0]

            # 3. Administration KPIs
            cur.execute("""
                SELECT COUNT(*) FROM servicio_requerimientos
                WHERE estado NOT IN ('Finalizado', 'Cancelado')
            """)
            active_requirements_count = cur.fetchone()[0]

            cur.execute("""
                SELECT categoria, SUM(total)
                FROM servicio_requerimiento_costos
                GROUP BY categoria
                ORDER BY SUM(total) DESC
            """)
            costos_por_categoria = [
                {"categoria": r[0], "total": float(r[1]) if r[1] is not None else 0.0}
                for r in cur.fetchall()
            ]

            # Planificación — conteos por estado
            cur.execute("""
                SELECT estado, COUNT(*)
                FROM planificacion_semanal
                GROUP BY estado
            """)
            planificacion_counts = {r[0]: r[1] for r in cur.fetchall()}

            # Productividad por persona (mes actual o rango)
            prod_filter = ""
            prod_params = []
            if desde:
                prod_filter += " AND rp.fecha >= %s::date"
                prod_params.append(desde)
            if hasta:
                prod_filter += " AND rp.fecha <= %s::date"
                prod_params.append(hasta)
            if not desde and not hasta:
                prod_filter += " AND rp.fecha >= date_trunc('month', CURRENT_DATE)"

            cur.execute(f"""
                SELECT u.username,
                       COUNT(rp.id) AS registros,
                       COALESCE(SUM(rp.duracion_minutos), 0) AS minutos_totales
                FROM registro_productividad rp
                JOIN users u ON u.id = rp.user_id
                WHERE 1=1 {prod_filter}
                GROUP BY u.id, u.username
                ORDER BY minutos_totales DESC
                LIMIT 15
            """, prod_params)
            productividad_por_persona = [
                {
                    "username": r[0],
                    "registros": r[1],
                    "horas_totales": round((r[2] or 0) / 60, 1),
                }
                for r in cur.fetchall()
            ]

            # 4. Overloaded users
            cur.execute("""
                SELECT u.id, u.username,
                       (SELECT COUNT(*) FROM planificacion_semanal
                        WHERE (responsable_id = u.id OR responsables_ids LIKE '%%' || u.id::text || '%%')
                          AND estado NOT IN ('Completado', 'Cancelado')) AS active_plan,
                       (SELECT COUNT(*) FROM ordenes_trabajo
                        WHERE asignado_a = u.id AND status IN ('PENDIENTE', 'EN_EJECUCION', 'PAUSADA')) AS active_ots
                FROM users u
                WHERE u.is_active = TRUE
            """)
            overloaded_users = []
            for uid, username, active_plan, active_ots in cur.fetchall():
                total_active = (active_plan or 0) + (active_ots or 0)
                if total_active >= 5:
                    overloaded_users.append({
                        "id": str(uid),
                        "username": username,
                        "active_plan": active_plan,
                        "active_ots": active_ots,
                        "total_active": total_active,
                    })

            cur.execute("""
                SELECT id, tarea, cliente, prioridad, estado, fecha_limite, responsable_id
                FROM planificacion_semanal
                WHERE estado NOT IN ('Completado', 'Cancelado')
                  AND (estado = 'Retraso' OR (fecha_limite IS NOT NULL AND fecha_limite < CURRENT_DATE))
                ORDER BY fecha_limite ASC NULLS LAST
            """)
            delayed_planning = [
                {
                    "id": str(r[0]),
                    "tarea": r[1],
                    "cliente": r[2],
                    "prioridad": r[3],
                    "estado": r[4],
                    "fecha_limite": r[5].isoformat() if r[5] else None,
                    "responsable_id": str(r[6]) if r[6] else None,
                }
                for r in cur.fetchall()
            ]

    return {
        "logistics": {
            "summary": logistics_summary,
            "sla": logistics_sla,
            "lead_time": logistics_lead,
            "materiales_bajo_stock": materiales_bajo_stock,
            "despachos_pendientes": despachos_pendientes,
        },
        "operations": {
            "total_ots": sum(ot_status_counts.values()),
            "status_counts": ot_status_counts,
            "delayed_count": len(delayed_ots),
            "delayed_ots": delayed_ots,
            "cotizaciones_status": cotizaciones_status,
            "total_cotizaciones": sum(cotizaciones_status.values()),
        },
        "compras": {
            "status_counts": oc_status_counts,
            "total_ocs": sum(oc_status_counts.values()),
            "gasto_recibido": gasto_compras,
        },
        "admin": {
            "active_requirements_count": active_requirements_count,
            "costos_por_categoria": costos_por_categoria,
            "total_requirements_cost": sum(c["total"] for c in costos_por_categoria),
            "planificacion_counts": planificacion_counts,
            "productividad_por_persona": productividad_por_persona,
        },
        "weak_points": {
            "delayed_ots": delayed_ots[:5],
            "delayed_planning": delayed_planning[:5],
            "overloaded_users": sorted(overloaded_users, key=lambda x: x["total_active"], reverse=True),
            "low_sla_alert": (logistics_sla.get("sla_compliance_rate", 100.0) < 80.0),
        },
    }


def export_dashboard_kpis_excel_service(
    desde: str = None,
    hasta: str = None,
    include_logistics: bool = True,
    include_operations: bool = True,
    include_compras: bool = True,
    include_admin: bool = True,
    include_weak_points: bool = True,
):
    from datetime import datetime
    from app.core.export_utils import (
        write_title_row, write_header_row, write_data_row,
        set_column_widths, excel_response, fmt_num,
    )
    import openpyxl

    data = get_dashboard_kpis_service(desde=desde, hasta=hasta)
    fecha = datetime.now().strftime("%d/%m/%Y %H:%M")
    periodo = ""
    if desde or hasta:
        periodo = f" — Período: {desde or 'inicio'} a {hasta or 'hoy'}"

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    def add_sheet(name, title, headers, rows, widths):
        ws = wb.create_sheet(name)
        write_title_row(ws, f"CeShark ERP — {title}{periodo} — {fecha}", len(headers))
        write_header_row(ws, headers, row=2)
        set_column_widths(ws, widths)
        for i, row in enumerate(rows, start=1):
            write_data_row(ws, i + 2, row, alternate=(i % 2 == 0))
        return ws

    if include_logistics:
        log = data["logistics"]
        add_sheet(
            "Logística",
            "KPIs de Logística",
            ["Indicador", "Valor"],
            [
                ["Solicitudes totales", log["summary"]["total_requests"]],
                ["Pendientes", log["summary"]["pending_requests"]],
                ["Aprobadas", log["summary"]["approved_requests"]],
                ["Rechazadas", log["summary"]["rejected_requests"]],
                ["Vencidas SLA", log["summary"]["overdue_requests"]],
                ["Cumplimiento SLA (%)", f"{log['sla']['sla_compliance_rate']:.1f}"],
                ["Lead time promedio (h)", f"{log['lead_time']['avg_lead_time_hours']:.1f}"],
                ["Materiales bajo stock mínimo", log["materiales_bajo_stock"]],
                ["Despachos pendientes", log["despachos_pendientes"]],
            ],
            [35, 18],
        )

    if include_operations:
        op = data["operations"]
        rows = [["Total OTs", op["total_ots"]], ["OTs retrasadas", op["delayed_count"]]]
        for status, count in op["status_counts"].items():
            rows.append([f"OT — {status}", count])
        rows.append(["", ""])
        rows.append(["Cotizaciones totales", op["total_cotizaciones"]])
        for status, count in op.get("cotizaciones_status", {}).items():
            rows.append([f"Cotización — {status}", count])
        add_sheet("Operaciones", "KPIs de Operaciones", ["Indicador", "Cantidad"], rows, [35, 15])

    if include_compras:
        comp = data["compras"]
        rows = [["Total OCs", comp["total_ocs"]], ["Gasto recibido (S/)", fmt_num(comp["gasto_recibido"])]]
        for status, count in comp["status_counts"].items():
            rows.append([f"OC — {status}", count])
        add_sheet("Compras", "KPIs de Compras", ["Indicador", "Valor"], rows, [35, 18])

    if include_admin:
        adm = data["admin"]
        rows = [
            ["Requerimientos activos", adm["active_requirements_count"]],
            ["Costo total requerimientos (S/)", fmt_num(adm["total_requirements_cost"])],
        ]
        for c in adm["costos_por_categoria"]:
            rows.append([f"Costo — {c['categoria']}", fmt_num(c["total"])])
        rows.append(["", ""])
        for estado, count in adm.get("planificacion_counts", {}).items():
            rows.append([f"Planificación — {estado}", count])
        rows.append(["", ""])
        for p in adm.get("productividad_por_persona", []):
            rows.append([p["username"], f"{p['horas_totales']} h ({p['registros']} reg.)"])
        add_sheet("Administración", "KPIs de Administración", ["Indicador / Persona", "Valor"], rows, [35, 22])

    if include_weak_points:
        wp = data["weak_points"]
        rows = [
            ["Alerta SLA bajo (<80%)", "SÍ" if wp["low_sla_alert"] else "NO"],
            ["", ""],
            ["Personal sobrecargado (>=5 ítems)", ""],
        ]
        for u in wp["overloaded_users"]:
            rows.append([u["username"], f"{u['total_active']} abiertas"])
        add_sheet("Diagnóstico", "Puntos Débiles", ["Alerta / Usuario", "Detalle"], rows, [32, 22])

    if not wb.sheetnames:
        ws = wb.create_sheet("Vacío")
        ws["A1"] = "Sin secciones seleccionadas"

    return excel_response(wb, f"reporte_kpis_{datetime.now().strftime('%Y%m%d')}.xlsx")