from datetime import datetime
from fastapi import HTTPException
from fastapi.responses import StreamingResponse
from app.core.database import db_connection
from app.core.utils import generate_sequential_code
import io

def _get_valid_tipos(conn) -> tuple:
    with conn.cursor() as cur:
        cur.execute("SELECT codigo FROM categorias_costo WHERE activo = TRUE ORDER BY orden")
        return tuple(r[0] for r in cur.fetchall())

VALID_STATUS = ("BORRADOR", "ENVIADA", "APROBADA", "RECHAZADA", "EXPIRADA")

VALID_TRANSITIONS = {
    "BORRADOR":  ["ENVIADA"],
    "ENVIADA":   ["APROBADA", "RECHAZADA", "EXPIRADA"],
    "APROBADA":  [],
    "RECHAZADA": ["BORRADOR"],
    "EXPIRADA":  ["BORRADOR"],
}


# ══════════════════════════════════════════════════════════════════════════════
# CATEGORÍAS DE COSTO
# ══════════════════════════════════════════════════════════════════════════════

def list_categorias_costo_service(solo_activas: bool = True) -> list:
    with db_connection() as conn:
        with conn.cursor() as cur:
            if solo_activas:
                cur.execute("""
                    SELECT id, codigo, nombre, es_directo, orden, color_hex, activo, created_at
                    FROM categorias_costo WHERE activo = TRUE ORDER BY orden
                """)
            else:
                cur.execute("""
                    SELECT id, codigo, nombre, es_directo, orden, color_hex, activo, created_at
                    FROM categorias_costo ORDER BY orden
                """)
            return [
                {
                    "id": str(r[0]), "codigo": r[1], "nombre": r[2],
                    "es_directo": r[3], "orden": r[4], "color_hex": r[5],
                    "activo": r[6], "created_at": r[7],
                }
                for r in cur.fetchall()
            ]


def create_categoria_costo_service(payload) -> dict:
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                INSERT INTO categorias_costo (codigo, nombre, es_directo, orden, color_hex)
                VALUES (%s, %s, %s, %s, %s)
                RETURNING id, codigo, nombre, es_directo, orden, color_hex, activo, created_at
            """, (payload.codigo.upper(), payload.nombre, payload.es_directo,
                  payload.orden, payload.color_hex))
            r = cur.fetchone()
            conn.commit()
    return {
        "id": str(r[0]), "codigo": r[1], "nombre": r[2],
        "es_directo": r[3], "orden": r[4], "color_hex": r[5],
        "activo": r[6], "created_at": r[7],
    }


def update_categoria_costo_service(categoria_id: str, payload) -> dict:
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT id FROM categorias_costo WHERE id = %s", (categoria_id,))
            if not cur.fetchone():
                raise HTTPException(404, "Categoría no encontrada")
            fields = {k: v for k, v in payload.model_dump().items() if v is not None}
            if not fields:
                raise HTTPException(400, "No hay campos para actualizar")
            set_clause = ", ".join(f"{k} = %s" for k in fields)
            cur.execute(
                f"UPDATE categorias_costo SET {set_clause} WHERE id = %s "
                "RETURNING id, codigo, nombre, es_directo, orden, color_hex, activo, created_at",
                (*fields.values(), categoria_id)
            )
            r = cur.fetchone()
            conn.commit()
    return {
        "id": str(r[0]), "codigo": r[1], "nombre": r[2],
        "es_directo": r[3], "orden": r[4], "color_hex": r[5],
        "activo": r[6], "created_at": r[7],
    }


# ══════════════════════════════════════════════════════════════════════════════
# RECURSOS DE MANO DE OBRA
# ══════════════════════════════════════════════════════════════════════════════

def _gen_mo_code(conn) -> str:
    return generate_sequential_code(conn, "recursos_mo", "MO", year_based=False, pad=3)


def list_recursos_mo_service() -> list:
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT id, codigo, descripcion, categoria, tarifa_hora, unidad, activo, created_at
                FROM recursos_mo
                ORDER BY codigo
            """)
            rows = cur.fetchall()
    return [
        {
            "id": str(r[0]), "codigo": r[1], "descripcion": r[2],
            "categoria": r[3], "tarifa_hora": float(r[4]),
            "unidad": r[5], "activo": r[6], "created_at": r[7],
        }
        for r in rows
    ]


def create_recurso_mo_service(payload) -> dict:
    with db_connection() as conn:
        codigo = payload.codigo.strip().upper()
        with conn.cursor() as cur:
            cur.execute("SELECT id FROM recursos_mo WHERE codigo = %s", (codigo,))
            if cur.fetchone():
                raise HTTPException(400, f"Ya existe un recurso con código {codigo}")
            cur.execute("""
                INSERT INTO recursos_mo (codigo, descripcion, categoria, tarifa_hora, unidad)
                VALUES (%s, %s, %s, %s, %s)
                RETURNING id, codigo, descripcion, categoria, tarifa_hora, unidad, activo, created_at
            """, (codigo, payload.descripcion.strip(), payload.categoria, payload.tarifa_hora, payload.unidad))
            row = cur.fetchone()
            conn.commit()
    return {
        "id": str(row[0]), "codigo": row[1], "descripcion": row[2],
        "categoria": row[3], "tarifa_hora": float(row[4]),
        "unidad": row[5], "activo": row[6], "created_at": row[7],
    }


def update_recurso_mo_service(recurso_id: str, payload) -> dict:
    fields, vals = [], []
    if payload.descripcion is not None:
        fields.append("descripcion = %s"); vals.append(payload.descripcion.strip())
    if payload.categoria is not None:
        fields.append("categoria = %s"); vals.append(payload.categoria)
    if payload.tarifa_hora is not None:
        fields.append("tarifa_hora = %s"); vals.append(payload.tarifa_hora)
    if payload.unidad is not None:
        fields.append("unidad = %s"); vals.append(payload.unidad)
    if payload.activo is not None:
        fields.append("activo = %s"); vals.append(payload.activo)
    if not fields:
        raise HTTPException(400, "Nada que actualizar")
    vals.append(recurso_id)
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                f"UPDATE recursos_mo SET {', '.join(fields)} WHERE id = %s "
                "RETURNING id, codigo, descripcion, categoria, tarifa_hora, unidad, activo",
                vals,
            )
            row = cur.fetchone()
            if not row:
                raise HTTPException(404, "Recurso MO no encontrado")
            conn.commit()
    return {
        "id": str(row[0]), "codigo": row[1], "descripcion": row[2],
        "categoria": row[3], "tarifa_hora": float(row[4]),
        "unidad": row[5], "activo": row[6],
    }


def delete_recurso_mo_service(recurso_id: str) -> dict:
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT COUNT(*) FROM presupuesto_apu_items WHERE recurso_mo_id = %s",
                (recurso_id,)
            )
            if cur.fetchone()[0] > 0:
                raise HTTPException(400, "No se puede eliminar: el recurso está en uso en partidas APU")
            cur.execute("DELETE FROM recursos_mo WHERE id = %s RETURNING id", (recurso_id,))
            if not cur.fetchone():
                raise HTTPException(404, "Recurso MO no encontrado")
            conn.commit()
    return {"ok": True}


# ══════════════════════════════════════════════════════════════════════════════
# PARTIDAS DEL PRESUPUESTO
# ══════════════════════════════════════════════════════════════════════════════

def _apu_subtotal(conn, partida_id: str) -> float:
    with conn.cursor() as cur:
        cur.execute("""
            SELECT COALESCE(SUM(cantidad * precio_unitario), 0)
            FROM presupuesto_apu_items
            WHERE partida_id = %s
        """, (partida_id,))
        return float(cur.fetchone()[0])


def _row_to_partida(r, apu_items=None) -> dict:
    return {
        "id": str(r[0]), "plan_id": str(r[1]),
        "codigo": r[2], "descripcion": r[3],
        "unidad": r[4], "cantidad": float(r[5]),
        "orden": r[6], "es_capitulo": r[7],
        "parent_id": str(r[8]) if r[8] else None,
        "precio_unitario_apu": float(r[9]) if r[9] is not None else 0.0,
        "subtotal": float(r[5]) * (float(r[9]) if r[9] else 0.0),
        "apu_items": apu_items or [],
    }


def list_partidas_service(plan_id: str) -> list:
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT id FROM project_plans WHERE id = %s", (plan_id,))
            if not cur.fetchone():
                raise HTTPException(404, "Plan no encontrado")
            cur.execute("""
                SELECT pp.id, pp.plan_id, pp.codigo, pp.descripcion,
                       pp.unidad, pp.cantidad, pp.orden, pp.es_capitulo, pp.parent_id,
                       COALESCE((
                           SELECT SUM(ai.cantidad * ai.precio_unitario)
                           FROM presupuesto_apu_items ai WHERE ai.partida_id = pp.id
                       ), 0) AS precio_unitario_apu
                FROM presupuesto_partidas pp
                WHERE pp.plan_id = %s
                ORDER BY pp.orden, pp.codigo
            """, (plan_id,))
            rows = cur.fetchall()
    return [_row_to_partida(r) for r in rows]


def create_partida_service(plan_id: str, payload) -> dict:
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT id FROM project_plans WHERE id = %s", (plan_id,))
            if not cur.fetchone():
                raise HTTPException(404, "Plan no encontrado")
            if payload.parent_id:
                cur.execute(
                    "SELECT id FROM presupuesto_partidas WHERE id = %s AND plan_id = %s",
                    (payload.parent_id, plan_id)
                )
                if not cur.fetchone():
                    raise HTTPException(400, "parent_id no existe en este plan")
            cur.execute("""
                INSERT INTO presupuesto_partidas
                    (plan_id, codigo, descripcion, unidad, cantidad, orden, es_capitulo, parent_id)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id, plan_id, codigo, descripcion, unidad, cantidad, orden, es_capitulo, parent_id, NULL
            """, (
                plan_id, payload.codigo.strip(), payload.descripcion.strip(),
                payload.unidad, payload.cantidad, payload.orden,
                payload.es_capitulo, payload.parent_id,
            ))
            row = cur.fetchone()
            conn.commit()
    return _row_to_partida(row)


def update_partida_service(plan_id: str, partida_id: str, payload) -> dict:
    fields, vals = [], []
    if payload.codigo is not None:
        fields.append("codigo = %s"); vals.append(payload.codigo.strip())
    if payload.descripcion is not None:
        fields.append("descripcion = %s"); vals.append(payload.descripcion.strip())
    if payload.unidad is not None:
        fields.append("unidad = %s"); vals.append(payload.unidad)
    if payload.cantidad is not None:
        fields.append("cantidad = %s"); vals.append(payload.cantidad)
    if payload.orden is not None:
        fields.append("orden = %s"); vals.append(payload.orden)
    if not fields:
        raise HTTPException(400, "Nada que actualizar")
    fields.append("updated_at = NOW()")
    vals += [partida_id, plan_id]
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                f"UPDATE presupuesto_partidas SET {', '.join(fields)} "
                "WHERE id = %s AND plan_id = %s "
                "RETURNING id, plan_id, codigo, descripcion, unidad, cantidad, orden, es_capitulo, parent_id",
                vals,
            )
            row = cur.fetchone()
            if not row:
                raise HTTPException(404, "Partida no encontrada")
            pu = _apu_subtotal(conn, str(row[0]))
            conn.commit()
    return _row_to_partida((*row, pu))


def delete_partida_service(plan_id: str, partida_id: str) -> dict:
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "DELETE FROM presupuesto_partidas WHERE id = %s AND plan_id = %s RETURNING id",
                (partida_id, plan_id)
            )
            if not cur.fetchone():
                raise HTTPException(404, "Partida no encontrada")
            conn.commit()
    return {"ok": True}


# ══════════════════════════════════════════════════════════════════════════════
# ÍTEMS APU
# ══════════════════════════════════════════════════════════════════════════════

def _row_to_apu(r) -> dict:
    return {
        "id": str(r[0]), "partida_id": str(r[1]),
        "tipo_recurso": r[2],
        "material_id": str(r[3]) if r[3] else None,
        "recurso_mo_id": str(r[4]) if r[4] else None,
        "descripcion": r[5], "unidad": r[6],
        "cantidad": float(r[7]), "precio_unitario": float(r[8]),
        "subtotal": float(r[7]) * float(r[8]),
        "material_nombre": r[9],
        "recurso_mo_codigo": r[10],
    }


def list_apu_service(partida_id: str) -> list:
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT ai.id, ai.partida_id, ai.tipo_recurso,
                       ai.material_id, ai.recurso_mo_id, ai.descripcion, ai.unidad,
                       ai.cantidad, ai.precio_unitario,
                       m.name AS material_nombre,
                       rmo.codigo AS recurso_mo_codigo
                FROM presupuesto_apu_items ai
                LEFT JOIN materials m ON m.id = ai.material_id
                LEFT JOIN recursos_mo rmo ON rmo.id = ai.recurso_mo_id
                WHERE ai.partida_id = %s
                ORDER BY ai.tipo_recurso, ai.created_at
            """, (partida_id,))
            rows = cur.fetchall()
    return [_row_to_apu(r) for r in rows]


def create_apu_item_service(partida_id: str, payload) -> dict:
    with db_connection() as conn:
        valid_tipos = _get_valid_tipos(conn)
        if payload.tipo_recurso not in valid_tipos:
            raise HTTPException(400, f"tipo_recurso debe ser uno de: {', '.join(valid_tipos)}")

        with conn.cursor() as cur:
            cur.execute("SELECT id FROM presupuesto_partidas WHERE id = %s", (partida_id,))
            if not cur.fetchone():
                raise HTTPException(404, "Partida no encontrada")

            mat_nombre, mo_codigo = None, None

            if payload.tipo_recurso != "MO" and payload.material_id:
                cur.execute("SELECT name FROM materials WHERE id = %s", (payload.material_id,))
                row = cur.fetchone()
                if not row:
                    raise HTTPException(400, "material_id no existe")
                mat_nombre = row[0]

            if payload.tipo_recurso == "MO" and payload.recurso_mo_id:
                cur.execute("SELECT codigo FROM recursos_mo WHERE id = %s", (payload.recurso_mo_id,))
                row = cur.fetchone()
                if not row:
                    raise HTTPException(400, "recurso_mo_id no existe")
                mo_codigo = row[0]

            cur.execute("""
                INSERT INTO presupuesto_apu_items
                    (partida_id, tipo_recurso, material_id, recurso_mo_id,
                     descripcion, unidad, cantidad, precio_unitario)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id, partida_id, tipo_recurso, material_id, recurso_mo_id,
                          descripcion, unidad, cantidad, precio_unitario
            """, (
                partida_id, payload.tipo_recurso,
                payload.material_id, payload.recurso_mo_id,
                payload.descripcion, payload.unidad,
                payload.cantidad, payload.precio_unitario,
            ))
            row = cur.fetchone()
            conn.commit()

    return _row_to_apu((*row, mat_nombre, mo_codigo))


def update_apu_item_service(partida_id: str, apu_id: str, payload) -> dict:
    fields, vals = [], []
    if payload.cantidad is not None:
        fields.append("cantidad = %s"); vals.append(payload.cantidad)
    if payload.precio_unitario is not None:
        fields.append("precio_unitario = %s"); vals.append(payload.precio_unitario)
    if payload.descripcion is not None:
        fields.append("descripcion = %s"); vals.append(payload.descripcion)
    if not fields:
        raise HTTPException(400, "Nada que actualizar")
    vals += [apu_id, partida_id]
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                f"UPDATE presupuesto_apu_items SET {', '.join(fields)} "
                "WHERE id = %s AND partida_id = %s "
                "RETURNING id, partida_id, tipo_recurso, material_id, recurso_mo_id, "
                "descripcion, unidad, cantidad, precio_unitario",
                vals,
            )
            row = cur.fetchone()
            if not row:
                raise HTTPException(404, "Ítem APU no encontrado")
            cur.execute(
                "SELECT m.name FROM materials m WHERE m.id = %s", (row[3],)
            ) if row[3] else None
            mat_nombre = cur.fetchone()[0] if row[3] else None
            cur.execute(
                "SELECT rmo.codigo FROM recursos_mo rmo WHERE rmo.id = %s", (row[4],)
            ) if row[4] else None
            mo_codigo = cur.fetchone()[0] if row[4] else None
            conn.commit()
    return _row_to_apu((*row, mat_nombre, mo_codigo))


def delete_apu_item_service(partida_id: str, apu_id: str) -> dict:
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "DELETE FROM presupuesto_apu_items WHERE id = %s AND partida_id = %s RETURNING id",
                (apu_id, partida_id)
            )
            if not cur.fetchone():
                raise HTTPException(404, "Ítem APU no encontrado")
            conn.commit()
    return {"ok": True}


# ══════════════════════════════════════════════════════════════════════════════
# CONFIGURACIÓN Y RESUMEN
# ══════════════════════════════════════════════════════════════════════════════

def _ensure_config(conn, plan_id: str):
    with conn.cursor() as cur:
        cur.execute("SELECT id FROM presupuesto_config WHERE plan_id = %s", (plan_id,))
        if not cur.fetchone():
            cur.execute(
                "INSERT INTO presupuesto_config (plan_id) VALUES (%s)", (plan_id,)
            )


def _row_to_config(r) -> dict:
    if not r:
        return {}
    return {
        "id": str(r[0]), "plan_id": str(r[1]),
        "gastos_generales_pct": float(r[2]), "utilidad_pct": float(r[3]),
        "igv_pct": float(r[4]), "moneda": r[5],
        "cliente_nombre": r[6], "cliente_ruc": r[7],
        "lugar_trabajo": r[8], "plazo_dias": r[9],
        "validez_dias": r[10], "notas": r[11],
        "cliente_id": str(r[12]) if r[12] else None,
        "numero_cotizacion": r[13],
        "status": r[14] or "BORRADOR",
        "fecha_envio": r[15].isoformat() if r[15] else None,
        "fecha_respuesta": r[16].isoformat() if r[16] else None,
        "notas_comerciales": r[17],
        "contacto_id": str(r[18]) if r[18] else None,
    }


_CONFIG_SELECT = """
    SELECT id, plan_id, gastos_generales_pct, utilidad_pct, igv_pct, moneda,
           cliente_nombre, cliente_ruc, lugar_trabajo, plazo_dias, validez_dias, notas,
           cliente_id, numero_cotizacion, status, fecha_envio, fecha_respuesta, notas_comerciales,
           contacto_id
    FROM presupuesto_config
"""


def get_config_service(plan_id: str) -> dict:
    with db_connection() as conn:
        _ensure_config(conn, plan_id)
        with conn.cursor() as cur:
            cur.execute(_CONFIG_SELECT + " WHERE plan_id = %s", (plan_id,))
            row = cur.fetchone()
            conn.commit()
    return _row_to_config(row)


def update_config_service(plan_id: str, payload) -> dict:
    with db_connection() as conn:
        _ensure_config(conn, plan_id)
        fields, vals = [], []
        mapping = [
            ("gastos_generales_pct", payload.gastos_generales_pct),
            ("utilidad_pct", payload.utilidad_pct),
            ("igv_pct", payload.igv_pct),
            ("moneda", payload.moneda),
            ("cliente_id", payload.cliente_id),
            ("cliente_nombre", payload.cliente_nombre),
            ("cliente_ruc", payload.cliente_ruc),
            ("contacto_id", payload.contacto_id),
            ("lugar_trabajo", payload.lugar_trabajo),
            ("plazo_dias", payload.plazo_dias),
            ("validez_dias", payload.validez_dias),
            ("notas", payload.notas),
            ("notas_comerciales", payload.notas_comerciales),
        ]
        for col, val in mapping:
            if val is not None:
                if col in ("cliente_id", "contacto_id") and val == "":
                    val = None
                fields.append(f"{col} = %s"); vals.append(val)
        if not fields:
            raise HTTPException(400, "Nada que actualizar")
        fields.append("updated_at = NOW()")
        vals.append(plan_id)
        with conn.cursor() as cur:
            cur.execute(
                f"UPDATE presupuesto_config SET {', '.join(fields)} WHERE plan_id = %s "
                "RETURNING id, plan_id, gastos_generales_pct, utilidad_pct, igv_pct, moneda, "
                "cliente_nombre, cliente_ruc, lugar_trabajo, plazo_dias, validez_dias, notas, "
                "cliente_id, numero_cotizacion, status, fecha_envio, fecha_respuesta, notas_comerciales, "
                "contacto_id",
                vals,
            )
            row = cur.fetchone()
            conn.commit()
    return _row_to_config(row)


# ══════════════════════════════════════════════════════════════════════════════
# CICLO DE ESTADOS DE COTIZACIÓN
# ══════════════════════════════════════════════════════════════════════════════

def _gen_cot_number(conn) -> str:
    year = datetime.now().year
    with conn.cursor() as cur:
        cur.execute(
            "SELECT COUNT(*) FROM presupuesto_config WHERE numero_cotizacion LIKE %s",
            (f"COT-{year}-%",)
        )
        seq = cur.fetchone()[0] + 1
    return f"COT-{year}-{seq:04d}"


def update_cotizacion_status_service(plan_id: str, nuevo_status: str) -> dict:
    if nuevo_status not in VALID_STATUS:
        raise HTTPException(400, f"Estado inválido. Válidos: {', '.join(VALID_STATUS)}")

    with db_connection() as conn:
        _ensure_config(conn, plan_id)
        with conn.cursor() as cur:
            cur.execute(
                "SELECT status, numero_cotizacion FROM presupuesto_config WHERE plan_id = %s",
                (plan_id,)
            )
            row = cur.fetchone()
            if not row:
                raise HTTPException(404, "Configuración no encontrada")
            estado_actual = row[0] or "BORRADOR"
            numero_actual = row[1]

        allowed = VALID_TRANSITIONS.get(estado_actual, [])
        if nuevo_status not in allowed:
            raise HTTPException(
                400,
                f"Transición inválida: {estado_actual} → {nuevo_status}. "
                f"Desde {estado_actual} solo se puede ir a: {allowed or 'ninguno'}"
            )

        extra_fields = []
        extra_vals = []

        if nuevo_status == "ENVIADA":
            if not numero_actual:
                numero_actual = _gen_cot_number(conn)
                extra_fields.append("numero_cotizacion = %s")
                extra_vals.append(numero_actual)
            extra_fields.append("fecha_envio = NOW()")

        if nuevo_status in ("APROBADA", "RECHAZADA", "EXPIRADA"):
            extra_fields.append("fecha_respuesta = NOW()")

        if nuevo_status == "BORRADOR":
            extra_fields.append("fecha_envio = NULL")
            extra_fields.append("fecha_respuesta = NULL")

        set_clause = "status = %s, updated_at = NOW()"
        set_vals = [nuevo_status]
        if extra_fields:
            set_clause += ", " + ", ".join(extra_fields)
            set_vals.extend(extra_vals)
        set_vals.append(plan_id)

        with conn.cursor() as cur:
            cur.execute(
                f"UPDATE presupuesto_config SET {set_clause} WHERE plan_id = %s "
                "RETURNING id, plan_id, gastos_generales_pct, utilidad_pct, igv_pct, moneda, "
                "cliente_nombre, cliente_ruc, lugar_trabajo, plazo_dias, validez_dias, notas, "
                "cliente_id, numero_cotizacion, status, fecha_envio, fecha_respuesta, notas_comerciales",
                set_vals,
            )
            row = cur.fetchone()
            conn.commit()
    return _row_to_config(row)


def list_cotizaciones_service(status: str = None, cliente_id: str = None) -> list:
    """Lista todas las cotizaciones con datos del plan y cliente."""
    with db_connection() as conn:
        with conn.cursor() as cur:
            q = """
                SELECT
                    pc.id, pc.plan_id, pc.numero_cotizacion, pc.status,
                    pc.fecha_envio, pc.fecha_respuesta, pc.moneda,
                    pc.gastos_generales_pct, pc.utilidad_pct, pc.igv_pct,
                    pc.lugar_trabajo, pc.plazo_dias, pc.validez_dias,
                    pc.cliente_id, pc.cliente_nombre, pc.cliente_ruc,
                    pc.notas_comerciales, pc.created_at,
                    pp.project_code AS plan_code, pp.title AS plan_title,
                    c.razon_social AS cliente_razon_social,
                    c.ruc AS cliente_ruc_reg,
                    c.contacto AS cliente_contacto
                FROM presupuesto_config pc
                JOIN project_plans pp ON pp.id = pc.plan_id
                LEFT JOIN clientes c ON c.id = pc.cliente_id
                WHERE 1=1
            """
            params = []
            if status:
                q += " AND pc.status = %s"
                params.append(status)
            if cliente_id:
                q += " AND pc.cliente_id = %s"
                params.append(cliente_id)
            q += " ORDER BY pc.created_at DESC"
            cur.execute(q, params)
            rows = cur.fetchall()

    result = []
    for r in rows:
        result.append({
            "id": str(r[0]),
            "plan_id": str(r[1]),
            "numero_cotizacion": r[2],
            "status": r[3] or "BORRADOR",
            "fecha_envio": r[4].isoformat() if r[4] else None,
            "fecha_respuesta": r[5].isoformat() if r[5] else None,
            "moneda": r[6],
            "gastos_generales_pct": float(r[7]),
            "utilidad_pct": float(r[8]),
            "igv_pct": float(r[9]),
            "lugar_trabajo": r[10],
            "plazo_dias": r[11],
            "validez_dias": r[12],
            "cliente_id": str(r[13]) if r[13] else None,
            "cliente_nombre": r[14],
            "cliente_ruc": r[15],
            "notas_comerciales": r[16],
            "created_at": r[17].isoformat() if r[17] else None,
            "plan_code": r[18],
            "plan_title": r[19],
            "cliente_razon_social": r[20],
            "cliente_ruc_reg": r[21],
            "cliente_contacto": r[22],
        })
    return result


def get_cotizaciones_stats_service() -> dict:
    """KPIs de cotizaciones: totales por estado y tasa de conversión."""
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT
                    COUNT(*) FILTER (WHERE status = 'BORRADOR')   AS borrador,
                    COUNT(*) FILTER (WHERE status = 'ENVIADA')    AS enviada,
                    COUNT(*) FILTER (WHERE status = 'APROBADA')   AS aprobada,
                    COUNT(*) FILTER (WHERE status = 'RECHAZADA')  AS rechazada,
                    COUNT(*) FILTER (WHERE status = 'EXPIRADA')   AS expirada,
                    COUNT(*)                                       AS total
                FROM presupuesto_config
            """)
            r = cur.fetchone()
    borrador, enviada, aprobada, rechazada, expirada, total = (int(x) for x in r)
    respondidas = aprobada + rechazada
    tasa = round((aprobada / respondidas) * 100, 1) if respondidas > 0 else 0
    return {
        "borrador": borrador,
        "enviada": enviada,
        "aprobada": aprobada,
        "rechazada": rechazada,
        "expirada": expirada,
        "total": total,
        "tasa_conversion": tasa,
    }


def _compute_resumen(conn, plan_id: str) -> dict:
    # Breakdown de costos por categoría (directo e indirecto)
    with conn.cursor() as cur:
        cur.execute("""
            SELECT
                COALESCE(cc.codigo, ai.tipo_recurso)  AS codigo,
                COALESCE(cc.nombre, ai.tipo_recurso)  AS nombre,
                COALESCE(cc.es_directo, TRUE)          AS es_directo,
                COALESCE(cc.orden, 99)                 AS orden,
                COALESCE(cc.color_hex, '#4F7C82')      AS color_hex,
                SUM(pp.cantidad * ai.cantidad * ai.precio_unitario) AS costo
            FROM presupuesto_partidas pp
            JOIN presupuesto_apu_items ai ON ai.partida_id = pp.id
            LEFT JOIN categorias_costo cc ON cc.codigo = ai.tipo_recurso
            WHERE pp.plan_id = %s AND pp.es_capitulo = FALSE
            GROUP BY cc.codigo, ai.tipo_recurso, cc.nombre, cc.es_directo, cc.orden, cc.color_hex
            ORDER BY COALESCE(cc.orden, 99)
        """, (plan_id,))
        categorias_rows = cur.fetchall()

    breakdown = [
        {
            "codigo": r[0], "nombre": r[1], "es_directo": r[2],
            "orden": r[3], "color_hex": r[4], "costo": round(float(r[5]), 2),
        }
        for r in categorias_rows
    ]
    costo_directo = sum(b["costo"] for b in breakdown if b["es_directo"])
    costos_indirectos = [b for b in breakdown if not b["es_directo"]]

    _ensure_config(conn, plan_id)
    with conn.cursor() as cur:
        cur.execute(
            "SELECT gastos_generales_pct, utilidad_pct, igv_pct FROM presupuesto_config WHERE plan_id = %s",
            (plan_id,)
        )
        cfg = cur.fetchone()
        conn.commit()

    gg_pct = float(cfg[0]) / 100
    ut_pct = float(cfg[1]) / 100
    igv_pct = float(cfg[2]) / 100

    sum_indirectos = sum(b["costo"] for b in breakdown if not b["es_directo"])
    gastos_generales = round(costo_directo * gg_pct, 2)
    sub_total = round(costo_directo + gastos_generales, 2)
    utilidad = round(sub_total * ut_pct, 2)
    valor_venta = round(sub_total + utilidad + sum_indirectos, 2)
    igv = round(valor_venta * igv_pct, 2)
    precio_total = round(valor_venta + igv, 2)

    return {
        "costo_directo": round(costo_directo, 2),
        "breakdown_categorias": breakdown,
        "costos_indirectos": costos_indirectos,
        "gastos_generales": gastos_generales,
        "gastos_generales_pct": float(cfg[0]),
        "sub_total": sub_total,
        "utilidad": utilidad,
        "utilidad_pct": float(cfg[1]),
        "valor_venta": valor_venta,
        "igv": igv,
        "igv_pct": float(cfg[2]),
        "precio_total": precio_total,
    }


def get_resumen_service(plan_id: str) -> dict:
    with db_connection() as conn:
        return _compute_resumen(conn, plan_id)


# ══════════════════════════════════════════════════════════════════════════════
# EXPORTACIÓN PDF
# ══════════════════════════════════════════════════════════════════════════════

def export_pdf_service(plan_id: str) -> StreamingResponse:
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
        )
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    except ImportError:
        raise HTTPException(500, "reportlab no está instalado. Ejecute: pip install reportlab")

    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT project_code, title FROM project_plans WHERE id = %s", (plan_id,)
            )
            plan_row = cur.fetchone()
            if not plan_row:
                raise HTTPException(404, "Plan no encontrado")
            plan_code, plan_title = plan_row
            plan_code = plan_code or f"PLAN_{str(plan_id)[:8]}"

        _ensure_config(conn, plan_id)
        with conn.cursor() as cur:
            cur.execute("""
                SELECT cliente_nombre, cliente_ruc, lugar_trabajo, plazo_dias,
                       validez_dias, gastos_generales_pct, utilidad_pct, igv_pct, moneda, notas
                FROM presupuesto_config WHERE plan_id = %s
            """, (plan_id,))
            cfg = cur.fetchone()
            conn.commit()

        with conn.cursor() as cur:
            cur.execute("""
                SELECT pp.id, pp.codigo, pp.descripcion, pp.unidad, pp.cantidad, pp.es_capitulo,
                       COALESCE((
                           SELECT SUM(ai.cantidad * ai.precio_unitario)
                           FROM presupuesto_apu_items ai WHERE ai.partida_id = pp.id
                       ), 0) AS precio_apu
                FROM presupuesto_partidas pp
                WHERE pp.plan_id = %s
                ORDER BY pp.orden, pp.codigo
            """, (plan_id,))
            partidas = cur.fetchall()

    resumen = _compute_resumen_raw(cfg)
    fecha = datetime.now().strftime("%d/%m/%Y")

    PRIMARY = colors.HexColor("#0B2E33")
    ACCENT = colors.HexColor("#B8E3E9")
    LIGHT = colors.HexColor("#F0F7F8")
    WHITE = colors.white
    moneda = cfg[8] or "S/"

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=1.5 * cm, rightMargin=1.5 * cm,
        topMargin=1.5 * cm, bottomMargin=1.5 * cm,
    )
    styles = getSampleStyleSheet()
    bold_white = ParagraphStyle("bw", parent=styles["Normal"], textColor=WHITE, fontName="Helvetica-Bold", fontSize=9)
    bold_dark = ParagraphStyle("bd", parent=styles["Normal"], textColor=PRIMARY, fontName="Helvetica-Bold", fontSize=8)
    normal = ParagraphStyle("n", parent=styles["Normal"], fontSize=8, leading=10)
    cap_style = ParagraphStyle("cap", parent=styles["Normal"], textColor=WHITE, fontName="Helvetica-Bold", fontSize=8)
    right_style = ParagraphStyle("r", parent=styles["Normal"], fontSize=8, alignment=TA_RIGHT)

    story = []

    # ── Cabecera ──────────────────────────────────────────────────────────────
    header_data = [[
        Paragraph("<b>CeShark ERP</b><br/><font size=7>Presupuesto / Cotización</font>", bold_white),
        Paragraph(f"<b>COT — {plan_code}</b><br/><font size=7>Fecha: {fecha}</font>", bold_white),
    ]]
    header_table = Table(header_data, colWidths=["60%", "40%"])
    header_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), PRIMARY),
        ("TEXTCOLOR", (0, 0), (-1, -1), WHITE),
        ("PADDING", (0, 0), (-1, -1), 10),
        ("ALIGN", (1, 0), (1, 0), "RIGHT"),
    ]))
    story.append(header_table)
    story.append(Spacer(1, 0.3 * cm))

    # ── Datos del cliente ─────────────────────────────────────────────────────
    cliente = cfg[0] or "—"
    ruc = cfg[1] or "—"
    lugar = cfg[2] or "—"
    plazo = f"{cfg[3]} días" if cfg[3] else "—"
    validez = f"{cfg[4]} días" if cfg[4] else "30 días"

    client_data = [
        [Paragraph(f"<b>Proyecto:</b> {plan_title}", normal),
         Paragraph(f"<b>Lugar:</b> {lugar}", normal)],
        [Paragraph(f"<b>Cliente:</b> {cliente} &nbsp; RUC: {ruc}", normal),
         Paragraph(f"<b>Plazo:</b> {plazo} &nbsp;&nbsp; <b>Validez:</b> {validez}", normal)],
    ]
    client_table = Table(client_data, colWidths=["55%", "45%"])
    client_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), LIGHT),
        ("PADDING", (0, 0), (-1, -1), 5),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#C5D8DB")),
    ]))
    story.append(client_table)
    story.append(Spacer(1, 0.4 * cm))

    # ── Tabla de partidas ─────────────────────────────────────────────────────
    col_w = [1.5 * cm, 8.5 * cm, 1.5 * cm, 2 * cm, 2.5 * cm, 2.5 * cm]
    table_data = [[
        Paragraph("<b>Ítem</b>", bold_white),
        Paragraph("<b>Descripción</b>", bold_white),
        Paragraph("<b>Und</b>", bold_white),
        Paragraph("<b>Cant.</b>", bold_white),
        Paragraph("<b>P.U. (S/)</b>", bold_white),
        Paragraph("<b>Parcial (S/)</b>", bold_white),
    ]]
    row_styles = []
    for i, p in enumerate(partidas, start=1):
        pid, cod, desc, und, cant, es_cap, precio_apu = p
        cant_f = float(cant)
        pu = float(precio_apu)
        parcial = cant_f * pu

        if es_cap:
            table_data.append([
                Paragraph(f"<b>{cod}</b>", cap_style),
                Paragraph(f"<b>{desc.upper()}</b>", cap_style),
                "", "", "", "",
            ])
            row_styles.append(("BACKGROUND", (0, i), (-1, i), PRIMARY))
            row_styles.append(("SPAN", (1, i), (5, i)))
        else:
            table_data.append([
                Paragraph(cod, normal),
                Paragraph(desc, normal),
                Paragraph(und, normal),
                Paragraph(f"{cant_f:,.3f}", right_style),
                Paragraph(f"{moneda} {pu:,.2f}", right_style),
                Paragraph(f"{moneda} {parcial:,.2f}", right_style),
            ])

    partidas_table = Table(table_data, colWidths=col_w, repeatRows=1)
    base_style = [
        ("BACKGROUND", (0, 0), (-1, 0), PRIMARY),
        ("TEXTCOLOR", (0, 0), (-1, 0), WHITE),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("PADDING", (0, 0), (-1, -1), 4),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [WHITE, LIGHT]),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D1DEE0")),
        ("ALIGN", (3, 1), (-1, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]
    partidas_table.setStyle(TableStyle(base_style + row_styles))
    story.append(partidas_table)
    story.append(Spacer(1, 0.5 * cm))

    # ── Resumen económico ─────────────────────────────────────────────────────
    r = resumen
    summary_data = [
        ["COSTO DIRECTO", f"{moneda} {r['costo_directo']:,.2f}"],
        [f"GASTOS GENERALES ({r['gg_pct']:.1f}%)", f"{moneda} {r['gastos_generales']:,.2f}"],
        [f"UTILIDAD ({r['ut_pct']:.1f}%)", f"{moneda} {r['utilidad']:,.2f}"],
        ["VALOR VENTA", f"{moneda} {r['valor_venta']:,.2f}"],
        [f"IGV ({r['igv_pct']:.1f}%)", f"{moneda} {r['igv']:,.2f}"],
        ["PRECIO TOTAL", f"{moneda} {r['precio_total']:,.2f}"],
    ]
    summary_table = Table(summary_data, colWidths=["70%", "30%"])
    summary_style = [
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("PADDING", (0, 0), (-1, -1), 5),
        ("ROWBACKGROUNDS", (0, 0), (-1, -1), [WHITE, LIGHT]),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D1DEE0")),
        ("FONTNAME", (0, 5), (-1, 5), "Helvetica-Bold"),
        ("BACKGROUND", (0, 5), (-1, 5), PRIMARY),
        ("TEXTCOLOR", (0, 5), (-1, 5), WHITE),
        ("FONTNAME", (0, 3), (-1, 3), "Helvetica-Bold"),
    ]
    summary_table.setStyle(TableStyle(summary_style))

    right_col = Table([[summary_table]], colWidths=["100%"])
    right_col.setStyle(TableStyle([("ALIGN", (0, 0), (0, 0), "RIGHT")]))
    story.append(right_col)

    if cfg[9]:
        story.append(Spacer(1, 0.4 * cm))
        story.append(Paragraph(f"<b>Notas:</b> {cfg[9]}", normal))

    doc.build(story)
    buffer.seek(0)

    filename = f"cotizacion_{plan_code.replace('-', '_')}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _compute_resumen_raw(cfg) -> dict:
    return {
        "gg_pct": float(cfg[5]),
        "ut_pct": float(cfg[6]),
        "igv_pct": float(cfg[7]),
        "costo_directo": 0,
        "gastos_generales": 0,
        "sub_total": 0,
        "utilidad": 0,
        "valor_venta": 0,
        "igv": 0,
        "precio_total": 0,
    }


# ══════════════════════════════════════════════════════════════════════════════
# EXPORTACIÓN EXCEL
# ══════════════════════════════════════════════════════════════════════════════

def export_excel_service(plan_id: str) -> StreamingResponse:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(500, "openpyxl no está instalado")

    from collections import defaultdict

    # Categorías que usan regla Desglose (fila por ítem) cuando son directas
    DESGLOSE_DIRECT = {"MO", "TRA"}

    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT project_code, title FROM project_plans WHERE id = %s", (plan_id,))
            plan_row = cur.fetchone()
            if not plan_row:
                raise HTTPException(404, "Plan no encontrado")
            plan_code, plan_title = plan_row
            plan_code = plan_code or f"PLAN_{str(plan_id)[:8]}"

        _ensure_config(conn, plan_id)
        with conn.cursor() as cur:
            cur.execute("""
                SELECT cliente_nombre, cliente_ruc, lugar_trabajo, plazo_dias,
                       validez_dias, gastos_generales_pct, utilidad_pct, igv_pct, moneda, notas
                FROM presupuesto_config WHERE plan_id = %s
            """, (plan_id,))
            cfg = cur.fetchone()
            conn.commit()

        # Categorías ordenadas
        with conn.cursor() as cur:
            cur.execute("SELECT codigo, nombre, es_directo, orden, color_hex FROM categorias_costo ORDER BY orden")
            cats = [{"codigo": r[0], "nombre": r[1], "es_directo": r[2], "orden": r[3], "color_hex": r[4]}
                    for r in cur.fetchall()]
        cats_by_code = {c["codigo"]: c for c in cats}

        # Items APU completos con info de partida y categoría
        with conn.cursor() as cur:
            cur.execute("""
                SELECT
                    pp.codigo, pp.descripcion, pp.cantidad AS part_cant,
                    ai.tipo_recurso,
                    COALESCE(m.name, rmo.codigo, ai.descripcion, '—') AS recurso_nombre,
                    ai.unidad, ai.cantidad AS item_cant, ai.precio_unitario,
                    COALESCE(cc.es_directo, TRUE), COALESCE(cc.orden, 99)
                FROM presupuesto_partidas pp
                JOIN presupuesto_apu_items ai ON ai.partida_id = pp.id
                LEFT JOIN categorias_costo cc ON cc.codigo = ai.tipo_recurso
                LEFT JOIN materials m ON m.id = ai.material_id
                LEFT JOIN recursos_mo rmo ON rmo.id = ai.recurso_mo_id
                WHERE pp.plan_id = %s AND pp.es_capitulo = FALSE
                ORDER BY COALESCE(cc.orden, 99), ai.tipo_recurso, pp.orden, pp.codigo
            """, (plan_id,))
            apu_full = cur.fetchall()

        # Items por partida para Hoja 2
        with conn.cursor() as cur:
            cur.execute("""
                SELECT pp.id, pp.codigo, pp.descripcion,
                       ai.tipo_recurso,
                       COALESCE(m.name, rmo.codigo, ai.descripcion, '—'),
                       ai.unidad, ai.cantidad, ai.precio_unitario
                FROM presupuesto_partidas pp
                JOIN presupuesto_apu_items ai ON ai.partida_id = pp.id
                LEFT JOIN materials m ON m.id = ai.material_id
                LEFT JOIN recursos_mo rmo ON rmo.id = ai.recurso_mo_id
                WHERE pp.plan_id = %s AND pp.es_capitulo = FALSE
                ORDER BY pp.orden, pp.codigo, ai.tipo_recurso
            """, (plan_id,))
            apu_by_partida = cur.fetchall()

        # Consolidado de insumos para Hoja 3
        with conn.cursor() as cur:
            cur.execute("""
                SELECT
                    COALESCE(cc.nombre, ai.tipo_recurso),
                    COALESCE(m.name, rmo.codigo, ai.descripcion, '—'),
                    ai.unidad,
                    SUM(pp.cantidad * ai.cantidad),
                    MAX(ai.precio_unitario),
                    COALESCE(cc.orden, 99)
                FROM presupuesto_partidas pp
                JOIN presupuesto_apu_items ai ON ai.partida_id = pp.id
                LEFT JOIN categorias_costo cc ON cc.codigo = ai.tipo_recurso
                LEFT JOIN materials m ON m.id = ai.material_id
                LEFT JOIN recursos_mo rmo ON rmo.id = ai.recurso_mo_id
                WHERE pp.plan_id = %s AND pp.es_capitulo = FALSE
                GROUP BY cc.nombre, ai.tipo_recurso, m.name, rmo.codigo, ai.descripcion, ai.unidad, cc.orden
                ORDER BY COALESCE(cc.orden, 99), 2
            """, (plan_id,))
            consolidado = cur.fetchall()

        resumen = _compute_resumen(conn, plan_id)

    fecha = datetime.now().strftime("%d/%m/%Y")
    moneda_sym = cfg[8] or "S/"
    gg_pct  = float(cfg[5])
    ut_pct  = float(cfg[6])
    igv_pct = float(cfg[7])

    # ── Estilos ──────────────────────────────────────────────────────────────
    thin   = Side(style="thin", color="C5D8DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    fill_primary = PatternFill("solid", fgColor="0B2E33")
    fill_light   = PatternFill("solid", fgColor="EEF7F8")
    fill_accent  = PatternFill("solid", fgColor="B8E3E9")

    font_title   = Font(bold=True, color="FFFFFF", size=12)
    font_col_hdr = Font(bold=True, color="FFFFFF", size=9)
    font_sec_hdr = Font(bold=True, color="0B2E33", size=9)
    font_normal  = Font(size=9)
    font_wht_bold = Font(bold=True, color="FFFFFF", size=10)

    def hex_fill(h: str) -> PatternFill:
        return PatternFill("solid", fgColor=h.lstrip("#"))

    def c_style(cell, fill=None, font=None, align="left", num_fmt=None):
        if fill:  cell.fill = fill
        if font:  cell.font = font
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        cell.border = border
        if num_fmt: cell.number_format = num_fmt

    wb = openpyxl.Workbook()

    # ══════════════════════════════════════════════════════════════════════════
    # HOJA 1 — RESUMEN SERVICIO
    # ══════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "RESUMEN SERVICIO"
    for col, w in zip("ABCDEFG", [8, 40, 7, 10, 8, 12, 14]):
        ws1.column_dimensions[col].width = w

    def w1_cell(r, col, val, fill=None, font=None, align="left", num_fmt=None):
        c = ws1.cell(row=r, column=col, value=val)
        c_style(c, fill, font, align, num_fmt)
        return c

    # Filas 1-5: cabeceras
    ws1.merge_cells("A1:G1")
    c = ws1["A1"]
    c.value = f"CeShark ERP  —  Cotización {plan_code}  —  {fecha}"
    c.fill = fill_primary; c.font = font_title
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border
    ws1.row_dimensions[1].height = 26

    for cells, val in [("A2:D2", f"Proyecto: {plan_title}"),
                        ("E2:G2", f"Cliente: {cfg[0] or '—'}   RUC: {cfg[1] or '—'}"),
                        ("A3:D3", f"Lugar: {cfg[2] or '—'}"),
                        ("E3:G3", f"Plazo: {cfg[3] or '—'} días   Validez: {cfg[4] or 30} días")]:
        ws1.merge_cells(cells)
        c = ws1[cells.split(":")[0]]
        c.value = val; c.fill = fill_light; c.font = font_sec_hdr
        c.alignment = Alignment(horizontal="left", vertical="center"); c.border = border
    ws1.row_dimensions[2].height = 14; ws1.row_dimensions[3].height = 13; ws1.row_dimensions[4].height = 6

    col_hdrs = ["Ítem", "Descripción", "Und", "Cantidad", "Días", f"P.U. ({moneda_sym})", f"Parcial ({moneda_sym})"]
    for col, h in enumerate(col_hdrs, 1):
        w1_cell(5, col, h, fill=fill_primary, font=font_col_hdr, align="center")
    ws1.row_dimensions[5].height = 16

    # Agrupar items por tipo_recurso
    items_by_cat = defaultdict(list)
    for row_data in apu_full:
        part_cod, part_desc, part_cant, tipo, recurso_nombre, unidad, item_cant, item_pu, es_directo, cat_orden = row_data
        items_by_cat[tipo].append({
            "part_cant": float(part_cant), "recurso_nombre": recurso_nombre,
            "unidad": unidad or "", "cantidad": float(item_cant),
            "precio_unitario": float(item_pu), "es_directo": es_directo,
        })

    present_cats_ordered = sorted(
        [cats_by_code[c] for c in items_by_cat if c in cats_by_code],
        key=lambda x: x["orden"]
    )

    r = 6
    item_num = 1
    last_direct_row = 5

    def write_row_formula(ws_, row, item_label, desc, und, cant, dias, pu, row_fill=None):
        ws_.cell(row=row, column=1, value=str(item_label)).fill = row_fill or PatternFill()
        ws_.cell(row=row, column=1).font = font_normal
        ws_.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws_.cell(row=row, column=1).border = border
        for col, (val, al, fmt) in enumerate([
            (desc, "left", None), (und, "center", None),
            (cant, "right", "#,##0.000"),
            (dias, "right", "#,##0.000" if dias else None),
            (pu,   "right", "#,##0.00"),
        ], 2):
            c = ws_.cell(row=row, column=col, value=val)
            if row_fill: c.fill = row_fill
            c.font = font_normal
            c.alignment = Alignment(horizontal=al, vertical="center")
            c.border = border
            if fmt: c.number_format = fmt
        g = ws_.cell(row=row, column=7, value=f"=D{row}*IF(ISBLANK(E{row}),1,E{row})*F{row}")
        if row_fill: g.fill = row_fill
        g.font = font_normal
        g.alignment = Alignment(horizontal="right", vertical="center")
        g.border = border; g.number_format = "#,##0.00"
        ws_.row_dimensions[row].height = 13

    # ── Categorías directas ──
    for cat in present_cats_ordered:
        if not cat["es_directo"]:
            continue
        cat_code  = cat["codigo"]
        cat_items = items_by_cat.get(cat_code, [])
        if not cat_items:
            continue

        # Header de sección
        ws1.merge_cells(f"A{r}:G{r}")
        hc = ws1.cell(row=r, column=1)
        hc.value = cat["nombre"].upper()
        hc.fill = fill_light
        hc.font = Font(bold=True, color=cat["color_hex"].lstrip("#"), size=9)
        hc.alignment = Alignment(horizontal="left", vertical="center")
        hc.border = border
        for col in range(2, 8):
            ws1.cell(row=r, column=col).fill = fill_light
            ws1.cell(row=r, column=col).border = border
        ws1.row_dimensions[r].height = 14
        r += 1

        use_desglose = cat_code in DESGLOSE_DIRECT

        if use_desglose:
            for it in cat_items:
                is_day = it["unidad"].lower().strip() in ("día", "dia", "días", "dias")
                cant_v = it["part_cant"] if is_day else it["part_cant"] * it["cantidad"]
                dias_v = it["cantidad"] if is_day else None
                write_row_formula(ws1, r, item_num, it["recurso_nombre"], it["unidad"],
                                  cant_v, dias_v, it["precio_unitario"])
                last_direct_row = r
                item_num += 1; r += 1
        else:
            total_pu = sum(it["part_cant"] * it["cantidad"] * it["precio_unitario"] for it in cat_items)
            write_row_formula(ws1, r, "", f"{cat['nombre']} Requerido", "Glob", 1, None, round(total_pu, 2))
            last_direct_row = r
            r += 1

    # ── Fila separador ──
    r += 1

    # ── Totales directos ──
    def total_row(ws_, row, label, formula, fill, font):
        ws_.merge_cells(f"A{row}:F{row}")
        c = ws_.cell(row=row, column=1, value=label)
        c.fill = fill; c.font = font
        c.alignment = Alignment(horizontal="right", vertical="center"); c.border = border
        for col in range(2, 7):
            ws_.cell(row=row, column=col).fill = fill; ws_.cell(row=row, column=col).border = border
        g = ws_.cell(row=row, column=7, value=formula)
        g.fill = fill; g.font = font
        g.alignment = Alignment(horizontal="right", vertical="center")
        g.border = border; g.number_format = "#,##0.00"
        ws_.row_dimensions[row].height = 15

    r_cd = r
    total_row(ws1, r, "COSTO DIRECTO", f"=SUM(G6:G{last_direct_row})", fill_accent, font_sec_hdr); r += 1
    r_gg = r
    total_row(ws1, r, f"GASTOS GENERALES ({gg_pct:.1f}%)", f"=G{r_cd}*({gg_pct}/100)", fill_accent, font_sec_hdr); r += 1
    r_ut = r
    total_row(ws1, r, f"UTILIDAD ({ut_pct:.1f}%)", f"=(G{r_cd}+G{r_gg})*({ut_pct}/100)", fill_accent, font_sec_hdr); r += 1

    # ── Categorías indirectas (en sección de totales) ──
    first_ind = None
    last_ind  = None
    ind_counter = 1
    for cat in present_cats_ordered:
        if cat["es_directo"]:
            continue
        cat_items = items_by_cat.get(cat["codigo"], [])
        if not cat_items:
            continue
        ws1.merge_cells(f"A{r}:G{r}")
        hc = ws1.cell(row=r, column=1)
        hc.value = cat["nombre"].upper(); hc.fill = fill_light
        hc.font = Font(bold=True, color=cat["color_hex"].lstrip("#"), size=9)
        hc.alignment = Alignment(horizontal="left", vertical="center"); hc.border = border
        for col in range(2, 8):
            ws1.cell(row=r, column=col).fill = fill_light; ws1.cell(row=r, column=col).border = border
        ws1.row_dimensions[r].height = 14; r += 1
        sub = 1
        for it in cat_items:
            is_day = it["unidad"].lower().strip() in ("día", "dia", "días", "dias")
            cant_v = it["part_cant"] if is_day else it["part_cant"] * it["cantidad"]
            dias_v = it["cantidad"] if is_day else None
            write_row_formula(ws1, r, f"{ind_counter}.{sub}", it["recurso_nombre"], it["unidad"],
                              cant_v, dias_v, it["precio_unitario"])
            if first_ind is None: first_ind = r
            last_ind = r
            sub += 1; r += 1
        ind_counter += 1

    # VALOR VENTA, IGV, PRECIO TOTAL
    r_vv = r
    vv_f = (f"=G{r_cd}+G{r_gg}+G{r_ut}+SUM(G{first_ind}:G{last_ind})"
            if first_ind else f"=G{r_cd}+G{r_gg}+G{r_ut}")
    total_row(ws1, r, "VALOR VENTA", vv_f, fill_primary, font_col_hdr); r += 1
    r_igv = r
    total_row(ws1, r, f"IGV ({igv_pct:.1f}%)", f"=G{r_vv}*({igv_pct}/100)", fill_accent, font_sec_hdr); r += 1
    total_row(ws1, r, "PRECIO TOTAL", f"=G{r_vv}+G{r_igv}", fill_primary, font_wht_bold)

    # ══════════════════════════════════════════════════════════════════════════
    # HOJA 2 — APU Detallado
    # ══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("APU Detallado")
    for col, w in zip("ABCDEF", [10, 14, 30, 8, 10, 14]):
        ws2.column_dimensions[col].width = w

    ws2.merge_cells("A1:F1")
    c = ws2["A1"]
    c.value = f"APU Detallado — {plan_code}"
    c.fill = fill_primary; c.font = font_title
    c.alignment = Alignment(horizontal="center", vertical="center"); c.border = border
    ws2.row_dimensions[1].height = 22

    # Agrupar por partida
    partida_items = defaultdict(list)
    partida_meta  = {}
    for row_data in apu_by_partida:
        pid, pcod, pdesc, tipo, rec_nombre, und, item_cant, item_pu = row_data
        partida_meta[pid] = (pcod, pdesc)
        partida_items[pid].append((tipo, rec_nombre, und, float(item_cant), float(item_pu)))

    r2 = 2
    for pid, (pcod, pdesc) in partida_meta.items():
        items = partida_items[pid]
        # Header partida
        ws2.merge_cells(f"A{r2}:F{r2}")
        c = ws2.cell(row=r2, column=1, value=f"{pcod} — {pdesc}")
        c.fill = PatternFill("solid", fgColor="1a4a52"); c.font = Font(bold=True, color="FFFFFF", size=9)
        c.alignment = Alignment(horizontal="left", vertical="center"); c.border = border
        for col in range(2, 7):
            ws2.cell(row=r2, column=col).fill = PatternFill("solid", fgColor="1a4a52")
            ws2.cell(row=r2, column=col).border = border
        ws2.row_dimensions[r2].height = 14; r2 += 1
        # Col headers
        for col, h in enumerate(["Tipo", "Recurso", "Descripción", "Und", "Cant", "P.U.", "Parcial"], 1):
            if col <= 6:
                c = ws2.cell(row=r2, column=col, value=h)
                c.fill = fill_accent; c.font = font_sec_hdr
                c.alignment = Alignment(horizontal="center", vertical="center"); c.border = border
        ws2.row_dimensions[r2].height = 13; r2 += 1
        first_item_row = r2
        for tipo, rec_nombre, und, cant, pu in items:
            cat_info = cats_by_code.get(tipo, {})
            chip_fill = hex_fill(cat_info.get("color_hex", "#4F7C82") + "20") if cat_info else fill_light
            for col, (val, al, fmt) in enumerate([
                (tipo,        "center", None),
                (rec_nombre,  "left",   None),
                (und or "",   "center", None),
                (cant,        "right",  "#,##0.000"),
                (pu,          "right",  "#,##0.00"),
            ], 1):
                c = ws2.cell(row=r2, column=col, value=val)
                c.fill = chip_fill if col == 1 else (fill_light if r2 % 2 == 0 else PatternFill())
                c.font = Font(bold=True, size=9) if col == 1 else font_normal
                c.alignment = Alignment(horizontal=al, vertical="center"); c.border = border
                if fmt: c.number_format = fmt
            f_cell = ws2.cell(row=r2, column=6, value=f"=D{r2}*E{r2}")
            f_cell.fill = fill_light if r2 % 2 == 0 else PatternFill()
            f_cell.font = font_normal
            f_cell.alignment = Alignment(horizontal="right", vertical="center")
            f_cell.border = border; f_cell.number_format = "#,##0.00"
            ws2.row_dimensions[r2].height = 13; r2 += 1
        # Total partida
        t = ws2.cell(row=r2, column=1, value="TOTAL PARTIDA")
        ws2.merge_cells(f"A{r2}:E{r2}")
        t.fill = fill_accent; t.font = font_sec_hdr
        t.alignment = Alignment(horizontal="right", vertical="center"); t.border = border
        for col in range(2, 6):
            ws2.cell(row=r2, column=col).fill = fill_accent; ws2.cell(row=r2, column=col).border = border
        tot = ws2.cell(row=r2, column=6, value=f"=SUM(F{first_item_row}:F{r2 - 1})")
        tot.fill = fill_accent; tot.font = font_sec_hdr
        tot.alignment = Alignment(horizontal="right", vertical="center")
        tot.border = border; tot.number_format = "#,##0.00"
        ws2.row_dimensions[r2].height = 14; r2 += 2

    # ══════════════════════════════════════════════════════════════════════════
    # HOJA 3 — Consolidado de Insumos
    # ══════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Consolidado de Insumos")
    for col, w in zip("ABCDEF", [16, 38, 8, 12, 12, 14]):
        ws3.column_dimensions[col].width = w

    ws3.merge_cells("A1:F1")
    c = ws3["A1"]
    c.value = f"Consolidado de Insumos — {plan_code}"
    c.fill = fill_primary; c.font = font_title
    c.alignment = Alignment(horizontal="center", vertical="center"); c.border = border
    ws3.row_dimensions[1].height = 22

    for col, h in enumerate(["Categoría", "Descripción", "Und", "Cantidad Total", f"P.U. ({moneda_sym})", "Total"], 1):
        c = ws3.cell(row=2, column=col, value=h)
        c.fill = fill_primary; c.font = font_col_hdr
        c.alignment = Alignment(horizontal="center", vertical="center"); c.border = border
    ws3.row_dimensions[2].height = 15

    for ri, row_data in enumerate(consolidado, start=3):
        cat_nombre, descripcion, und, cantidad_total, precio_unitario, _ = row_data
        row_fill = fill_light if ri % 2 == 0 else PatternFill()
        for col, (val, al, fmt) in enumerate([
            (cat_nombre,          "center", None),
            (descripcion,         "left",   None),
            (und or "",           "center", None),
            (float(cantidad_total), "right", "#,##0.000"),
            (float(precio_unitario), "right", "#,##0.00"),
        ], 1):
            c = ws3.cell(row=ri, column=col, value=val)
            c.fill = row_fill; c.font = font_normal
            c.alignment = Alignment(horizontal=al, vertical="center"); c.border = border
            if fmt: c.number_format = fmt
        f_cell = ws3.cell(row=ri, column=6, value=f"=D{ri}*E{ri}")
        f_cell.fill = row_fill; f_cell.font = font_normal
        f_cell.alignment = Alignment(horizontal="right", vertical="center")
        f_cell.border = border; f_cell.number_format = "#,##0.00"
        ws3.row_dimensions[ri].height = 13

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"cotizacion_{plan_code.replace('-', '_')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ══════════════════════════════════════════════════════════════════════════════
# APU BULK — guardar múltiples ítems de una vez
# ══════════════════════════════════════════════════════════════════════════════

def bulk_create_apu_service(partida_id: str, items: list) -> list:
    if not items:
        raise HTTPException(400, "Lista de ítems vacía")

    with db_connection() as conn:
        valid_tipos = _get_valid_tipos(conn)
        with conn.cursor() as cur:
            cur.execute("SELECT id FROM presupuesto_partidas WHERE id = %s", (partida_id,))
            if not cur.fetchone():
                raise HTTPException(404, "Partida no encontrada")

            results = []
            for payload in items:
                if payload.tipo_recurso not in valid_tipos:
                    raise HTTPException(400, f"tipo_recurso inválido: {payload.tipo_recurso}")

                mat_nombre, mo_codigo = None, None

                if payload.tipo_recurso != "MO" and payload.material_id:
                    cur.execute("SELECT name FROM materials WHERE id = %s", (payload.material_id,))
                    row = cur.fetchone()
                    if not row:
                        raise HTTPException(400, f"material_id no existe: {payload.material_id}")
                    mat_nombre = row[0]

                if payload.tipo_recurso == "MO" and payload.recurso_mo_id:
                    cur.execute("SELECT codigo FROM recursos_mo WHERE id = %s", (payload.recurso_mo_id,))
                    row = cur.fetchone()
                    if not row:
                        raise HTTPException(400, f"recurso_mo_id no existe: {payload.recurso_mo_id}")
                    mo_codigo = row[0]

                cur.execute("""
                    INSERT INTO presupuesto_apu_items
                        (partida_id, tipo_recurso, material_id, recurso_mo_id,
                         descripcion, unidad, cantidad, precio_unitario)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    RETURNING id, partida_id, tipo_recurso, material_id, recurso_mo_id,
                              descripcion, unidad, cantidad, precio_unitario
                """, (
                    partida_id, payload.tipo_recurso,
                    payload.material_id if payload.tipo_recurso != "MO" else None,
                    payload.recurso_mo_id if payload.tipo_recurso == "MO" else None,
                    payload.descripcion, payload.unidad,
                    payload.cantidad, payload.precio_unitario,
                ))
                row = cur.fetchone()
                results.append(_row_to_apu((*row, mat_nombre, mo_codigo)))

            conn.commit()
    return results


# ══════════════════════════════════════════════════════════════════════════════
# BAÚLES APU — kits preconfigurados
# ══════════════════════════════════════════════════════════════════════════════

def _row_to_baul(r) -> dict:
    return {
        "id": str(r[0]), "nombre": r[1], "descripcion": r[2],
        "categoria": r[3], "activo": r[4],
        "created_at": r[5].isoformat() if r[5] else None,
    }

def _row_to_baul_item(r) -> dict:
    return {
        "id": str(r[0]), "baul_id": str(r[1]),
        "tipo_recurso": r[2],
        "material_id": str(r[3]) if r[3] else None,
        "recurso_mo_id": str(r[4]) if r[4] else None,
        "descripcion": r[5], "unidad": r[6],
        "cantidad_base": float(r[7]), "precio_unitario": float(r[8]),
        "orden": r[9],
        "material_nombre": r[10],
        "recurso_mo_codigo": r[11],
    }


def list_baules_service() -> list:
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT b.id, b.nombre, b.descripcion, b.categoria, b.activo, b.created_at
                FROM apu_baules b
                WHERE b.activo = TRUE
                ORDER BY b.categoria, b.nombre
            """)
            baules = [_row_to_baul(r) for r in cur.fetchall()]

            for baul in baules:
                cur.execute("""
                    SELECT bi.id, bi.baul_id, bi.tipo_recurso,
                           bi.material_id, bi.recurso_mo_id,
                           bi.descripcion, bi.unidad,
                           bi.cantidad_base, bi.precio_unitario, bi.orden,
                           m.name, rmo.codigo
                    FROM apu_baul_items bi
                    LEFT JOIN materials m ON m.id = bi.material_id
                    LEFT JOIN recursos_mo rmo ON rmo.id = bi.recurso_mo_id
                    WHERE bi.baul_id = %s
                    ORDER BY bi.orden, bi.created_at
                """, (baul["id"],))
                baul["items"] = [_row_to_baul_item(r) for r in cur.fetchall()]

    return baules


def create_baul_service(payload) -> dict:
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                INSERT INTO apu_baules (nombre, descripcion, categoria)
                VALUES (%s, %s, %s)
                RETURNING id, nombre, descripcion, categoria, activo, created_at
            """, (payload.nombre, payload.descripcion, payload.categoria))
            row = cur.fetchone()
            conn.commit()
    baul = _row_to_baul(row)
    baul["items"] = []
    return baul


def update_baul_service(baul_id: str, payload) -> dict:
    fields, vals = [], []
    if payload.nombre is not None:
        fields.append("nombre = %s"); vals.append(payload.nombre)
    if payload.descripcion is not None:
        fields.append("descripcion = %s"); vals.append(payload.descripcion)
    if payload.categoria is not None:
        fields.append("categoria = %s"); vals.append(payload.categoria)
    if payload.activo is not None:
        fields.append("activo = %s"); vals.append(payload.activo)
    if not fields:
        raise HTTPException(400, "Nada que actualizar")
    vals.append(baul_id)

    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                f"UPDATE apu_baules SET {', '.join(fields)}, updated_at = NOW() "
                "WHERE id = %s RETURNING id, nombre, descripcion, categoria, activo, created_at",
                vals,
            )
            row = cur.fetchone()
            if not row:
                raise HTTPException(404, "Baúl no encontrado")
            conn.commit()
    return _row_to_baul(row)


def delete_baul_service(baul_id: str) -> dict:
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "UPDATE apu_baules SET activo = FALSE WHERE id = %s RETURNING id", (baul_id,)
            )
            if not cur.fetchone():
                raise HTTPException(404, "Baúl no encontrado")
            conn.commit()
    return {"ok": True}


def add_baul_item_service(baul_id: str, payload) -> dict:
    with db_connection() as conn:
        valid_tipos = _get_valid_tipos(conn)
        if payload.tipo_recurso not in valid_tipos:
            raise HTTPException(400, f"tipo_recurso inválido: {payload.tipo_recurso}")

        with conn.cursor() as cur:
            cur.execute("SELECT id FROM apu_baules WHERE id = %s", (baul_id,))
            if not cur.fetchone():
                raise HTTPException(404, "Baúl no encontrado")

            mat_nombre, mo_codigo = None, None
            if payload.material_id:
                cur.execute("SELECT name FROM materials WHERE id = %s", (payload.material_id,))
                r = cur.fetchone()
                mat_nombre = r[0] if r else None
            if payload.recurso_mo_id:
                cur.execute("SELECT codigo FROM recursos_mo WHERE id = %s", (payload.recurso_mo_id,))
                r = cur.fetchone()
                mo_codigo = r[0] if r else None

            cur.execute("""
                INSERT INTO apu_baul_items
                    (baul_id, tipo_recurso, material_id, recurso_mo_id,
                     descripcion, unidad, cantidad_base, precio_unitario, orden)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id, baul_id, tipo_recurso, material_id, recurso_mo_id,
                          descripcion, unidad, cantidad_base, precio_unitario, orden
            """, (
                baul_id, payload.tipo_recurso,
                payload.material_id if payload.tipo_recurso != "MO" else None,
                payload.recurso_mo_id if payload.tipo_recurso == "MO" else None,
                payload.descripcion, payload.unidad,
                payload.cantidad_base, payload.precio_unitario, payload.orden,
            ))
            row = cur.fetchone()
            conn.commit()
    return _row_to_baul_item((*row, mat_nombre, mo_codigo))


def update_baul_item_service(baul_id: str, item_id: str, payload) -> dict:
    fields, vals = [], []
    if payload.descripcion is not None:
        fields.append("descripcion = %s"); vals.append(payload.descripcion)
    if payload.unidad is not None:
        fields.append("unidad = %s"); vals.append(payload.unidad)
    if payload.cantidad_base is not None:
        fields.append("cantidad_base = %s"); vals.append(payload.cantidad_base)
    if payload.precio_unitario is not None:
        fields.append("precio_unitario = %s"); vals.append(payload.precio_unitario)
    if payload.orden is not None:
        fields.append("orden = %s"); vals.append(payload.orden)
    if not fields:
        raise HTTPException(400, "Nada que actualizar")
    vals += [item_id, baul_id]

    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                f"UPDATE apu_baul_items SET {', '.join(fields)} "
                "WHERE id = %s AND baul_id = %s "
                "RETURNING id, baul_id, tipo_recurso, material_id, recurso_mo_id, "
                "descripcion, unidad, cantidad_base, precio_unitario, orden",
                vals,
            )
            row = cur.fetchone()
            if not row:
                raise HTTPException(404, "Ítem de baúl no encontrado")
            conn.commit()
    return _row_to_baul_item((*row, None, None))


def delete_baul_item_service(baul_id: str, item_id: str) -> dict:
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "DELETE FROM apu_baul_items WHERE id = %s AND baul_id = %s RETURNING id",
                (item_id, baul_id)
            )
            if not cur.fetchone():
                raise HTTPException(404, "Ítem no encontrado")
            conn.commit()
    return {"ok": True}


# ══════════════════════════════════════════════════════════════════════════════
# 📥 IMPORTAR BAÚLES DESDE EXCEL
# ══════════════════════════════════════════════════════════════════════════════

def import_baules_from_excel_service(file) -> dict:
    """
    Formato del Excel (una fila por ítem):
    Baúl Nombre | Categoría | Descripción baúl | Tipo recurso |
    Descripción ítem | Unidad | Cantidad base | Precio unitario | Cód. material
    """
    from app.modules.logistics.utils import read_excel

    df = read_excel(file)

    COL_MAP = {
        "Baúl Nombre":       "baul_nombre",
        "Categoría":         "categoria",
        "Descripción baúl":  "descripcion_baul",
        "Tipo recurso":      "tipo_recurso",
        "Descripción ítem":  "descripcion_item",
        "Unidad":            "unidad",
        "Cantidad base":     "cantidad_base",
        "Precio unitario":   "precio_unitario",
        "Cód. material":     "cod_material",
    }
    df = df.rename(columns=COL_MAP)

    if "baul_nombre" not in df.columns:
        raise HTTPException(400, "Falta la columna 'Baúl Nombre'")
    if "tipo_recurso" not in df.columns:
        raise HTTPException(400, "Falta la columna 'Tipo recurso'")
    if "descripcion_item" not in df.columns:
        raise HTTPException(400, "Falta la columna 'Descripción ítem'")

    created_baules  = 0
    updated_baules  = 0
    total_items     = 0
    errors          = []

    # Agrupar filas por baúl
    groups: dict[str, list] = {}
    for idx, row in df.iterrows():
        nombre = str(row.get("baul_nombre") or "").strip()
        if not nombre:
            errors.append(f"Fila {int(idx)+2}: 'Baúl Nombre' vacío — fila ignorada")
            continue
        groups.setdefault(nombre, []).append((int(idx) + 2, row))

    with db_connection() as conn:
        with conn.cursor() as cur:
            TIPOS_VALIDOS = set(_get_valid_tipos(conn))
            for baul_nombre, filas in groups.items():
                # Tomar metadatos del baúl de la primera fila
                first_row = filas[0][1]
                categoria = str(first_row.get("categoria") or "").strip() or None
                desc_baul = str(first_row.get("descripcion_baul") or "").strip() or None

                # UPSERT del baúl por nombre
                cur.execute(
                    "SELECT id FROM apu_baules WHERE nombre = %s", (baul_nombre,)
                )
                existing = cur.fetchone()
                if existing:
                    baul_id = existing[0]
                    cur.execute(
                        "UPDATE apu_baules SET categoria=%s, descripcion=%s, activo=TRUE, updated_at=NOW() WHERE id=%s",
                        (categoria, desc_baul, baul_id),
                    )
                    cur.execute("DELETE FROM apu_baul_items WHERE baul_id=%s", (baul_id,))
                    updated_baules += 1
                else:
                    cur.execute(
                        "INSERT INTO apu_baules (nombre, descripcion, categoria) VALUES (%s,%s,%s) RETURNING id",
                        (baul_nombre, desc_baul, categoria),
                    )
                    baul_id = cur.fetchone()[0]
                    created_baules += 1

                # Insertar ítems
                for orden, (excel_row, row_data) in enumerate(filas, start=1):
                    try:
                        tipo = str(row_data.get("tipo_recurso") or "").strip().upper()
                        if tipo not in TIPOS_VALIDOS:
                            raise ValueError(f"Tipo recurso inválido '{tipo}' — debe ser uno de: {', '.join(sorted(TIPOS_VALIDOS))}")

                        desc_item = str(row_data.get("descripcion_item") or "").strip()
                        if not desc_item:
                            raise ValueError("'Descripción ítem' está vacío")

                        unidad    = str(row_data.get("unidad") or "").strip() or "und"
                        cant_base = float(row_data["cantidad_base"]) if row_data.get("cantidad_base") is not None else 1.0
                        precio    = float(row_data["precio_unitario"]) if row_data.get("precio_unitario") is not None else 0.0

                        # Resolver material por código si viene el campo
                        material_id = None
                        cod_mat = str(row_data.get("cod_material") or "").strip()
                        if cod_mat:
                            cur.execute("SELECT id FROM materials WHERE code=%s", (cod_mat,))
                            mat_row = cur.fetchone()
                            if mat_row:
                                material_id = mat_row[0]
                            else:
                                errors.append(f"Fila {excel_row}: Cód. material '{cod_mat}' no existe — ítem importado sin vincular")

                        cur.execute("""
                            INSERT INTO apu_baul_items
                                (baul_id, tipo_recurso, material_id, descripcion, unidad,
                                 cantidad_base, precio_unitario, orden)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                        """, (baul_id, tipo, material_id, desc_item, unidad,
                              cant_base, precio, orden))
                        total_items += 1

                    except Exception as e:
                        errors.append(f"Fila {excel_row}: {e}")

        conn.commit()

    return {
        "created":     created_baules,
        "updated":     updated_baules,
        "total_baules": created_baules + updated_baules,
        "total_items":  total_items,
        "errors":       errors,
    }


# ══════════════════════════════════════════════════════════════════════════════
# TARIFAS DE PERSONAL — Matriz contextual (Fase 6B)
# ══════════════════════════════════════════════════════════════════════════════

VALID_CONTEXTOS  = ("PARADA", "PROYECTO", "SERVICIO", "INGENIERIA")
VALID_UBICACIONES = ("MINA", "AREQUIPA", "INDUSTRIA", "CUALQUIERA")
VALID_MODALIDADES = ("HORA", "DIA")


def _row_to_tarifa(r) -> dict:
    return {
        "id":                   str(r[0]),
        "rol":                  r[1],
        "contexto":             r[2],
        "ubicacion":            r[3],
        "modalidad":            r[4],
        "horas_por_dia":        r[5],
        "tarifa":               float(r[6]),
        "tarifa_hora_extra":    float(r[7]) if r[7] is not None else None,
        "moneda":               r[8],
        "incluye_epp":          r[9],
        "incluye_herramientas": r[10],
        "notas":                r[11],
        "activo":               r[12],
        "created_at":           r[13].isoformat() if r[13] else None,
    }

_TARIFA_SELECT = """
    SELECT id, rol, contexto, ubicacion, modalidad, horas_por_dia,
           tarifa, tarifa_hora_extra, moneda,
           incluye_epp, incluye_herramientas, notas, activo, created_at
    FROM tarifas_personal
"""


def list_tarifas_personal_service(
    rol: str = None,
    contexto: str = None,
    ubicacion: str = None,
    modalidad: str = None,
    activo: bool = True,
) -> list:
    with db_connection() as conn:
        with conn.cursor() as cur:
            q = _TARIFA_SELECT + " WHERE 1=1"
            params = []
            if activo is not None:
                q += " AND activo = %s"
                params.append(activo)
            if rol:
                q += " AND LOWER(rol) LIKE LOWER(%s)"
                params.append(f"%{rol}%")
            if contexto:
                q += " AND contexto = %s"
                params.append(contexto.upper())
            if ubicacion:
                q += " AND ubicacion = %s"
                params.append(ubicacion.upper())
            if modalidad:
                q += " AND modalidad = %s"
                params.append(modalidad.upper())
            q += " ORDER BY rol, contexto, ubicacion, modalidad"
            cur.execute(q, params)
            return [_row_to_tarifa(r) for r in cur.fetchall()]


def create_tarifa_personal_service(payload) -> dict:
    contexto  = payload.contexto.upper()
    ubicacion = payload.ubicacion.upper()
    modalidad = payload.modalidad.upper()

    if contexto not in VALID_CONTEXTOS:
        raise HTTPException(400, f"contexto inválido: {contexto}. Válidos: {VALID_CONTEXTOS}")
    if ubicacion not in VALID_UBICACIONES:
        raise HTTPException(400, f"ubicacion inválida: {ubicacion}. Válidos: {VALID_UBICACIONES}")
    if modalidad not in VALID_MODALIDADES:
        raise HTTPException(400, f"modalidad inválida: {modalidad}. Válidos: {VALID_MODALIDADES}")

    with db_connection() as conn:
        with conn.cursor() as cur:
            try:
                cur.execute("""
                    INSERT INTO tarifas_personal
                        (rol, contexto, ubicacion, modalidad, horas_por_dia,
                         tarifa, tarifa_hora_extra, moneda,
                         incluye_epp, incluye_herramientas, notas)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    RETURNING id, rol, contexto, ubicacion, modalidad, horas_por_dia,
                              tarifa, tarifa_hora_extra, moneda,
                              incluye_epp, incluye_herramientas, notas, activo, created_at
                """, (
                    payload.rol.strip(), contexto, ubicacion, modalidad,
                    payload.horas_por_dia, payload.tarifa, payload.tarifa_hora_extra,
                    payload.moneda, payload.incluye_epp, payload.incluye_herramientas,
                    payload.notas,
                ))
                row = cur.fetchone()
                conn.commit()
            except Exception as e:
                if "idx_tarifas_unico" in str(e):
                    raise HTTPException(
                        409,
                        f"Ya existe una tarifa activa para {payload.rol} / {contexto} / {ubicacion} / {modalidad}"
                    )
                raise
    return _row_to_tarifa(row)


def update_tarifa_personal_service(tarifa_id: str, payload) -> dict:
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT id FROM tarifas_personal WHERE id = %s", (tarifa_id,))
            if not cur.fetchone():
                raise HTTPException(404, "Tarifa no encontrada")

            fields, vals = [], []
            for field in ("rol", "contexto", "ubicacion", "modalidad",
                          "horas_por_dia", "tarifa", "tarifa_hora_extra",
                          "moneda", "incluye_epp", "incluye_herramientas",
                          "notas", "activo"):
                val = getattr(payload, field, None)
                if val is not None:
                    if field in ("contexto", "ubicacion", "modalidad"):
                        val = val.upper()
                    fields.append(f"{field} = %s")
                    vals.append(val)

            if not fields:
                raise HTTPException(400, "Sin campos para actualizar")

            vals.append(tarifa_id)
            try:
                cur.execute(
                    f"UPDATE tarifas_personal SET {', '.join(fields)} WHERE id = %s "
                    "RETURNING id, rol, contexto, ubicacion, modalidad, horas_por_dia, "
                    "tarifa, tarifa_hora_extra, moneda, "
                    "incluye_epp, incluye_herramientas, notas, activo, created_at",
                    vals,
                )
                row = cur.fetchone()
                conn.commit()
            except Exception as e:
                if "idx_tarifas_unico" in str(e):
                    raise HTTPException(409, "Ya existe una tarifa activa con esa combinación")
                raise
    return _row_to_tarifa(row)


def delete_tarifa_personal_service(tarifa_id: str) -> dict:
    """Soft delete — marca activo=FALSE."""
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "UPDATE tarifas_personal SET activo = FALSE WHERE id = %s RETURNING id",
                (tarifa_id,)
            )
            if not cur.fetchone():
                raise HTTPException(404, "Tarifa no encontrada")
            conn.commit()
    return {"ok": True}


def buscar_tarifa_personal_service(rol: str, contexto: str, ubicacion: str, modalidad: str):
    """
    Retorna la tarifa más específica para la combinación dada.
    Si no hay match exacto en ubicacion, busca CUALQUIERA como fallback.
    Retorna None si no hay ninguna tarifa aplicable.
    """
    contexto  = contexto.upper()
    ubicacion = ubicacion.upper()
    modalidad = modalidad.upper()

    with db_connection() as conn:
        with conn.cursor() as cur:
            # Intento 1: match exacto
            cur.execute(
                _TARIFA_SELECT + """
                WHERE LOWER(rol) = LOWER(%s)
                  AND contexto = %s AND ubicacion = %s AND modalidad = %s
                  AND activo = TRUE
                LIMIT 1
                """,
                (rol, contexto, ubicacion, modalidad),
            )
            row = cur.fetchone()
            if row:
                return _row_to_tarifa(row)

            # Intento 2: fallback a CUALQUIERA
            cur.execute(
                _TARIFA_SELECT + """
                WHERE LOWER(rol) = LOWER(%s)
                  AND contexto = %s AND ubicacion = 'CUALQUIERA' AND modalidad = %s
                  AND activo = TRUE
                LIMIT 1
                """,
                (rol, contexto, modalidad),
            )
            row = cur.fetchone()
            if row:
                return {**_row_to_tarifa(row), "_fallback": True}

    return None


def list_roles_tarifas_service() -> list:
    """Lista única de roles disponibles en la matriz (para el selector APU)."""
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT DISTINCT rol FROM tarifas_personal WHERE activo = TRUE ORDER BY rol"
            )
            return [r[0] for r in cur.fetchall()]
