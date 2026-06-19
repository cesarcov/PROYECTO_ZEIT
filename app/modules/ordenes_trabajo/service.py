from datetime import datetime, timezone
from fastapi import HTTPException
from app.core.database import db_connection
from app.core.utils import generate_sequential_code
from app.modules.ordenes_trabajo.schemas import (
    VALID_TIPOS, VALID_PRIORIDAD, VALID_STATUS,
)


# ══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def _gen_ot_code(conn) -> str:
    return generate_sequential_code(conn, "ordenes_trabajo", "OT")


def _row_to_ot(r) -> dict:
    return {
        "id":               str(r[0]),
        "code":             r[1],
        "plan_id":          str(r[2]) if r[2] else None,
        "partida_id":       str(r[3]) if r[3] else None,
        "titulo":           r[4],
        "descripcion":      r[5],
        "tipo":             r[6],
        "prioridad":        r[7],
        "status":           r[8],
        "asignado_a":       str(r[9])  if r[9]  else None,
        "creado_por":       str(r[10]) if r[10] else None,
        "fecha_inicio_plan": r[11].isoformat() if r[11] else None,
        "fecha_fin_plan":   r[12].isoformat() if r[12] else None,
        "fecha_inicio_real": r[13].isoformat() if r[13] else None,
        "fecha_fin_real":   r[14].isoformat() if r[14] else None,
        "horas_estimadas":  float(r[15]) if r[15] is not None else None,
        "horas_reales":     float(r[16]) if r[16] is not None else None,
        "lugar_trabajo":    r[17],
        "observaciones":    r[18],
        "created_at":       r[19].isoformat() if r[19] else None,
        "updated_at":       r[20].isoformat() if r[20] else None,
        # joined
        "asignado_nombre":  r[21],
        "plan_code":        r[22],
    }


def _row_to_checklist(r) -> dict:
    return {
        "id":           str(r[0]),
        "ot_id":        str(r[1]),
        "orden":        r[2],
        "descripcion":  r[3],
        "completado":   r[4],
        "completado_por": str(r[5]) if r[5] else None,
        "completado_at": r[6].isoformat() if r[6] else None,
        "notas":        r[7],
        "completado_nombre": r[8],
    }


def _row_to_material(r) -> dict:
    return {
        "id":               str(r[0]),
        "ot_id":            str(r[1]),
        "material_id":      str(r[2]),
        "almacen_id":       str(r[3]) if r[3] else None,
        "cantidad_plan":    float(r[4]) if r[4] is not None else None,
        "cantidad_real":    float(r[5]),
        "stock_movement_id": str(r[6]) if r[6] else None,
        "created_at":       r[7].isoformat() if r[7] else None,
        "material_nombre":  r[8],
        "material_unidad":  r[9],
        "almacen_nombre":   r[10],
        "registrado_nombre": r[11],
        "oc_id":            str(r[12]) if r[12] else None,
        "oc_code":          r[13],
        "stock_disponible": float(r[14]) if r[14] is not None else 0.0,
    }


def _row_to_tiempo(r) -> dict:
    return {
        "id":          str(r[0]),
        "ot_id":       str(r[1]),
        "tecnico_id":  str(r[2]) if r[2] else None,
        "inicio":      r[3].isoformat() if r[3] else None,
        "fin":         r[4].isoformat() if r[4] else None,
        "horas":       float(r[5]) if r[5] is not None else None,
        "notas":       r[6],
        "tecnico_nombre": r[7],
    }


OT_SELECT = """
    SELECT ot.id, ot.code, ot.plan_id, ot.partida_id,
           ot.titulo, ot.descripcion, ot.tipo, ot.prioridad, ot.status,
           ot.asignado_a, ot.creado_por,
           ot.fecha_inicio_plan, ot.fecha_fin_plan,
           ot.fecha_inicio_real, ot.fecha_fin_real,
           ot.horas_estimadas, ot.horas_reales,
           ot.lugar_trabajo, ot.observaciones,
           ot.created_at, ot.updated_at,
           u.username   AS asignado_nombre,
           pp.project_code AS plan_code
    FROM ordenes_trabajo ot
    LEFT JOIN users u  ON u.id = ot.asignado_a
    LEFT JOIN project_plans pp ON pp.id = ot.plan_id
"""


# ══════════════════════════════════════════════════════════════════════════════
# LISTA Y DETALLE
# ══════════════════════════════════════════════════════════════════════════════

def list_ot_service(status: str = None, plan_id: str = None,
                    tipo: str = None, asignado_a: str = None) -> list:
    conditions, vals = [], []
    if status:
        conditions.append("ot.status = %s"); vals.append(status)
    if plan_id:
        conditions.append("ot.plan_id = %s::uuid"); vals.append(plan_id)
    if tipo:
        conditions.append("ot.tipo = %s"); vals.append(tipo)
    if asignado_a:
        conditions.append("ot.asignado_a = %s::uuid"); vals.append(asignado_a)

    where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(f"{OT_SELECT} {where} ORDER BY ot.created_at DESC", vals)
            rows = cur.fetchall()
    return [_row_to_ot(r) for r in rows]


def get_ot_service(ot_id: str) -> dict:
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(f"{OT_SELECT} WHERE ot.id = %s::uuid", (ot_id,))
            row = cur.fetchone()
            if not row:
                raise HTTPException(404, "OT no encontrada")
            ot = _row_to_ot(row)

            # Checklist
            cur.execute("""
                SELECT cl.id, cl.ot_id, cl.orden, cl.descripcion,
                       cl.completado, cl.completado_por, cl.completado_at, cl.notas,
                       u.username
                FROM ot_checklist cl
                LEFT JOIN users u ON u.id = cl.completado_por
                WHERE cl.ot_id = %s::uuid
                ORDER BY cl.orden, cl.id
            """, (ot_id,))
            ot["checklist"] = [_row_to_checklist(r) for r in cur.fetchall()]

            # Materiales
            cur.execute("""
                SELECT om.id, om.ot_id, om.material_id, om.almacen_id,
                       om.cantidad_plan, om.cantidad_real, om.stock_movement_id,
                       om.created_at,
                       m.name, m.unit, w.name, u.username,
                       om.oc_id, oc.code AS oc_code,
                       COALESCE((
                           SELECT SUM(sl.quantity)
                           FROM stock_locations sl
                           WHERE sl.material_id = om.material_id
                       ), 0) AS stock_disponible
                FROM ot_materiales om
                JOIN materials m ON m.id = om.material_id
                LEFT JOIN warehouses w ON w.id = om.almacen_id
                LEFT JOIN users u ON u.id = om.registrado_por
                LEFT JOIN ordenes_compra oc ON oc.id = om.oc_id
                WHERE om.ot_id = %s::uuid
                ORDER BY om.created_at
            """, (ot_id,))
            ot["materiales"] = [_row_to_material(r) for r in cur.fetchall()]

            # Tiempos
            cur.execute("""
                SELECT t.id, t.ot_id, t.tecnico_id,
                       t.inicio, t.fin, t.horas, t.notas,
                       u.username
                FROM ot_tiempos t
                LEFT JOIN users u ON u.id = t.tecnico_id
                WHERE t.ot_id = %s::uuid
                ORDER BY t.inicio
            """, (ot_id,))
            ot["tiempos"] = [_row_to_tiempo(r) for r in cur.fetchall()]

            # Tiempo activo (sin fin)
            ot["tiene_tiempo_activo"] = any(t["fin"] is None for t in ot["tiempos"])

    return ot


# ══════════════════════════════════════════════════════════════════════════════
# CREAR / EDITAR
# ══════════════════════════════════════════════════════════════════════════════

def create_ot_service(payload, user: dict) -> dict:
    if payload.tipo not in VALID_TIPOS:
        raise HTTPException(400, f"tipo debe ser uno de: {', '.join(VALID_TIPOS)}")
    if payload.prioridad not in VALID_PRIORIDAD:
        raise HTTPException(400, f"prioridad debe ser uno de: {', '.join(VALID_PRIORIDAD)}")

    with db_connection() as conn:
        code = _gen_ot_code(conn)
        with conn.cursor() as cur:
            if payload.plan_id:
                cur.execute("SELECT id FROM project_plans WHERE id = %s::uuid", (payload.plan_id,))
                if not cur.fetchone():
                    raise HTTPException(400, "plan_id no existe")
            if payload.partida_id:
                cur.execute("SELECT id FROM presupuesto_partidas WHERE id = %s::uuid", (payload.partida_id,))
                if not cur.fetchone():
                    raise HTTPException(400, "partida_id no existe")
            if payload.asignado_a:
                cur.execute("SELECT id FROM users WHERE id = %s::uuid", (payload.asignado_a,))
                if not cur.fetchone():
                    raise HTTPException(400, "asignado_a no existe")

            cur.execute("""
                INSERT INTO ordenes_trabajo
                    (code, plan_id, partida_id, titulo, descripcion, tipo, prioridad,
                     asignado_a, creado_por, fecha_inicio_plan, fecha_fin_plan,
                     horas_estimadas, lugar_trabajo, observaciones)
                VALUES (%s, %s::uuid, %s::uuid, %s, %s, %s, %s,
                        %s::uuid, %s::uuid, %s, %s, %s, %s, %s)
                RETURNING id
            """, (
                code,
                payload.plan_id, payload.partida_id,
                payload.titulo.strip(), payload.descripcion,
                payload.tipo, payload.prioridad,
                payload.asignado_a, str(user["id"]),
                payload.fecha_inicio_plan, payload.fecha_fin_plan,
                payload.horas_estimadas, payload.lugar_trabajo, payload.observaciones,
            ))
            ot_id = str(cur.fetchone()[0])
            conn.commit()

    return get_ot_service(ot_id)


def update_ot_service(ot_id: str, payload, user: dict) -> dict:
    fields, vals = [], []
    if payload.titulo is not None:
        fields.append("titulo = %s"); vals.append(payload.titulo.strip())
    if payload.descripcion is not None:
        fields.append("descripcion = %s"); vals.append(payload.descripcion)
    if payload.tipo is not None:
        if payload.tipo not in VALID_TIPOS:
            raise HTTPException(400, f"tipo inválido")
        fields.append("tipo = %s"); vals.append(payload.tipo)
    if payload.prioridad is not None:
        if payload.prioridad not in VALID_PRIORIDAD:
            raise HTTPException(400, f"prioridad inválida")
        fields.append("prioridad = %s"); vals.append(payload.prioridad)
    if payload.asignado_a is not None:
        fields.append("asignado_a = %s::uuid"); vals.append(payload.asignado_a)
    if payload.fecha_inicio_plan is not None:
        fields.append("fecha_inicio_plan = %s"); vals.append(payload.fecha_inicio_plan)
    if payload.fecha_fin_plan is not None:
        fields.append("fecha_fin_plan = %s"); vals.append(payload.fecha_fin_plan)
    if payload.horas_estimadas is not None:
        fields.append("horas_estimadas = %s"); vals.append(payload.horas_estimadas)
    if payload.lugar_trabajo is not None:
        fields.append("lugar_trabajo = %s"); vals.append(payload.lugar_trabajo)
    if payload.observaciones is not None:
        fields.append("observaciones = %s"); vals.append(payload.observaciones)
    if not fields:
        raise HTTPException(400, "Nada que actualizar")
    fields.append("updated_at = NOW()")
    vals.append(ot_id)

    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                f"UPDATE ordenes_trabajo SET {', '.join(fields)} WHERE id = %s::uuid RETURNING id",
                vals,
            )
            if not cur.fetchone():
                raise HTTPException(404, "OT no encontrada")
            conn.commit()

    return get_ot_service(ot_id)


# ══════════════════════════════════════════════════════════════════════════════
# CAMBIO DE ESTADO
# ══════════════════════════════════════════════════════════════════════════════

TRANSITIONS = {
    "PENDIENTE":    ["EN_EJECUCION", "CANCELADA"],
    "EN_EJECUCION": ["PAUSADA", "COMPLETADA", "CANCELADA"],
    "PAUSADA":      ["EN_EJECUCION", "CANCELADA"],
    "COMPLETADA":   ["CERRADA"],
    "CERRADA":      [],
    "CANCELADA":    [],
}


def change_status_service(ot_id: str, payload, user: dict) -> dict:
    new_status = payload.status
    if new_status not in VALID_STATUS:
        raise HTTPException(400, f"Estado inválido: {new_status}")

    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT status FROM ordenes_trabajo WHERE id = %s::uuid", (ot_id,))
            row = cur.fetchone()
            if not row:
                raise HTTPException(404, "OT no encontrada")
            current = row[0]

        allowed = TRANSITIONS.get(current, [])
        if new_status not in allowed:
            raise HTTPException(400, f"Transición no permitida: {current} → {new_status}")

        extra_fields = ""
        if new_status == "EN_EJECUCION" and current == "PENDIENTE":
            extra_fields = ", fecha_inicio_real = NOW()"
        elif new_status == "COMPLETADA":
            extra_fields = ", fecha_fin_real = NOW()"

        with conn.cursor() as cur:
            cur.execute(
                f"UPDATE ordenes_trabajo SET status = %s, updated_at = NOW() {extra_fields} "
                "WHERE id = %s::uuid",
                (new_status, ot_id),
            )
            conn.commit()

    return get_ot_service(ot_id)


# ══════════════════════════════════════════════════════════════════════════════
# CHECKLIST
# ══════════════════════════════════════════════════════════════════════════════

def add_checklist_item_service(ot_id: str, payload) -> dict:
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT id FROM ordenes_trabajo WHERE id = %s::uuid", (ot_id,))
            if not cur.fetchone():
                raise HTTPException(404, "OT no encontrada")
            cur.execute("""
                INSERT INTO ot_checklist (ot_id, descripcion, orden)
                VALUES (%s::uuid, %s, %s)
                RETURNING id, ot_id, orden, descripcion, completado, completado_por, completado_at, notas
            """, (ot_id, payload.descripcion.strip(), payload.orden))
            row = cur.fetchone()
            conn.commit()
    return _row_to_checklist((*row, None))


def toggle_checklist_service(ot_id: str, item_id: str, payload, user: dict) -> dict:
    with db_connection() as conn:
        with conn.cursor() as cur:
            if payload.completado:
                cur.execute("""
                    UPDATE ot_checklist
                    SET completado = TRUE, completado_por = %s::uuid,
                        completado_at = NOW(), notas = COALESCE(%s, notas)
                    WHERE id = %s::uuid AND ot_id = %s::uuid
                    RETURNING id, ot_id, orden, descripcion, completado,
                              completado_por, completado_at, notas
                """, (str(user["id"]), payload.notas, item_id, ot_id))
            else:
                cur.execute("""
                    UPDATE ot_checklist
                    SET completado = FALSE, completado_por = NULL,
                        completado_at = NULL
                    WHERE id = %s::uuid AND ot_id = %s::uuid
                    RETURNING id, ot_id, orden, descripcion, completado,
                              completado_por, completado_at, notas
                """, (item_id, ot_id))
            row = cur.fetchone()
            if not row:
                raise HTTPException(404, "Ítem del checklist no encontrado")

            user_name = None
            if row[5]:
                cur.execute("SELECT username FROM users WHERE id = %s", (row[5],))
                r = cur.fetchone()
                user_name = r[0] if r else None
            conn.commit()
    return _row_to_checklist((*row, user_name))


def delete_checklist_item_service(ot_id: str, item_id: str) -> dict:
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "DELETE FROM ot_checklist WHERE id = %s::uuid AND ot_id = %s::uuid RETURNING id",
                (item_id, ot_id)
            )
            if not cur.fetchone():
                raise HTTPException(404, "Ítem del checklist no encontrado")
            conn.commit()
    return {"ok": True}


# ══════════════════════════════════════════════════════════════════════════════
# MATERIALES
# ══════════════════════════════════════════════════════════════════════════════

def add_material_service(ot_id: str, payload, user: dict) -> dict:
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT status FROM ordenes_trabajo WHERE id = %s::uuid", (ot_id,))
            ot_row = cur.fetchone()
            if not ot_row:
                raise HTTPException(404, "OT no encontrada")
            if ot_row[0] in ("CERRADA", "CANCELADA"):
                raise HTTPException(400, "No se pueden agregar materiales a una OT cerrada o cancelada")

            cur.execute("SELECT id, name, unit FROM materials WHERE id = %s::uuid", (payload.material_id,))
            mat = cur.fetchone()
            if not mat:
                raise HTTPException(400, "Material no encontrado")

            if payload.almacen_id:
                cur.execute("SELECT id FROM warehouses WHERE id = %s::uuid", (payload.almacen_id,))
                if not cur.fetchone():
                    raise HTTPException(400, "Almacén no encontrado")

            cur.execute("""
                INSERT INTO ot_materiales
                    (ot_id, material_id, almacen_id, cantidad_plan, cantidad_real, registrado_por)
                VALUES (%s::uuid, %s::uuid, %s::uuid, %s, %s, %s::uuid)
                RETURNING id, ot_id, material_id, almacen_id,
                          cantidad_plan, cantidad_real, stock_movement_id, created_at
            """, (
                ot_id, payload.material_id, payload.almacen_id,
                payload.cantidad_plan, payload.cantidad_real,
                str(user["id"]),
            ))
            row = cur.fetchone()

            wh_name = None
            if payload.almacen_id:
                cur.execute("SELECT name FROM warehouses WHERE id = %s::uuid", (payload.almacen_id,))
                r = cur.fetchone()
                wh_name = r[0] if r else None

            conn.commit()

    return _row_to_material((*row, mat[1], mat[2], wh_name, None, None, None, 0.0))


def update_material_service(ot_id: str, mat_id: str, payload, user: dict) -> dict:
    fields, vals = [], []
    if payload.cantidad_real is not None:
        fields.append("cantidad_real = %s"); vals.append(payload.cantidad_real)
    if payload.almacen_id is not None:
        fields.append("almacen_id = %s::uuid"); vals.append(payload.almacen_id)
    if not fields:
        raise HTTPException(400, "Nada que actualizar")
    vals += [mat_id, ot_id]
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                f"UPDATE ot_materiales SET {', '.join(fields)} "
                "WHERE id = %s::uuid AND ot_id = %s::uuid RETURNING id",
                vals,
            )
            if not cur.fetchone():
                raise HTTPException(404, "Material de OT no encontrado")
            conn.commit()
    return get_ot_service(ot_id)


def delete_material_service(ot_id: str, mat_id: str) -> dict:
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT status FROM ordenes_trabajo WHERE id = %s::uuid", (ot_id,))
            ot_row = cur.fetchone()
            if ot_row and ot_row[0] in ("CERRADA",):
                raise HTTPException(400, "No se puede eliminar material de una OT cerrada")
            cur.execute(
                "DELETE FROM ot_materiales WHERE id = %s::uuid AND ot_id = %s::uuid RETURNING id",
                (mat_id, ot_id)
            )
            if not cur.fetchone():
                raise HTTPException(404, "Material de OT no encontrado")
            conn.commit()
    return {"ok": True}


# ══════════════════════════════════════════════════════════════════════════════
# CRONÓMETRO
# ══════════════════════════════════════════════════════════════════════════════

def iniciar_tiempo_service(ot_id: str, user: dict) -> dict:
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT status FROM ordenes_trabajo WHERE id = %s::uuid", (ot_id,))
            ot_row = cur.fetchone()
            if not ot_row:
                raise HTTPException(404, "OT no encontrada")
            if ot_row[0] not in ("EN_EJECUCION",):
                raise HTTPException(400, "La OT debe estar EN_EJECUCION para iniciar el cronómetro")

            # Verificar que no haya un intervalo abierto para este técnico
            cur.execute("""
                SELECT id FROM ot_tiempos
                WHERE ot_id = %s::uuid AND tecnico_id = %s::uuid AND fin IS NULL
            """, (ot_id, str(user["id"])))
            if cur.fetchone():
                raise HTTPException(400, "Ya tienes un cronómetro activo en esta OT")

            cur.execute("""
                INSERT INTO ot_tiempos (ot_id, tecnico_id, inicio)
                VALUES (%s::uuid, %s::uuid, NOW())
                RETURNING id, ot_id, tecnico_id, inicio, fin, horas, notas
            """, (ot_id, str(user["id"])))
            row = cur.fetchone()
            conn.commit()

    return _row_to_tiempo((*row, None))


def pausar_tiempo_service(ot_id: str, user: dict, notas: str = None) -> dict:
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT id, inicio FROM ot_tiempos
                WHERE ot_id = %s::uuid AND tecnico_id = %s::uuid AND fin IS NULL
                ORDER BY inicio DESC LIMIT 1
            """, (ot_id, str(user["id"])))
            row = cur.fetchone()
            if not row:
                raise HTTPException(400, "No tienes un cronómetro activo en esta OT")

            tiempo_id, inicio = row
            fin = datetime.now(timezone.utc)
            if inicio.tzinfo is None:
                from datetime import timezone as tz
                inicio = inicio.replace(tzinfo=timezone.utc)
            diff = fin - inicio
            horas = round(diff.total_seconds() / 3600, 4)

            cur.execute("""
                UPDATE ot_tiempos
                SET fin = NOW(), horas = %s, notas = %s
                WHERE id = %s::uuid
                RETURNING id, ot_id, tecnico_id, inicio, fin, horas, notas
            """, (horas, notas, tiempo_id))
            updated = cur.fetchone()
            conn.commit()

    return _row_to_tiempo((*updated, None))


# ══════════════════════════════════════════════════════════════════════════════
# CIERRE DE OT → descuenta stock automáticamente
# ══════════════════════════════════════════════════════════════════════════════

def cerrar_ot_service(ot_id: str, user: dict) -> dict:
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT status FROM ordenes_trabajo WHERE id = %s::uuid", (ot_id,))
            ot_row = cur.fetchone()
            if not ot_row:
                raise HTTPException(404, "OT no encontrada")
            if ot_row[0] not in ("COMPLETADA", "EN_EJECUCION"):
                raise HTTPException(400, f"Solo se puede cerrar una OT COMPLETADA o EN_EJECUCION (estado actual: {ot_row[0]})")

            # 1. Pausar cualquier cronómetro abierto
            cur.execute("""
                SELECT id, inicio FROM ot_tiempos
                WHERE ot_id = %s::uuid AND fin IS NULL
            """, (ot_id,))
            open_timers = cur.fetchall()
            for tid, inicio in open_timers:
                fin = datetime.now(timezone.utc)
                if inicio.tzinfo is None:
                    inicio = inicio.replace(tzinfo=timezone.utc)
                horas = round((fin - inicio).total_seconds() / 3600, 4)
                cur.execute(
                    "UPDATE ot_tiempos SET fin = NOW(), horas = %s WHERE id = %s",
                    (horas, tid)
                )

            # 2. Registrar movimientos de salida para cada material
            cur.execute("""
                SELECT id, material_id, almacen_id, cantidad_real
                FROM ot_materiales
                WHERE ot_id = %s::uuid AND cantidad_real > 0 AND stock_movement_id IS NULL
            """, (ot_id,))
            materiales = cur.fetchall()

            for mat_row in materiales:
                mat_ot_id, material_id, almacen_id, cantidad = mat_row
                cur.execute("""
                    INSERT INTO stock_movements
                        (material_id, movement_type, quantity, from_warehouse, reference, notes, created_by)
                    VALUES (%s::uuid, 'OUT', %s, %s::uuid, %s, %s, %s::uuid)
                    RETURNING id
                """, (
                    str(material_id),
                    float(cantidad),
                    str(almacen_id) if almacen_id else None,
                    f"OT-CIERRE-{ot_id[:8]}",
                    f"Consumo registrado al cerrar OT",
                    str(user["id"]),
                ))
                movement_id = cur.fetchone()[0]
                cur.execute(
                    "UPDATE ot_materiales SET stock_movement_id = %s WHERE id = %s",
                    (str(movement_id), mat_ot_id)
                )

            # 3. Calcular horas totales
            cur.execute(
                "SELECT COALESCE(SUM(horas), 0) FROM ot_tiempos WHERE ot_id = %s::uuid",
                (ot_id,)
            )
            horas_totales = float(cur.fetchone()[0])

            # 4. Cerrar OT
            cur.execute("""
                UPDATE ordenes_trabajo
                SET status = 'CERRADA',
                    fecha_fin_real = NOW(),
                    horas_reales = %s,
                    updated_at = NOW()
                WHERE id = %s::uuid
            """, (horas_totales, ot_id))
            conn.commit()

    return get_ot_service(ot_id)


# ══════════════════════════════════════════════════════════════════════════════
# COMPARATIVA PLAN APU vs REAL OT
# ══════════════════════════════════════════════════════════════════════════════

def get_comparativa_service(ot_id: str) -> dict:
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT ot.id, ot.code, ot.titulo, ot.horas_estimadas, ot.horas_reales,
                       ot.plan_id, ot.partida_id, ot.status
                FROM ordenes_trabajo ot
                WHERE ot.id = %s::uuid
            """, (ot_id,))
            ot_row = cur.fetchone()
            if not ot_row:
                raise HTTPException(404, "OT no encontrada")

            # Materiales plan (del APU de la partida asociada) vs real
            apu_materiales = []
            if ot_row[6]:  # partida_id
                cur.execute("""
                    SELECT ai.descripcion, ai.unidad, ai.cantidad AS plan_cantidad,
                           ai.precio_unitario, m.name AS material_nombre
                    FROM presupuesto_apu_items ai
                    LEFT JOIN materials m ON m.id = ai.material_id
                    WHERE ai.partida_id = %s::uuid AND ai.tipo_recurso IN ('MATERIAL', 'EQUIPO')
                    ORDER BY ai.tipo_recurso, ai.descripcion
                """, (str(ot_row[6]),))
                apu_materiales = [
                    {
                        "descripcion": r[0],
                        "unidad": r[1],
                        "plan_cantidad": float(r[2]),
                        "precio_unitario": float(r[3]),
                        "material_nombre": r[4],
                        "real_cantidad": 0.0,
                        "desviacion": 0.0,
                    }
                    for r in cur.fetchall()
                ]

            # Reales de OT
            cur.execute("""
                SELECT m.name, om.cantidad_real, m.unit
                FROM ot_materiales om
                JOIN materials m ON m.id = om.material_id
                WHERE om.ot_id = %s::uuid
            """, (ot_id,))
            reales = {r[0]: float(r[1]) for r in cur.fetchall()}

            # Merge plan vs real
            for item in apu_materiales:
                key = item["material_nombre"] or item["descripcion"]
                item["real_cantidad"] = reales.get(key, 0.0)
                item["desviacion"] = item["real_cantidad"] - item["plan_cantidad"]

            # Materiales reales sin APU correspondiente
            extra_reales = []
            nombres_plan = {it["material_nombre"] or it["descripcion"] for it in apu_materiales}
            cur.execute("""
                SELECT m.name, om.cantidad_real, m.unit
                FROM ot_materiales om
                JOIN materials m ON m.id = om.material_id
                WHERE om.ot_id = %s::uuid
            """, (ot_id,))
            for r in cur.fetchall():
                if r[0] not in nombres_plan:
                    extra_reales.append({
                        "descripcion": r[0],
                        "unidad": r[2],
                        "plan_cantidad": 0.0,
                        "real_cantidad": float(r[1]),
                        "desviacion": float(r[1]),
                        "material_nombre": r[0],
                        "precio_unitario": 0.0,
                    })

    return {
        "ot_id":            str(ot_row[0]),
        "ot_code":          ot_row[1],
        "titulo":           ot_row[2],
        "horas_plan":       float(ot_row[3]) if ot_row[3] else None,
        "horas_real":       float(ot_row[4]) if ot_row[4] else None,
        "horas_desviacion": (float(ot_row[4]) - float(ot_row[3])) if ot_row[3] and ot_row[4] else None,
        "materiales_plan":  apu_materiales + extra_reales,
        "tiene_partida":    ot_row[6] is not None,
    }


# ══════════════════════════════════════════════════════════════════════════════
# 📥 EXPORTAR OTs A EXCEL
# ══════════════════════════════════════════════════════════════════════════════

def export_ot_excel_service(status: str = None, tipo: str = None):
    from app.core.export_utils import (
        write_title_row, write_header_row, write_data_row,
        set_column_widths, fmt_date, fmt_num, excel_response,
    )
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    ots = list_ot_service(status=status, tipo=tipo)
    fecha = datetime.now().strftime("%d/%m/%Y %H:%M")

    STATUS_COLORS = {
        "PENDIENTE":    ("FEF3C7", "92400E"),
        "EN_EJECUCION": ("DBEAFE", "1E40AF"),
        "PAUSADA":      ("FEE2E2", "991B1B"),
        "COMPLETADA":   ("D1FAE5", "065F46"),
        "CERRADA":      ("F3F4F6", "374151"),
        "CANCELADA":    ("F3F4F6", "6B7280"),
    }
    PRIO_COLORS = {
        "URGENTE": ("FEE2E2", "DC2626"),
        "ALTA":    ("FEF3C7", "EA580C"),
        "NORMAL":  ("FFFBEB", "CA8A04"),
        "BAJA":    ("DCFCE7", "16A34A"),
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Órdenes de Trabajo"

    headers = [
        "Código OT", "Título", "Tipo", "Prioridad", "Estado",
        "Asignado a", "Plan / Proyecto",
        "Horas Est.", "Horas Reales",
        "Lugar de Trabajo",
        "Inicio Planeado", "Fin Planeado",
        "Inicio Real", "Fin Real",
        "Observaciones",
    ]
    widths = [12, 38, 12, 10, 14, 18, 16, 10, 11, 28, 14, 14, 14, 14, 32]

    write_title_row(ws, f"CeShark ERP — Órdenes de Trabajo — {fecha}", len(headers))
    write_header_row(ws, headers, row=2)
    set_column_widths(ws, widths)
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 18

    for i, ot in enumerate(ots, start=1):
        write_data_row(ws, i + 2, [
            ot.get("code", ""),
            ot.get("titulo", ""),
            ot.get("tipo", ""),
            ot.get("prioridad", ""),
            ot.get("status", ""),
            ot.get("asignado_nombre") or "",
            ot.get("plan_code") or "",
            fmt_num(ot.get("horas_estimadas"), 1) if ot.get("horas_estimadas") is not None else "",
            fmt_num(ot.get("horas_reales"), 1) if ot.get("horas_reales") is not None else "",
            ot.get("lugar_trabajo") or "",
            fmt_date(ot.get("fecha_inicio_plan")),
            fmt_date(ot.get("fecha_fin_plan")),
            fmt_date(ot.get("fecha_inicio_real")),
            fmt_date(ot.get("fecha_fin_real")),
            ot.get("observaciones") or "",
        ], alternate=(i % 2 == 0))

        data_row = i + 2
        # Color de Estado (col 5)
        status_val = ot.get("status", "")
        if status_val in STATUS_COLORS:
            bg, fg = STATUS_COLORS[status_val]
            c = ws.cell(row=data_row, column=5)
            c.fill = PatternFill("solid", fgColor=bg)
            c.font = Font(size=9, bold=True, color=fg)

        # Color de Prioridad (col 4)
        prio_val = ot.get("prioridad", "")
        if prio_val in PRIO_COLORS:
            bg, fg = PRIO_COLORS[prio_val]
            c = ws.cell(row=data_row, column=4)
            c.fill = PatternFill("solid", fgColor=bg)
            c.font = Font(size=9, bold=True, color=fg)

    return excel_response(wb, f"ordenes_trabajo_{datetime.now().strftime('%Y%m%d')}.xlsx")


# ══════════════════════════════════════════════════════════════════════════════
# GENERAR OC DESDE OT (Fase 5A)
# ══════════════════════════════════════════════════════════════════════════════

def generar_oc_service(ot_id: str, payload, user: dict) -> dict:
    with db_connection() as conn:
        with conn.cursor() as cur:
            # 1. Verificar OT
            cur.execute(
                "SELECT id, plan_id, status FROM ordenes_trabajo WHERE id = %s::uuid",
                (ot_id,)
            )
            ot = cur.fetchone()
            if not ot:
                raise HTTPException(404, "OT no encontrada")

            # 2. Verificar que mat_item pertenece a esta OT y no tiene OC ya
            cur.execute(
                "SELECT id, material_id, cantidad_real, oc_id FROM ot_materiales "
                "WHERE id = %s::uuid AND ot_id = %s::uuid",
                (payload.mat_item_id, ot_id)
            )
            mat_item = cur.fetchone()
            if not mat_item:
                raise HTTPException(404, "Material no encontrado en esta OT")
            if mat_item[3]:
                raise HTTPException(400, "Este material ya tiene una OC generada")

            # 3. Datos del material
            cur.execute("SELECT name, unit FROM materials WHERE id = %s::uuid", (payload.material_id,))
            mat = cur.fetchone()
            if not mat:
                raise HTTPException(404, "Material no encontrado")

            # 4. Buscar proveedor: si no viene en payload, buscar el principal
            proveedor_id = payload.proveedor_id
            proveedor_nombre = None
            precio_unitario = 0.0

            if not proveedor_id:
                cur.execute("""
                    SELECT mp.proveedor_id, p.nombre, mp.precio_unitario
                    FROM material_proveedores mp
                    JOIN proveedores p ON p.id = mp.proveedor_id
                    WHERE mp.material_id = %s::uuid AND mp.es_principal = TRUE AND p.activo = TRUE
                    LIMIT 1
                """, (payload.material_id,))
                prov_row = cur.fetchone()

                if not prov_row:
                    # Sin proveedor principal: devolver lista de opciones
                    cur.execute("""
                        SELECT mp.proveedor_id, p.nombre, mp.precio_unitario
                        FROM material_proveedores mp
                        JOIN proveedores p ON p.id = mp.proveedor_id
                        WHERE mp.material_id = %s::uuid AND p.activo = TRUE
                        ORDER BY p.nombre
                    """, (payload.material_id,))
                    opciones = cur.fetchall()
                    return {
                        "necesita_proveedor": True,
                        "proveedores": [
                            {"id": str(r[0]), "nombre": r[1], "precio_unitario": float(r[2])}
                            for r in opciones
                        ],
                    }

                proveedor_id = str(prov_row[0])
                proveedor_nombre = prov_row[1]
                precio_unitario = float(prov_row[2])
            else:
                cur.execute("SELECT nombre FROM proveedores WHERE id = %s::uuid", (proveedor_id,))
                p = cur.fetchone()
                if p:
                    proveedor_nombre = p[0]
                cur.execute(
                    "SELECT precio_unitario FROM material_proveedores "
                    "WHERE material_id = %s::uuid AND proveedor_id = %s::uuid",
                    (payload.material_id, proveedor_id)
                )
                pr = cur.fetchone()
                if pr:
                    precio_unitario = float(pr[0])

            # 5. Crear la OC en BORRADOR
            oc_code = generate_sequential_code(conn, "ordenes_compra", "OC")
            total = precio_unitario * payload.cantidad_faltante

            cur.execute("""
                INSERT INTO ordenes_compra
                    (code, proveedor_id, plan_id, status, solicitado_por, notas,
                     total_estimado, ot_origen_id)
                VALUES (%s, %s::uuid, %s, 'BORRADOR', %s::uuid, %s, %s, %s::uuid)
                RETURNING id, code
            """, (
                oc_code,
                proveedor_id,
                str(ot[1]) if ot[1] else None,
                str(user["id"]),
                f"Generada desde OT. Material: {mat[0]}",
                total,
                ot_id,
            ))
            oc_row = cur.fetchone()
            oc_id = str(oc_row[0])
            oc_code_real = oc_row[1]

            # 6. Agregar ítem a la OC
            cur.execute("""
                INSERT INTO ordenes_compra_items
                    (oc_id, material_id, cantidad_pedida, precio_unitario)
                VALUES (%s::uuid, %s::uuid, %s, %s)
            """, (oc_id, payload.material_id, payload.cantidad_faltante, precio_unitario))

            # 7. Vincular ot_materiales → oc_id
            cur.execute(
                "UPDATE ot_materiales SET oc_id = %s::uuid WHERE id = %s::uuid AND ot_id = %s::uuid",
                (oc_id, payload.mat_item_id, ot_id)
            )

            conn.commit()

    return {
        "necesita_proveedor": False,
        "oc_id": oc_id,
        "oc_code": oc_code_real,
        "proveedor_nombre": proveedor_nombre,
        "redirect": f"/compras/oc/{oc_id}",
    }
