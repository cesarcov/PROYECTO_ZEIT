import json
from datetime import date, datetime
from typing import Optional
from psycopg2 import sql
from app.core.database import db_connection


# ─── helpers ────────────────────────────────────────────────────────────────

def get_first_id(ids_str: Optional[str]) -> Optional[str]:
    if not ids_str:
        return None
    parts = [p.strip() for p in ids_str.split(",") if p.strip()]
    return parts[0] if parts else None

def _row_to_actividad(r):
    return {
        "id":               str(r[0]),
        "prioridad":        r[1],
        "tarea":            r[2],
        "cliente":          r[3],
        "contacto":         r[4],
        "fecha_solicitud":  r[5].isoformat() if r[5] else None,
        "responsable_id":   str(r[6]) if r[6] else None,
        "responsable":      r[7],
        "etapa":            r[8],
        "estado":           r[9],
        "fecha_limite":     r[10].isoformat() if r[10] else None,
        "seguimiento_id":   str(r[11]) if r[11] else None,
        "seguimiento":      r[12],
        "notas":            r[13],
        "progreso_pct":     float(r[14]) if r[14] is not None else 0.0,
        "created_at":       r[15].isoformat() if r[15] else None,
        "responsables_ids": r[16] if len(r) > 16 else None,
        "seguimientos_ids": r[17] if len(r) > 17 else None,
        "contactos_ids":    r[18] if len(r) > 18 else None,
        "subtareas":        [],
    }

def _row_to_subtarea(r):
    return {
        "id":           str(r[0]),
        "actividad_id": str(r[1]),
        "descripcion":  r[2],
        "culminado":    r[3],
        "created_at":   r[4].isoformat() if r[4] else None,
        "responsable_id": str(r[5]) if len(r) > 5 and r[5] else None,
        "responsable_username": r[6] if len(r) > 6 and r[6] else None,
    }

def get_subtarea_service(actividad_id: str, subtarea_id: str):
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """SELECT pst.id, pst.actividad_id, pst.descripcion, pst.culminado, pst.created_at, pst.responsable_id, u.username
                   FROM planificacion_subtareas pst
                   LEFT JOIN users u ON u.id = pst.responsable_id
                   WHERE pst.id = %s AND pst.actividad_id = %s""",
                (subtarea_id, actividad_id)
            )
            row = cur.fetchone()
            return _row_to_subtarea(row) if row else None


def _row_to_prod(r):
    return {
        "id":                    str(r[0]),
        "user_id":               str(r[1]),
        "username":              r[2],
        "fecha":                 r[3].isoformat() if r[3] else None,
        "actividad":             r[4],
        "hora_inicio":           str(r[5]) if r[5] else None,
        "hora_fin":              str(r[6]) if r[6] else None,
        "duracion_minutos":      r[7],
        "estado":                r[8],
        "actividad_semanal_id":  str(r[9]) if r[9] else None,
        "tarea_vinculada":       r[10],
        "created_at":            r[11].isoformat() if r[11] else None,
    }

_ACTIVIDAD_SELECT = """
    SELECT ps.id, ps.prioridad, ps.tarea, ps.cliente, ps.contacto,
           ps.fecha_solicitud, ps.responsable_id, ur.username,
           ps.etapa, ps.estado, ps.fecha_limite,
           ps.seguimiento_id, us.username,
           ps.notas, ps.progreso_pct, ps.created_at,
           ps.responsables_ids, ps.seguimientos_ids, ps.contactos_ids
    FROM planificacion_semanal ps
    LEFT JOIN users ur ON ur.id = ps.responsable_id
    LEFT JOIN users us ON us.id = ps.seguimiento_id
"""

def _recalc_progreso(conn, actividad_id: str):
    """Recalcula progreso_pct según subtareas culminadas y actualiza la tabla."""
    with conn.cursor() as cur:
        cur.execute(
            """SELECT COUNT(*), COUNT(*) FILTER (WHERE culminado = TRUE)
               FROM planificacion_subtareas WHERE actividad_id = %s""",
            (actividad_id,)
        )
        total, done = cur.fetchone()
        if total == 0:
            pct = 0.0
        else:
            pct = round((done / total) * 100, 2)
        cur.execute(
            "UPDATE planificacion_semanal SET progreso_pct = %s, updated_at = NOW() WHERE id = %s",
            (pct, actividad_id)
        )


# ─── CRUD Planificación Semanal ──────────────────────────────────────────────

def list_actividades_service(responsable_id=None, etapa=None, estado=None, solo_mias=False, solo_seguimiento=False, user=None):
    with db_connection() as conn:
        with conn.cursor() as cur:
            conditions = []
            params = []

            if solo_seguimiento and user:
                conditions.append("ps.seguimiento_id = %s")
                params.append(user["id"])
            elif solo_mias and user:
                conditions.append("(ps.responsable_id = %s OR EXISTS (SELECT 1 FROM planificacion_subtareas pst WHERE pst.actividad_id = ps.id AND pst.responsable_id = %s))")
                params.extend([user["id"], user["id"]])
            elif responsable_id:
                conditions.append("ps.responsable_id = %s")
                params.append(responsable_id)

            if etapa:
                conditions.append("ps.etapa ILIKE %s")
                params.append(f"%{etapa}%")
            if estado:
                conditions.append("ps.estado = %s")
                params.append(estado)

            where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
            cur.execute(
                f"{_ACTIVIDAD_SELECT} {where} ORDER BY ps.fecha_limite ASC NULLS LAST, ps.created_at DESC",
                params
            )
            actividades = [_row_to_actividad(r) for r in cur.fetchall()]

            if not actividades:
                return actividades

            # Cargar subtareas de una vez (evitar N+1)
            ids = [a["id"] for a in actividades]
            placeholders = ",".join(["%s"] * len(ids))
            cur.execute(
                f"""SELECT pst.id, pst.actividad_id, pst.descripcion, pst.culminado, pst.created_at, pst.responsable_id, u.username
                    FROM planificacion_subtareas pst
                    LEFT JOIN users u ON u.id = pst.responsable_id
                    WHERE pst.actividad_id IN ({placeholders})
                    ORDER BY pst.created_at ASC""",
                ids
            )
            subs_map: dict = {}
            for row in cur.fetchall():
                sub = _row_to_subtarea(row)
                subs_map.setdefault(sub["actividad_id"], []).append(sub)

            for a in actividades:
                a["subtareas"] = subs_map.get(a["id"], [])

    return actividades


def get_actividad_service(actividad_id: str):
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(f"{_ACTIVIDAD_SELECT} WHERE ps.id = %s", (actividad_id,))
            row = cur.fetchone()
            if not row:
                return None
            act = _row_to_actividad(row)
            cur.execute(
                """SELECT pst.id, pst.actividad_id, pst.descripcion, pst.culminado, pst.created_at, pst.responsable_id, u.username
                   FROM planificacion_subtareas pst
                   LEFT JOIN users u ON u.id = pst.responsable_id
                   WHERE pst.actividad_id = %s
                   ORDER BY pst.created_at ASC""",
                (actividad_id,)
            )
            act["subtareas"] = [_row_to_subtarea(r) for r in cur.fetchall()]
    return act


def create_actividad_service(data, user):
    resp_id = get_first_id(data.responsables_ids) or data.responsable_id or None
    seg_id = get_first_id(data.seguimientos_ids) or data.seguimiento_id or None
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """INSERT INTO planificacion_semanal
                   (prioridad, tarea, cliente, contacto, fecha_solicitud,
                    responsable_id, etapa, estado, fecha_limite, seguimiento_id,
                    responsables_ids, seguimientos_ids, notas)
                   VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                   RETURNING id""",
                (data.prioridad, data.tarea, data.cliente, data.contacto,
                 data.fecha_solicitud, resp_id, data.etapa,
                 data.estado or "En Progreso", data.fecha_limite, seg_id,
                 data.responsables_ids, data.seguimientos_ids, data.notas)
            )
            new_id = str(cur.fetchone()[0])
        conn.commit()
    return get_actividad_service(new_id)


def update_actividad_service(actividad_id: str, data, user):
    with db_connection() as conn:
        with conn.cursor() as cur:
            fields = {k: v for k, v in data.model_dump(exclude_none=True).items()}
            if not fields:
                return get_actividad_service(actividad_id)
            if "responsables_ids" in fields:
                fields["responsable_id"] = get_first_id(fields["responsables_ids"])
            if "seguimientos_ids" in fields:
                fields["seguimiento_id"] = get_first_id(fields["seguimientos_ids"])
            dyn = sql.SQL(", ").join(
                sql.SQL("{} = %s").format(sql.Identifier(k)) for k in fields
            )
            cur.execute(
                sql.SQL("UPDATE planificacion_semanal SET {}, updated_at = NOW() WHERE id = %s").format(dyn),
                list(fields.values()) + [actividad_id],
            )
        conn.commit()
    return get_actividad_service(actividad_id)


def delete_actividad_service(actividad_id: str):
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM planificacion_semanal WHERE id = %s", (actividad_id,))
        conn.commit()
    return {"ok": True}


def _resolve_contacto_label(cur, contactos_ids: Optional[str], fallback: Optional[str]) -> Optional[str]:
    if not contactos_ids:
        return fallback
    ids = [x.strip() for x in contactos_ids.split(",") if x.strip()]
    if not ids:
        return fallback
    cur.execute(
        "SELECT nombre FROM cliente_contactos WHERE id = ANY(%s::uuid[]) ORDER BY nombre",
        (ids,),
    )
    names = [r[0] for r in cur.fetchall()]
    return ", ".join(names) if names else fallback


def list_historial_service(actividad_id: str):
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """SELECT id, snapshot, guardado_por, created_at
                   FROM planificacion_historial
                   WHERE actividad_id = %s
                   ORDER BY created_at DESC""",
                (actividad_id,),
            )
            return [
                {
                    "id": str(r[0]),
                    "snapshot": r[1],
                    "guardado_por": r[2],
                    "created_at": r[3].isoformat() if r[3] else None,
                }
                for r in cur.fetchall()
            ]


def export_planificacion_excel_service(
    fecha_inicio: str = None,
    fecha_fin: str = None,
    prioridad: str = None,
    estado: str = None,
    cliente: str = None,
    responsable: str = None,
):
    from app.core.export_utils import (
        write_title_row, write_header_row, write_data_row,
        set_column_widths, excel_response, fmt_date,
    )
    import openpyxl

    acts = list_actividades_service()
    filtered = []
    for a in acts:
        if prioridad and a.get("prioridad") != prioridad:
            continue
        if estado and a.get("estado") != estado:
            continue
        if cliente and cliente.lower() not in (a.get("cliente") or "").lower():
            continue
        if responsable:
            # Mismos criterios que el filtro de responsable del tablero:
            # coincide si el id está en responsables_ids (multi) o es el responsable_id.
            resp_ids = [x.strip() for x in (a.get("responsables_ids") or "").split(",") if x.strip()]
            tiene_responsable = bool(resp_ids) or bool(a.get("responsable_id"))
            if responsable == "__none__":
                if tiene_responsable:
                    continue
            elif responsable not in resp_ids and a.get("responsable_id") != responsable:
                continue
        ftarget = a.get("fecha_limite") or a.get("fecha_solicitud")
        if fecha_inicio and (not ftarget or ftarget < fecha_inicio):
            continue
        if fecha_fin and (not ftarget or ftarget > fecha_fin):
            continue
        filtered.append(a)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Planificación"
    headers = [
        "Prioridad", "Tarea", "Cliente", "Contacto", "F. Solicitud",
        "Responsable", "Etapa", "Estado", "F. Límite", "Seguimiento", "Notas", "Progreso %",
    ]
    widths = [12, 38, 20, 22, 12, 22, 18, 14, 12, 22, 24, 10]
    titulo = f"CeShark ERP — Planificación Semanal — {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    write_title_row(ws, titulo, len(headers))
    write_header_row(ws, headers, row=2)
    set_column_widths(ws, widths)

    for i, a in enumerate(filtered, start=1):
        write_data_row(ws, i + 2, [
            a.get("prioridad") or "",
            a.get("tarea") or "",
            a.get("cliente") or "",
            a.get("contacto") or "",
            fmt_date(a.get("fecha_solicitud")),
            a.get("responsable") or "",
            a.get("etapa") or "",
            a.get("estado") or "",
            fmt_date(a.get("fecha_limite")),
            a.get("seguimiento") or "",
            a.get("notas") or "",
            f"{a.get('progreso_pct', 0):.0f}%",
        ], alternate=(i % 2 == 0))

    return excel_response(wb, f"planificacion_{datetime.now().strftime('%Y%m%d')}.xlsx")


def bulk_save_actividades_service(payload, user):
    """Guardado masivo: elimina IDs marcados e inserta/actualiza filas del grid."""
    from fastapi import HTTPException
    guardado_por = user.get("username") or "sistema"
    with db_connection() as conn:
        try:
            with conn.cursor() as cur:
                for act_id in payload.delete:
                    cur.execute("DELETE FROM planificacion_semanal WHERE id = %s", (act_id,))

                for item in payload.upsert:
                    is_new = not item.id or item.id.startswith("temp-")
                    fecha_sol = item.fecha_solicitud or None
                    fecha_lim = item.fecha_limite or None
                    resp_id = get_first_id(item.responsables_ids) or item.responsable_id or None
                    seg_id = get_first_id(item.seguimientos_ids) or item.seguimiento_id or None
                    contacto_label = _resolve_contacto_label(cur, item.contactos_ids, item.contacto)

                    if is_new:
                        cur.execute(
                            """INSERT INTO planificacion_semanal
                               (prioridad, tarea, cliente, contacto, fecha_solicitud,
                                responsable_id, etapa, estado, fecha_limite, seguimiento_id,
                                responsables_ids, seguimientos_ids, contactos_ids, notas)
                               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                            (item.prioridad, item.tarea, item.cliente, contacto_label,
                             fecha_sol, resp_id, item.etapa,
                             item.estado or "En Progreso",
                             fecha_lim, seg_id, item.responsables_ids, item.seguimientos_ids,
                             item.contactos_ids, item.notas),
                        )
                    else:
                        cur.execute(_ACTIVIDAD_SELECT + " WHERE ps.id = %s", (item.id,))
                        prev = cur.fetchone()
                        if prev:
                            snapshot = _row_to_actividad(prev)
                            cur.execute(
                                """INSERT INTO planificacion_historial
                                   (actividad_id, snapshot, guardado_por)
                                   VALUES (%s, %s, %s)""",
                                (item.id, json.dumps(snapshot), guardado_por),
                            )
                        cur.execute(
                            """UPDATE planificacion_semanal
                               SET prioridad=%s, tarea=%s, cliente=%s, contacto=%s,
                                   fecha_solicitud=%s, responsable_id=%s, etapa=%s,
                                   estado=%s, fecha_limite=%s, seguimiento_id=%s,
                                   responsables_ids=%s, seguimientos_ids=%s,
                                   contactos_ids=%s, notas=%s, updated_at=NOW()
                               WHERE id = %s""",
                            (item.prioridad, item.tarea, item.cliente, contacto_label,
                             fecha_sol, resp_id, item.etapa,
                             item.estado or "En Progreso",
                             fecha_lim, seg_id, item.responsables_ids, item.seguimientos_ids,
                             item.contactos_ids, item.notas, item.id),
                        )
            conn.commit()
        except Exception as exc:
            conn.rollback()
            raise HTTPException(status_code=500, detail=f"Error en guardado masivo: {exc}")

    return {"ok": True, "upserted": len(payload.upsert), "deleted": len(payload.delete)}


# ─── Subtareas ───────────────────────────────────────────────────────────────

def create_subtarea_service(actividad_id: str, descripcion: str):
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "INSERT INTO planificacion_subtareas (actividad_id, descripcion) VALUES (%s,%s) RETURNING id",
                (actividad_id, descripcion)
            )
            new_id = str(cur.fetchone()[0])
        _recalc_progreso(conn, actividad_id)
        conn.commit()
    return get_subtarea_service(actividad_id, new_id)


def toggle_subtarea_service(actividad_id: str, subtarea_id: str, user):
    with db_connection() as conn:
        with conn.cursor() as cur:
            # Check permissions: activity owner or subtask assignee or admin
            cur.execute(
                "SELECT responsable_id FROM planificacion_semanal WHERE id = %s",
                (actividad_id,)
            )
            act_row = cur.fetchone()
            if not act_row:
                return None
            act_resp_id = str(act_row[0]) if act_row[0] else None

            cur.execute(
                "SELECT responsable_id FROM planificacion_subtareas WHERE id = %s AND actividad_id = %s",
                (subtarea_id, actividad_id)
            )
            sub_row = cur.fetchone()
            if not sub_row:
                return None
            sub_resp_id = str(sub_row[0]) if sub_row[0] else None

            is_admin = user.get("role") == "admin"
            is_owner = act_resp_id == user["id"]
            is_assignee = sub_resp_id == user["id"]

            if not (is_admin or is_owner or is_assignee):
                from fastapi import HTTPException
                raise HTTPException(status_code=403, detail="No tienes permiso para marcar esta subtarea")

            cur.execute(
                """UPDATE planificacion_subtareas
                   SET culminado = NOT culminado, updated_at = NOW()
                   WHERE id = %s AND actividad_id = %s
                   RETURNING id""",
                (subtarea_id, actividad_id)
            )
            row = cur.fetchone()
            if not row:
                return None
        _recalc_progreso(conn, actividad_id)
        conn.commit()
    return get_subtarea_service(actividad_id, subtarea_id)


def assign_subtarea_service(actividad_id: str, subtarea_id: str, responsable_id: Optional[str], user):
    with db_connection() as conn:
        with conn.cursor() as cur:
            # Check permissions: activity owner or admin
            cur.execute(
                "SELECT responsable_id FROM planificacion_semanal WHERE id = %s",
                (actividad_id,)
            )
            act_row = cur.fetchone()
            if not act_row:
                return None
            act_resp_id = str(act_row[0]) if act_row[0] else None

            is_admin = user.get("role") == "admin"
            is_owner = act_resp_id == user["id"]

            if not (is_admin or is_owner):
                from fastapi import HTTPException
                raise HTTPException(status_code=403, detail="Solo el encargado de la actividad o un administrador puede asignar subtareas")

            cur.execute(
                """UPDATE planificacion_subtareas
                   SET responsable_id = %s, updated_at = NOW()
                   WHERE id = %s AND actividad_id = %s
                   RETURNING id""",
                (responsable_id or None, subtarea_id, actividad_id)
            )
            row = cur.fetchone()
            if not row:
                return None
        conn.commit()
    return get_subtarea_service(actividad_id, subtarea_id)



def delete_subtarea_service(actividad_id: str, subtarea_id: str):
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "DELETE FROM planificacion_subtareas WHERE id = %s AND actividad_id = %s",
                (subtarea_id, actividad_id)
            )
        _recalc_progreso(conn, actividad_id)
        conn.commit()
    return {"ok": True}


# ─── Registro de Productividad ───────────────────────────────────────────────

def create_productividad_service(data, user):
    inicio = data.hora_inicio
    fin = data.hora_fin
    # Calcular duración en minutos
    minutos = (fin.hour * 60 + fin.minute) - (inicio.hour * 60 + inicio.minute)
    if minutos < 0:
        minutos = 0

    fecha = data.fecha or date.today()

    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """INSERT INTO registro_productividad
                   (user_id, fecha, actividad, hora_inicio, hora_fin, duracion_minutos, actividad_semanal_id)
                   VALUES (%s,%s,%s,%s,%s,%s,%s) RETURNING id""",
                (user["user_id"], fecha, data.actividad,
                 inicio, fin, minutos, data.actividad_semanal_id)
            )
            new_id = str(cur.fetchone()[0])
        conn.commit()
    return {"id": new_id, "duracion_minutos": minutos}


def list_productividad_service(user_id: str, fecha: Optional[date] = None):
    """Logs del usuario para una fecha (default hoy)."""
    fecha = fecha or date.today()
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """SELECT rp.id, rp.user_id, u.username, rp.fecha,
                          rp.actividad, rp.hora_inicio, rp.hora_fin,
                          rp.duracion_minutos, rp.estado, rp.actividad_semanal_id,
                          ps.tarea, rp.created_at
                   FROM registro_productividad rp
                   JOIN users u ON u.id = rp.user_id
                   LEFT JOIN planificacion_semanal ps ON ps.id = rp.actividad_semanal_id
                   WHERE rp.user_id = %s AND rp.fecha = %s
                   ORDER BY rp.hora_inicio ASC""",
                (user_id, fecha)
            )
            return [_row_to_prod(r) for r in cur.fetchall()]


def delete_productividad_service(log_id: str, user_id: str):
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "DELETE FROM registro_productividad WHERE id = %s AND user_id = %s",
                (log_id, user_id)
            )
        conn.commit()
    return {"ok": True}


def list_productividad_admin_service(user_id_filter: Optional[str] = None, fecha: Optional[date] = None, mes: Optional[str] = None):
    """Logs agrupados para el panel de administración."""
    with db_connection() as conn:
        with conn.cursor() as cur:
            conditions = []
            params = []
            if user_id_filter:
                conditions.append("rp.user_id = %s")
                params.append(user_id_filter)
            if fecha:
                conditions.append("rp.fecha = %s")
                params.append(fecha)
            elif mes:
                conditions.append("TO_CHAR(rp.fecha, 'YYYY-MM') = %s")
                params.append(mes)
            where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
            cur.execute(
                f"""SELECT rp.id, rp.user_id, u.username, rp.fecha,
                           rp.actividad, rp.hora_inicio, rp.hora_fin,
                           rp.duracion_minutos, rp.estado, rp.actividad_semanal_id,
                           ps.tarea, rp.created_at
                    FROM registro_productividad rp
                    JOIN users u ON u.id = rp.user_id
                    LEFT JOIN planificacion_semanal ps ON ps.id = rp.actividad_semanal_id
                    {where}
                    ORDER BY rp.fecha DESC, u.username, rp.hora_inicio ASC""",
                params
            )
            return [_row_to_prod(r) for r in cur.fetchall()]


def export_productividad_excel_service(user_id_filter: Optional[str] = None, fecha: Optional[date] = None, mes: Optional[str] = None):
    from app.core.export_utils import (
        write_title_row, write_header_row, write_data_row, write_total_row,
        set_column_widths, fmt_num, excel_response,
    )
    from app.core.database import db_connection
    import openpyxl
    import re
    from datetime import datetime

    # 1. Obtener los logs de productividad
    logs = list_productividad_admin_service(user_id_filter, fecha, mes)

    # 2. Obtener info del usuario si hay filtro para personalizar el título
    username_filter = None
    if user_id_filter:
        with db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT username FROM users WHERE id = %s", (user_id_filter,))
                row = cur.fetchone()
                if row:
                    username_filter = row[0]

    # 3. Formatear título del reporte
    report_title = "CeShark ERP — Reporte de Productividad"
    if username_filter:
        formatted_uname = username_filter.replace("_", " ").title()
        if username_filter.lower() == "admin":
            formatted_uname = "TI - Ceshark"
        report_title += f" ({formatted_uname})"
    if fecha:
        report_title += f" — Fecha: {fecha.strftime('%d/%m/%Y')}"
    elif mes:
        try:
            year_part, month_part = mes.split("-")
            month_names = {
                "01": "Enero", "02": "Febrero", "03": "Marzo", "04": "Abril",
                "05": "Mayo", "06": "Junio", "07": "Julio", "08": "Agosto",
                "09": "Septiembre", "10": "Octubre", "11": "Noviembre", "12": "Diciembre"
            }
            report_title += f" — Mes: {month_names.get(month_part, month_part)} {year_part}"
        except Exception:
            report_title += f" — Mes: {mes}"
    else:
        report_title += f" — Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Logs de Actividades"

    headers = [
        "Usuario", "Fecha", "Actividad", "Hora Inicio", "Hora Fin", "Duración (minutos)", "Duración (horas)", "Tarea Vinculada"
    ]
    widths = [20, 14, 45, 13, 13, 20, 18, 30]

    write_title_row(ws, report_title, len(headers))
    write_header_row(ws, headers, row=2)
    set_column_widths(ws, widths)
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 18

    total_minutos = 0
    for i, log in enumerate(logs, start=1):
        username = log.get("username", "")
        formatted_username = username.replace("_", " ").title()
        if username.lower() == "admin":
            formatted_username = "TI - Ceshark"
            
        duracion_min = log.get("duracion_minutos") or 0
        total_minutos += duracion_min
        
        # Horas formato decimal
        duracion_h = duracion_min / 60.0
        
        # Limpiar prefijo redundante [CUALQUIERA] - de la descripción
        actividad = log.get("actividad", "")
        if actividad:
            actividad = re.sub(r"^\[[^\]]+\]\s*[-:]?\s*", "", actividad)
        
        write_data_row(ws, i + 2, [
            formatted_username,
            log.get("fecha") or "",
            actividad,
            log.get("hora_inicio")[:5] if log.get("hora_inicio") else "",
            log.get("hora_fin")[:5] if log.get("hora_fin") else "",
            fmt_num(duracion_min, 0),
            fmt_num(duracion_h, 2),
            log.get("tarea_vinculada") or "—",
        ], alternate=(i % 2 == 0))

    # Fila de totales
    total_row_idx = len(logs) + 3
    total_val = ["TOTAL ACUMULADO", "", "", "", "", fmt_num(total_minutos, 0), fmt_num(total_minutos/60.0, 2), ""]
    write_total_row(ws, total_row_idx, total_val, len(headers))
    ws.row_dimensions[total_row_idx].height = 20

    filename_suffix = username_filter if username_filter else "equipo"
    return excel_response(wb, f"productividad_{filename_suffix}_{datetime.now().strftime('%Y%m%d')}.xlsx")


def start_productividad_service(data, user):
    now = datetime.now()
    fecha = now.date()
    hora_inicio = now.time()

    with db_connection() as conn:
        with conn.cursor() as cur:
            # 1. Buscar si tiene temporizador activo
            cur.execute(
                """SELECT id, fecha, hora_inicio FROM registro_productividad 
                   WHERE user_id = %s AND estado = 'A'""",
                (user["id"],)
            )
            active = cur.fetchone()
            if active:
                # Detener temporizador activo anterior
                active_id, active_fecha, active_hora_inicio = active
                start_dt = datetime.combine(active_fecha, active_hora_inicio)
                elapsed = now - start_dt
                minutos = max(1, int(elapsed.total_seconds() / 60))
                hora_fin = now.time()
                cur.execute(
                    """UPDATE registro_productividad 
                       SET hora_fin = %s, duracion_minutos = %s, estado = 'F' 
                       WHERE id = %s""",
                    (hora_fin, minutos, active_id)
                )
            
            # 2. Iniciar el nuevo temporizador
            cur.execute(
                """INSERT INTO registro_productividad
                   (user_id, fecha, actividad, hora_inicio, hora_fin, duracion_minutos, estado, actividad_semanal_id)
                   VALUES (%s, %s, %s, %s, NULL, NULL, 'A', %s) RETURNING id""",
                (user["id"], fecha, data.actividad, hora_inicio, data.actividad_semanal_id or None)
            )
            new_id = str(cur.fetchone()[0])
        conn.commit()
    
    return {"id": new_id, "estado": "A", "hora_inicio": str(hora_inicio)}


def stop_productividad_service(user):
    now = datetime.now()
    hora_fin = now.time()
    
    with db_connection() as conn:
        with conn.cursor() as cur:
            # Buscar el activo
            cur.execute(
                """SELECT id, fecha, hora_inicio, actividad, actividad_semanal_id 
                   FROM registro_productividad 
                   WHERE user_id = %s AND estado = 'A'""",
                (user["id"],)
            )
            active = cur.fetchone()
            if not active:
                return None
            
            active_id, active_fecha, active_hora_inicio, actividad, act_semanal_id = active
            start_dt = datetime.combine(active_fecha, active_hora_inicio)
            elapsed = now - start_dt
            minutos = max(1, int(elapsed.total_seconds() / 60))
            
            cur.execute(
                """UPDATE registro_productividad 
                   SET hora_fin = %s, duracion_minutos = %s, estado = 'F' 
                   WHERE id = %s""",
                (hora_fin, minutos, active_id)
            )
        conn.commit()
    
    return {"id": active_id, "estado": "F", "duracion_minutos": minutos, "hora_fin": str(hora_fin)}


def get_active_productividad_service(user):
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """SELECT rp.id, rp.fecha, rp.actividad, rp.hora_inicio, rp.actividad_semanal_id, ps.tarea
                   FROM registro_productividad rp
                   LEFT JOIN planificacion_semanal ps ON ps.id = rp.actividad_semanal_id
                   WHERE rp.user_id = %s AND rp.estado = 'A'""",
                (user["id"],)
            )
            row = cur.fetchone()
            if not row:
                return None
            
            return {
                "id": str(row[0]),
                "fecha": row[1].isoformat(),
                "actividad": row[2],
                "hora_inicio": str(row[3]),
                "actividad_semanal_id": str(row[4]) if row[4] else None,
                "tarea_vinculada": row[5]
            }


# ─── KPIs / Métricas ─────────────────────────────────────────────────────────

def get_kpis_productividad_service():
    """Retorna los KPIs simplificados de planificación: ratio finalización, tareas retrasadas,
       distribución de tareas por usuario (carga) y tareas pendientes por cliente/etapa."""
    with db_connection() as conn:
        with conn.cursor() as cur:
            # 1. Ratio de finalización de tareas
            cur.execute(
                """SELECT
                     COUNT(*) FILTER (WHERE estado = 'Completado') as completadas,
                     COUNT(*) as total
                   FROM planificacion_semanal"""
            )
            row = cur.fetchone()
            completadas, total = row[0] or 0, row[1] or 0
            ratio = round((completadas / total * 100), 1) if total > 0 else 0

            # 2. Tareas retrasadas
            cur.execute(
                """SELECT COUNT(*) FROM planificacion_semanal
                   WHERE fecha_limite < CURRENT_DATE AND estado != 'Completado'"""
            )
            retrasadas = cur.fetchone()[0] or 0

            # 3. Distribución de tareas por usuario
            cur.execute(
                """SELECT ps.responsable_id, u.username, ps.estado, COUNT(*)
                   FROM planificacion_semanal ps
                   JOIN users u ON u.id = ps.responsable_id
                   GROUP BY ps.responsable_id, u.username, ps.estado"""
            )
            dist_usuario = {}
            for r_id, username, estado, count in cur.fetchall():
                if not r_id:
                    continue
                uid = str(r_id)
                if uid not in dist_usuario:
                    dist_usuario[uid] = {
                        "username": username,
                        "Completado": 0,
                        "Retraso": 0,
                        "En espera": 0,
                        "En Progreso": 0
                    }
                if estado in dist_usuario[uid]:
                    dist_usuario[uid][estado] = count

            # 4. Tareas pendientes por cliente y etapa
            cur.execute("SELECT id, razon_social, codigo FROM clientes WHERE activo = TRUE")
            clientes_db = cur.fetchall()

            cur.execute(
                """SELECT ps.cliente, ps.etapa, COUNT(*)
                   FROM planificacion_semanal ps
                   WHERE ps.estado != 'Completado' AND ps.cliente IS NOT NULL AND ps.cliente != ''
                   GROUP BY ps.cliente, ps.etapa
                   ORDER BY ps.cliente, ps.etapa"""
            )
            dist_cliente = {}
            for client_str, etapa, count in cur.fetchall():
                if not client_str:
                    continue
                c_key = client_str.strip().upper()
                if not c_key:
                    continue
                if c_key not in dist_cliente:
                    cliente_id = None
                    for c_id, c_razon, c_code in clientes_db:
                        if c_key in c_razon.upper() or c_key in c_code.upper():
                            cliente_id = str(c_id)
                            break
                    dist_cliente[c_key] = {
                        "cliente": client_str,
                        "cliente_id": cliente_id,
                        "pendientes_por_etapa": {}
                    }
                etapa_name = etapa.strip().upper() if etapa else "OTROS"
                dist_cliente[c_key]["pendientes_por_etapa"][etapa_name] = count

    return {
        "ratio_finalizacion_pct": ratio,
        "tareas_completadas": completadas,
        "tareas_total": total,
        "tareas_retrasadas": retrasadas,
        "tareas_por_usuario": list(dist_usuario.values()),
        "pendientes_por_cliente": list(dist_cliente.values())
    }



# ─── Importador de Excel Semanal ─────────────────────────────────────────────

def import_planificacion_excel_service(file_bytes: bytes):
    """Lee hojas SEM* del Excel y crea tareas en planificacion_semanal."""
    try:
        import openpyxl
        from io import BytesIO
    except ImportError:
        return {"error": "openpyxl no instalado"}

    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    sheets_sem = [s for s in wb.sheetnames if s.upper().startswith("SEM")]

    if not sheets_sem:
        return {"error": "No se encontraron hojas que empiecen con 'SEM'"}

    # Cargar usuarios una sola vez para el mapeo
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT id, username FROM users WHERE is_active = TRUE")
            users_db = [(str(r[0]), r[1].lower()) for r in cur.fetchall()]

    def _find_user(name_excel: str):
        if not name_excel:
            return None
        name_lower = str(name_excel).lower().strip()
        for uid, uname in users_db:
            if uname in name_lower or name_lower in uname:
                return uid
        return None

    inserted = 0
    errors = []

    with db_connection() as conn:
        for sheet_name in sheets_sem:
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=2, values_only=True):
                # Columnas esperadas: ITEM, PRIORIDAD, TAREA, CLIENTE, CONTACTO,
                # FECHA SOLICITUD, RESPONSABLE COTIZACIÓN, ETAPA, ESTADO,
                # FECHA LÍMITE, RESPONSABLE SEGUIMIENTO, NOTAS
                if not row or not row[2]:  # tarea vacía → skip
                    continue
                try:
                    prioridad   = str(row[1]).strip() if row[1] else None
                    tarea       = str(row[2]).strip()
                    cliente     = str(row[3]).strip() if row[3] else None
                    contacto    = str(row[4]).strip() if row[4] else None

                    fecha_sol = None
                    if row[5]:
                        if hasattr(row[5], "date"):
                            fecha_sol = row[5].date()
                        else:
                            try:
                                from datetime import datetime as dt
                                fecha_sol = dt.strptime(str(row[5]), "%Y-%m-%d").date()
                            except Exception:
                                fecha_sol = None

                    responsable_id = _find_user(row[6])
                    etapa          = str(row[7]).strip() if row[7] else None
                    estado         = str(row[8]).strip() if row[8] else "En Progreso"

                    fecha_lim = None
                    if row[9]:
                        if hasattr(row[9], "date"):
                            fecha_lim = row[9].date()
                        else:
                            try:
                                from datetime import datetime as dt
                                fecha_lim = dt.strptime(str(row[9]), "%Y-%m-%d").date()
                            except Exception:
                                fecha_lim = None

                    seguimiento_id = _find_user(row[10])
                    notas          = str(row[11]).strip() if len(row) > 11 and row[11] else None

                    with conn.cursor() as cur:
                        cur.execute(
                            """INSERT INTO planificacion_semanal
                               (prioridad, tarea, cliente, contacto, fecha_solicitud,
                                responsable_id, etapa, estado, fecha_limite, seguimiento_id, notas)
                               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                            (prioridad, tarea, cliente, contacto, fecha_sol,
                             responsable_id, etapa, estado, fecha_lim, seguimiento_id, notas)
                        )
                    inserted += 1
                except Exception as e:
                    errors.append(str(e))

        conn.commit()

    return {"inserted": inserted, "errors": errors, "sheets_procesadas": sheets_sem}


def list_active_users_service():
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT id, username FROM users WHERE is_active = TRUE ORDER BY username ASC")
            return [{"id": str(r[0]), "username": r[1]} for r in cur.fetchall()]


def count_my_pending_tasks_service(user: dict) -> dict:
    user_id = str(user["id"])
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT COUNT(*) FROM planificacion_semanal
                WHERE estado != 'Completado'
                  AND (responsable_id = %s OR responsables_ids LIKE %s)
            """, (user_id, f"%{user_id}%"))
            count = cur.fetchone()[0]
    return {"count": int(count)}

