from fastapi import HTTPException
from app.core.database import db_connection
from app.core.utils import generate_sequential_code
from app.modules.compras.schemas import TRANSICIONES_OC


# ══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def _gen_prov_code(conn) -> str:
    return generate_sequential_code(conn, "proveedores", "PROV", year_based=False, pad=3)


def _gen_oc_code(conn) -> str:
    return generate_sequential_code(conn, "ordenes_compra", "OC")


def _row_to_proveedor(r) -> dict:
    return {
        "id": str(r[0]), "codigo": r[1], "nombre": r[2], "ruc": r[3],
        "direccion": r[4], "telefono": r[5], "email": r[6], "contacto": r[7],
        "tipo": r[8], "activo": r[9], "created_at": r[10],
    }


def _row_to_oc(r) -> dict:
    return {
        "id": str(r[0]), "code": r[1],
        "proveedor_id": str(r[2]) if r[2] else None,
        "proveedor_nombre": r[3],
        "plan_id": str(r[4]) if r[4] else None,
        "status": r[5],
        "solicitado_por": str(r[6]) if r[6] else None,
        "aprobado_por": str(r[7]) if r[7] else None,
        "almacen_destino": str(r[8]) if r[8] else None,
        "almacen_nombre": r[9],
        "fecha_solicitud": r[10],
        "fecha_entrega_est": r[11],
        "fecha_recepcion": r[12],
        "notas": r[13],
        "total_estimado": float(r[14]) if r[14] is not None else None,
        "total_real": float(r[15]) if r[15] is not None else None,
        "created_at": r[16],
        "updated_at": r[17],
    }


def _row_to_oc_item(r) -> dict:
    return {
        "id": str(r[0]),
        "oc_id": str(r[1]),
        "material_id": str(r[2]) if r[2] else None,
        "material_nombre": r[3],
        "material_unidad": r[4],
        "cantidad_pedida": float(r[5]),
        "precio_unitario": float(r[6]),
        "subtotal": float(r[5]) * float(r[6]),
        "cantidad_recibida": float(r[7]),
        "pendiente": float(r[5]) - float(r[7]),
        "stock_movement_id": str(r[8]) if r[8] else None,
        "notas": r[9],
    }


def _recalc_total(conn, oc_id: str):
    with conn.cursor() as cur:
        cur.execute("""
            UPDATE ordenes_compra
            SET total_estimado = (
                SELECT COALESCE(SUM(cantidad_pedida * precio_unitario), 0)
                FROM ordenes_compra_items WHERE oc_id = %s
            ),
            updated_at = NOW()
            WHERE id = %s
        """, (oc_id, oc_id))


# ══════════════════════════════════════════════════════════════════════════════
# PROVEEDORES
# ══════════════════════════════════════════════════════════════════════════════

def list_proveedores_service(activo: bool = None) -> list:
    with db_connection() as conn:
        with conn.cursor() as cur:
            sql = """
                SELECT id, codigo, nombre, ruc, direccion, telefono, email, contacto,
                       tipo, activo, created_at
                FROM proveedores
            """
            params = []
            if activo is not None:
                sql += " WHERE activo = %s"
                params.append(activo)
            sql += " ORDER BY nombre"
            cur.execute(sql, params)
            rows = cur.fetchall()
    return [_row_to_proveedor(r) for r in rows]


def create_proveedor_service(payload) -> dict:
    with db_connection() as conn:
        codigo = _gen_prov_code(conn)
        with conn.cursor() as cur:
            if payload.ruc:
                cur.execute("SELECT id FROM proveedores WHERE ruc = %s", (payload.ruc,))
                if cur.fetchone():
                    raise HTTPException(400, f"Ya existe un proveedor con RUC {payload.ruc}")
            cur.execute("""
                INSERT INTO proveedores
                    (codigo, nombre, ruc, direccion, telefono, email, contacto, tipo)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id, codigo, nombre, ruc, direccion, telefono, email, contacto,
                          tipo, activo, created_at
            """, (
                codigo,
                payload.nombre.strip(),
                payload.ruc.strip() if payload.ruc else None,
                payload.direccion,
                payload.telefono,
                payload.email,
                payload.contacto,
                payload.tipo,
            ))
            row = cur.fetchone()
            conn.commit()
    return _row_to_proveedor(row)


def update_proveedor_service(proveedor_id: str, payload) -> dict:
    fields, vals = [], []
    mapping = [
        ("nombre", payload.nombre),
        ("ruc", payload.ruc),
        ("direccion", payload.direccion),
        ("telefono", payload.telefono),
        ("email", payload.email),
        ("contacto", payload.contacto),
        ("tipo", payload.tipo),
        ("activo", payload.activo),
    ]
    for col, val in mapping:
        if val is not None:
            fields.append(f"{col} = %s")
            vals.append(val.strip() if isinstance(val, str) else val)
    if not fields:
        raise HTTPException(400, "Nada que actualizar")
    vals.append(proveedor_id)
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                f"UPDATE proveedores SET {', '.join(fields)} WHERE id = %s "
                "RETURNING id, codigo, nombre, ruc, direccion, telefono, email, contacto, tipo, activo, created_at",
                vals,
            )
            row = cur.fetchone()
            if not row:
                raise HTTPException(404, "Proveedor no encontrado")
            conn.commit()
    return _row_to_proveedor(row)


def get_proveedor_materiales_service(proveedor_id: str) -> list:
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT id FROM proveedores WHERE id = %s", (proveedor_id,))
            if not cur.fetchone():
                raise HTTPException(404, "Proveedor no encontrado")
            cur.execute("""
                SELECT mp.id, mp.material_id, m.name AS material_nombre, m.unit,
                       mp.precio_unitario, mp.moneda, mp.tiempo_entrega_dias, mp.es_principal, mp.updated_at
                FROM material_proveedores mp
                JOIN materials m ON m.id = mp.material_id
                WHERE mp.proveedor_id = %s
                ORDER BY m.name
            """, (proveedor_id,))
            rows = cur.fetchall()
    return [
        {
            "id": str(r[0]), "material_id": str(r[1]),
            "material_nombre": r[2], "material_unidad": r[3],
            "precio_unitario": float(r[4]), "moneda": r[5],
            "tiempo_entrega_dias": r[6], "es_principal": r[7], "updated_at": r[8],
        }
        for r in rows
    ]


# ══════════════════════════════════════════════════════════════════════════════
# CATÁLOGO MATERIAL-PROVEEDOR
# ══════════════════════════════════════════════════════════════════════════════

def add_material_proveedor_service(material_id: str, payload) -> dict:
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT id FROM materials WHERE id = %s", (material_id,))
            if not cur.fetchone():
                raise HTTPException(404, "Material no encontrado")
            cur.execute("SELECT id FROM proveedores WHERE id = %s", (payload.proveedor_id,))
            if not cur.fetchone():
                raise HTTPException(404, "Proveedor no encontrado")

            if payload.es_principal:
                cur.execute(
                    "UPDATE material_proveedores SET es_principal = FALSE WHERE material_id = %s",
                    (material_id,)
                )

            cur.execute("""
                INSERT INTO material_proveedores
                    (material_id, proveedor_id, precio_unitario, moneda, tiempo_entrega_dias, es_principal)
                VALUES (%s, %s, %s, %s, %s, %s)
                ON CONFLICT (material_id, proveedor_id)
                DO UPDATE SET
                    precio_unitario = EXCLUDED.precio_unitario,
                    moneda = EXCLUDED.moneda,
                    tiempo_entrega_dias = EXCLUDED.tiempo_entrega_dias,
                    es_principal = EXCLUDED.es_principal,
                    updated_at = NOW()
                RETURNING id, material_id, proveedor_id, precio_unitario, moneda,
                          tiempo_entrega_dias, es_principal, updated_at
            """, (
                material_id, payload.proveedor_id, payload.precio_unitario,
                payload.moneda, payload.tiempo_entrega_dias, payload.es_principal,
            ))
            row = cur.fetchone()
            conn.commit()
    return {
        "id": str(row[0]), "material_id": str(row[1]), "proveedor_id": str(row[2]),
        "precio_unitario": float(row[3]), "moneda": row[4],
        "tiempo_entrega_dias": row[5], "es_principal": row[6], "updated_at": row[7],
    }


def delete_material_proveedor_service(material_id: str, mp_id: str) -> dict:
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "DELETE FROM material_proveedores WHERE id = %s AND material_id = %s RETURNING id",
                (mp_id, material_id)
            )
            if not cur.fetchone():
                raise HTTPException(404, "Vínculo no encontrado")
            conn.commit()
    return {"ok": True}


def list_material_proveedores_service(material_id: str) -> list:
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT mp.id, mp.proveedor_id, p.nombre AS proveedor_nombre, p.codigo,
                       mp.precio_unitario, mp.moneda, mp.tiempo_entrega_dias, mp.es_principal, mp.updated_at
                FROM material_proveedores mp
                JOIN proveedores p ON p.id = mp.proveedor_id
                WHERE mp.material_id = %s
                ORDER BY mp.es_principal DESC, p.nombre
            """, (material_id,))
            rows = cur.fetchall()
    return [
        {
            "id": str(r[0]), "proveedor_id": str(r[1]),
            "proveedor_nombre": r[2], "proveedor_codigo": r[3],
            "precio_unitario": float(r[4]), "moneda": r[5],
            "tiempo_entrega_dias": r[6], "es_principal": r[7], "updated_at": r[8],
        }
        for r in rows
    ]


# ══════════════════════════════════════════════════════════════════════════════
# ÓRDENES DE COMPRA
# ══════════════════════════════════════════════════════════════════════════════

def list_oc_service(status=None, proveedor_id=None, plan_id=None, solicitado_por=None) -> list:
    with db_connection() as conn:
        with conn.cursor() as cur:
            sql = """
                SELECT oc.id, oc.code, oc.proveedor_id, p.nombre AS proveedor_nombre,
                       oc.plan_id, oc.status, oc.solicitado_por, oc.aprobado_por,
                       oc.almacen_destino, w.name AS almacen_nombre,
                       oc.fecha_solicitud, oc.fecha_entrega_est, oc.fecha_recepcion,
                       oc.notas, oc.total_estimado, oc.total_real,
                       oc.created_at, oc.updated_at
                FROM ordenes_compra oc
                JOIN proveedores p ON p.id = oc.proveedor_id
                LEFT JOIN warehouses w ON w.id = oc.almacen_destino
            """
            filters, params = [], []
            if status:
                filters.append("oc.status = %s"); params.append(status)
            if proveedor_id:
                filters.append("oc.proveedor_id = %s"); params.append(proveedor_id)
            if plan_id:
                filters.append("oc.plan_id = %s"); params.append(plan_id)
            if solicitado_por:
                filters.append("oc.solicitado_por = %s"); params.append(solicitado_por)
            if filters:
                sql += " WHERE " + " AND ".join(filters)
            sql += " ORDER BY oc.created_at DESC"
            cur.execute(sql, params)
            rows = cur.fetchall()
    return [_row_to_oc(r) for r in rows]


def create_oc_service(payload, user) -> dict:
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT id FROM proveedores WHERE id = %s AND activo = TRUE", (payload.proveedor_id,))
            if not cur.fetchone():
                raise HTTPException(404, "Proveedor no encontrado o inactivo")

            code = _gen_oc_code(conn)

            cur.execute("""
                INSERT INTO ordenes_compra
                    (code, proveedor_id, plan_id, almacen_destino, fecha_entrega_est,
                     notas, solicitado_por, total_estimado)
                VALUES (%s, %s, %s, %s, %s, %s, %s, 0)
                RETURNING id
            """, (
                code,
                payload.proveedor_id,
                payload.plan_id or None,
                payload.almacen_destino or None,
                payload.fecha_entrega_est or None,
                payload.notas,
                user["id"],
            ))
            oc_id = str(cur.fetchone()[0])

            for item in payload.items:
                cur.execute("SELECT id FROM materials WHERE id = %s", (item.material_id,))
                if not cur.fetchone():
                    raise HTTPException(400, f"Material {item.material_id} no existe")
                cur.execute("""
                    INSERT INTO ordenes_compra_items
                        (oc_id, material_id, cantidad_pedida, precio_unitario, notas)
                    VALUES (%s, %s, %s, %s, %s)
                """, (oc_id, item.material_id, item.cantidad_pedida, item.precio_unitario, item.notas))

            _recalc_total(conn, oc_id)
            conn.commit()

    return get_oc_service(oc_id)


def get_oc_service(oc_id: str) -> dict:
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT oc.id, oc.code, oc.proveedor_id, p.nombre,
                       oc.plan_id, oc.status, oc.solicitado_por, oc.aprobado_por,
                       oc.almacen_destino, w.name,
                       oc.fecha_solicitud, oc.fecha_entrega_est, oc.fecha_recepcion,
                       oc.notas, oc.total_estimado, oc.total_real,
                       oc.created_at, oc.updated_at
                FROM ordenes_compra oc
                JOIN proveedores p ON p.id = oc.proveedor_id
                LEFT JOIN warehouses w ON w.id = oc.almacen_destino
                WHERE oc.id = %s
            """, (oc_id,))
            row = cur.fetchone()
            if not row:
                raise HTTPException(404, "Orden de Compra no encontrada")
            oc = _row_to_oc(row)

            cur.execute("""
                SELECT oci.id, oci.oc_id, oci.material_id, m.name, m.unit,
                       oci.cantidad_pedida, oci.precio_unitario,
                       oci.cantidad_recibida, oci.stock_movement_id, oci.notas
                FROM ordenes_compra_items oci
                JOIN materials m ON m.id = oci.material_id
                WHERE oci.oc_id = %s
                ORDER BY m.name
            """, (oc_id,))
            oc["items"] = [_row_to_oc_item(r) for r in cur.fetchall()]

    return oc


def update_oc_service(oc_id: str, payload) -> dict:
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT status FROM ordenes_compra WHERE id = %s", (oc_id,))
            row = cur.fetchone()
            if not row:
                raise HTTPException(404, "OC no encontrada")
            if row[0] not in ("BORRADOR",):
                raise HTTPException(400, "Solo se puede editar una OC en estado BORRADOR")

        fields, vals = [], []
        mapping = [
            ("proveedor_id", payload.proveedor_id),
            ("plan_id", payload.plan_id),
            ("almacen_destino", payload.almacen_destino),
            ("fecha_entrega_est", payload.fecha_entrega_est),
            ("notas", payload.notas),
        ]
        for col, val in mapping:
            if val is not None:
                fields.append(f"{col} = %s"); vals.append(val)
        if not fields:
            raise HTTPException(400, "Nada que actualizar")
        fields.append("updated_at = NOW()")
        vals.append(oc_id)
        with conn.cursor() as cur:
            cur.execute(
                f"UPDATE ordenes_compra SET {', '.join(fields)} WHERE id = %s",
                vals,
            )
            conn.commit()

    return get_oc_service(oc_id)


def change_oc_status_service(oc_id: str, payload, user) -> dict:
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT status FROM ordenes_compra WHERE id = %s", (oc_id,))
            row = cur.fetchone()
            if not row:
                raise HTTPException(404, "OC no encontrada")
            current = row[0]

        new_status = payload.status
        if new_status not in TRANSICIONES_OC.get(current, []):
            raise HTTPException(400, f"Transición inválida: {current} → {new_status}")

        extra_fields = ""
        extra_vals = []
        if new_status == "APROBADA":
            extra_fields = ", aprobado_por = %s"
            extra_vals = [user["id"]]
        elif new_status in ("RECIBIDA", "CERRADA"):
            extra_fields = ", fecha_recepcion = NOW()"

        with conn.cursor() as cur:
            cur.execute(
                f"UPDATE ordenes_compra SET status = %s{extra_fields}, updated_at = NOW() WHERE id = %s",
                [new_status] + extra_vals + [oc_id],
            )
            conn.commit()

    return get_oc_service(oc_id)


# ── Ítems de OC ───────────────────────────────────────────────────────────────

def add_oc_item_service(oc_id: str, payload) -> dict:
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT status FROM ordenes_compra WHERE id = %s", (oc_id,))
            row = cur.fetchone()
            if not row:
                raise HTTPException(404, "OC no encontrada")
            if row[0] != "BORRADOR":
                raise HTTPException(400, "Solo se pueden agregar ítems a una OC en estado BORRADOR")

            cur.execute("SELECT id FROM materials WHERE id = %s", (payload.material_id,))
            if not cur.fetchone():
                raise HTTPException(404, "Material no encontrado")

            cur.execute("""
                INSERT INTO ordenes_compra_items (oc_id, material_id, cantidad_pedida, precio_unitario, notas)
                VALUES (%s, %s, %s, %s, %s)
                RETURNING id
            """, (oc_id, payload.material_id, payload.cantidad_pedida, payload.precio_unitario, payload.notas))

        _recalc_total(conn, oc_id)
        conn.commit()

    return get_oc_service(oc_id)


def update_oc_item_service(oc_id: str, item_id: str, payload) -> dict:
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT status FROM ordenes_compra WHERE id = %s", (oc_id,))
            row = cur.fetchone()
            if not row:
                raise HTTPException(404, "OC no encontrada")
            if row[0] != "BORRADOR":
                raise HTTPException(400, "Solo se pueden editar ítems de una OC en BORRADOR")

        fields, vals = [], []
        if payload.cantidad_pedida is not None:
            fields.append("cantidad_pedida = %s"); vals.append(payload.cantidad_pedida)
        if payload.precio_unitario is not None:
            fields.append("precio_unitario = %s"); vals.append(payload.precio_unitario)
        if payload.notas is not None:
            fields.append("notas = %s"); vals.append(payload.notas)
        if not fields:
            raise HTTPException(400, "Nada que actualizar")
        vals += [item_id, oc_id]

        with conn.cursor() as cur:
            cur.execute(
                f"UPDATE ordenes_compra_items SET {', '.join(fields)} "
                "WHERE id = %s AND oc_id = %s RETURNING id",
                vals,
            )
            if not cur.fetchone():
                raise HTTPException(404, "Ítem no encontrado")

        _recalc_total(conn, oc_id)
        conn.commit()

    return get_oc_service(oc_id)


def delete_oc_item_service(oc_id: str, item_id: str) -> dict:
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT status FROM ordenes_compra WHERE id = %s", (oc_id,))
            row = cur.fetchone()
            if not row:
                raise HTTPException(404, "OC no encontrada")
            if row[0] != "BORRADOR":
                raise HTTPException(400, "Solo se pueden eliminar ítems de una OC en BORRADOR")

            cur.execute(
                "DELETE FROM ordenes_compra_items WHERE id = %s AND oc_id = %s RETURNING id",
                (item_id, oc_id)
            )
            if not cur.fetchone():
                raise HTTPException(404, "Ítem no encontrado")

        _recalc_total(conn, oc_id)
        conn.commit()

    return {"ok": True}


# ── Recepción de OC ───────────────────────────────────────────────────────────

def recibir_oc_service(oc_id: str, payload, user) -> dict:
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT status, almacen_destino FROM ordenes_compra WHERE id = %s",
                (oc_id,)
            )
            row = cur.fetchone()
            if not row:
                raise HTTPException(404, "OC no encontrada")
            if row[0] not in ("APROBADA", "EN_TRANSITO"):
                raise HTTPException(400, "Solo se puede recibir una OC en estado APROBADA o EN_TRANSITO")

            oc_almacen = str(row[1]) if row[1] else None

        total_real = 0.0

        for item_recv in payload.items:
            if item_recv.cantidad_recibida <= 0:
                continue

            almacen_id = item_recv.almacen_id or (payload.almacen_id if hasattr(payload, "almacen_id") else None) or oc_almacen
            if not almacen_id:
                raise HTTPException(400, "Se requiere almacen_id para la recepción")

            with conn.cursor() as cur:
                cur.execute("""
                    SELECT oci.id, oci.material_id, oci.cantidad_pedida, oci.cantidad_recibida, oci.precio_unitario
                    FROM ordenes_compra_items oci
                    WHERE oci.id = %s AND oci.oc_id = %s
                """, (item_recv.item_id, oc_id))
                item_row = cur.fetchone()
                if not item_row:
                    raise HTTPException(404, f"Ítem {item_recv.item_id} no pertenece a esta OC")

                _, material_id, cant_pedida, cant_ya_recibida, precio_u = item_row
                nueva_cantidad = float(cant_ya_recibida) + item_recv.cantidad_recibida

                # Crear stock_movement ENTRADA
                cur.execute("""
                    INSERT INTO stock_movements
                        (material_id, movement_type, quantity,
                         from_warehouse, to_warehouse, reference, created_by)
                    VALUES (%s::uuid, 'IN', %s, NULL, %s::uuid, %s, %s)
                    RETURNING id
                """, (
                    material_id, item_recv.cantidad_recibida,
                    almacen_id,
                    f"OC-{oc_id[:8]}",
                    user.get("username", str(user["id"])),
                ))
                movement_id = str(cur.fetchone()[0])

                # Actualizar stock_locations (sin ubicación física específica)
                cur.execute("""
                    SELECT id FROM stock_locations
                    WHERE material_id = %s::uuid AND warehouse_id = %s::uuid
                      AND rack = '' AND level = '' AND box = '' AND position = ''
                    FOR UPDATE
                """, (material_id, almacen_id))
                loc = cur.fetchone()
                if loc:
                    cur.execute(
                        "UPDATE stock_locations SET quantity = quantity + %s, updated_at = NOW() WHERE id = %s",
                        (item_recv.cantidad_recibida, loc[0])
                    )
                else:
                    cur.execute("""
                        INSERT INTO stock_locations
                            (material_id, warehouse_id, rack, level, box, position, quantity)
                        VALUES (%s::uuid, %s::uuid, '', '', '', '', %s)
                    """, (material_id, almacen_id, item_recv.cantidad_recibida))

                # Marcar ítem recibido
                cur.execute("""
                    UPDATE ordenes_compra_items
                    SET cantidad_recibida = %s, stock_movement_id = %s
                    WHERE id = %s
                """, (nueva_cantidad, movement_id, item_recv.item_id))

                total_real += item_recv.cantidad_recibida * float(precio_u)

        # Determinar nuevo estado
        with conn.cursor() as cur:
            cur.execute("""
                SELECT SUM(cantidad_pedida), SUM(cantidad_recibida)
                FROM ordenes_compra_items WHERE oc_id = %s
            """, (oc_id,))
            totals = cur.fetchone()
            pedida_total = float(totals[0] or 0)
            recibida_total = float(totals[1] or 0)

        new_status = "RECIBIDA" if recibida_total >= pedida_total else "EN_TRANSITO"

        with conn.cursor() as cur:
            cur.execute("""
                UPDATE ordenes_compra
                SET status = %s,
                    total_real = COALESCE(total_real, 0) + %s,
                    fecha_recepcion = CASE WHEN %s = 'RECIBIDA' THEN NOW() ELSE fecha_recepcion END,
                    updated_at = NOW()
                WHERE id = %s
            """, (new_status, total_real, new_status, oc_id))
            conn.commit()

    return get_oc_service(oc_id)


# ══════════════════════════════════════════════════════════════════════════════
# 📥 EXPORTAR OCs A EXCEL
# ══════════════════════════════════════════════════════════════════════════════

def export_oc_excel_service(status=None, proveedor_id=None):
    from app.core.export_utils import (
        write_title_row, write_header_row, write_data_row,
        set_column_widths, fmt_date, fmt_num, excel_response, write_total_row,
    )
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from datetime import datetime

    ocs = list_oc_service(status=status, proveedor_id=proveedor_id)
    fecha = datetime.now().strftime("%d/%m/%Y %H:%M")

    STATUS_COLORS = {
        "BORRADOR":    ("F3F4F6", "374151"),
        "ENVIADA":     ("DBEAFE", "1E40AF"),
        "APROBADA":    ("D1FAE5", "065F46"),
        "EN_TRANSITO": ("FEF3C7", "92400E"),
        "RECIBIDA":    ("E0E7FF", "3730A3"),
        "CERRADA":     ("F0FDF4", "14532D"),
        "CANCELADA":   ("FEE2E2", "991B1B"),
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Órdenes de Compra"

    headers = [
        "Código OC", "Proveedor", "Estado", "Almacén Destino",
        "Total Estimado (S/)", "Total Real (S/)",
        "Fecha Solicitud", "Fecha Entrega Est.", "Fecha Recepción",
        "Notas",
    ]
    widths = [13, 32, 14, 24, 18, 16, 16, 18, 16, 32]

    write_title_row(ws, f"CeShark ERP — Órdenes de Compra — {fecha}", len(headers))
    write_header_row(ws, headers, row=2)
    set_column_widths(ws, widths)
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 18

    total_est_sum = 0.0
    total_real_sum = 0.0

    for i, oc in enumerate(ocs, start=1):
        t_est = oc.get("total_estimado") or 0
        t_real = oc.get("total_real") or 0
        total_est_sum  += float(t_est)
        total_real_sum += float(t_real)

        write_data_row(ws, i + 2, [
            oc.get("code", ""),
            oc.get("proveedor_nombre", ""),
            oc.get("status", ""),
            oc.get("almacen_nombre") or "",
            fmt_num(t_est),
            fmt_num(t_real) if t_real else "",
            fmt_date(oc.get("fecha_solicitud")),
            fmt_date(oc.get("fecha_entrega_est")),
            fmt_date(oc.get("fecha_recepcion")),
            oc.get("notas") or "",
        ], alternate=(i % 2 == 0))

        status_val = oc.get("status", "")
        if status_val in STATUS_COLORS:
            bg, fg = STATUS_COLORS[status_val]
            c = ws.cell(row=i + 2, column=3)
            c.fill = PatternFill("solid", fgColor=bg)
            c.font = Font(size=9, bold=True, color=fg)

    tot_row = len(ocs) + 3
    write_total_row(ws, tot_row, [
        f"TOTAL ({len(ocs)} OCs)", "", "", "",
        fmt_num(total_est_sum), fmt_num(total_real_sum),
        "", "", "", "",
    ], len(headers))

    return excel_response(wb, f"ordenes_compra_{datetime.now().strftime('%Y%m%d')}.xlsx")


# ══════════════════════════════════════════════════════════════════════════════
# 📥 EXPORTAR PROVEEDORES A EXCEL
# ══════════════════════════════════════════════════════════════════════════════

def export_proveedores_excel_service():
    from app.core.export_utils import (
        write_title_row, write_header_row, write_data_row,
        set_column_widths, fmt_date, excel_response,
    )
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from datetime import datetime

    proveedores = list_proveedores_service()
    fecha = datetime.now().strftime("%d/%m/%Y %H:%M")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Proveedores"

    headers = [
        "Código", "Nombre / Razón Social", "RUC",
        "Teléfono", "Email", "Contacto",
        "Tipo", "Estado", "Dirección", "Registrado",
    ]
    widths = [11, 36, 13, 15, 28, 22, 16, 10, 32, 14]

    write_title_row(ws, f"CeShark ERP — Catálogo de Proveedores — {fecha}", len(headers))
    write_header_row(ws, headers, row=2)
    set_column_widths(ws, widths)
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 18

    for i, p in enumerate(proveedores, start=1):
        activo = p.get("activo", True)
        write_data_row(ws, i + 2, [
            p.get("codigo", ""),
            p.get("nombre", ""),
            p.get("ruc") or "",
            p.get("telefono") or "",
            p.get("email") or "",
            p.get("contacto") or "",
            p.get("tipo", ""),
            "Activo" if activo else "Inactivo",
            p.get("direccion") or "",
            fmt_date(p.get("created_at")),
        ], alternate=(i % 2 == 0))

        c = ws.cell(row=i + 2, column=8)
        if activo:
            c.fill = PatternFill("solid", fgColor="D1FAE5")
            c.font = Font(size=9, bold=True, color="065F46")
        else:
            c.fill = PatternFill("solid", fgColor="FEE2E2")
            c.font = Font(size=9, bold=True, color="991B1B")

        tipo_c = ws.cell(row=i + 2, column=7)
        if p.get("tipo") == "SUBCONTRATISTA":
            tipo_c.fill = PatternFill("solid", fgColor="FEF3C7")
            tipo_c.font = Font(size=9, color="92400E")
        else:
            tipo_c.fill = PatternFill("solid", fgColor="DCFCE7")
            tipo_c.font = Font(size=9, color="166534")

    return excel_response(wb, f"proveedores_{datetime.now().strftime('%Y%m%d')}.xlsx")
